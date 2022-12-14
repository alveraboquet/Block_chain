#用于检查那么多Gmail邮箱，哪些是好的，哪些是不好的
#思路是去订阅这个网站https://www.storkapp.me/

# pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pandas
from sqlite3 import Time
import pandas as pd
import random
import time, os, re
import requests
import json
import re
import sys
import ast # 用于将字典型字符串转为字典
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver import ActionChains  # 用于操作鼠标
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ChromeOptions  #
import logging
import pyperclip  # 用于读取剪切板
import string
from itertools import chain
from random import choice, sample
from openpyxl import Workbook,load_workbook
import linecache #用于读取txt文档
from faker import Faker
from random_word import RandomWords
fake = Faker()
# from bs4 import BeautifulSoup
from lxml import etree

##===========================脆球邮箱相关
#从脆球官网获取
cuiqiu_token = '88434c9de6ef45b0b8f360a190f60abd'
cuiqiu_mail_id = '608142'

#找alchemy的注册激活的邮件 id
def cuiqiu_find_alchemy_activate_email_id(email_to_be_activate, email_from, email_subject):
    #循环检索邮箱. email_to_be_activate 表示待激活的邮箱
    not_find_yet = True
    try_times = 0
    while not_find_yet:
        #获取邮件列表
        url = "https://domain-open-api.cuiqiu.com/v1/box/list"
        payload={'mail_id': cuiqiu_mail_id,
        'token': cuiqiu_token,
        'start_time': '2022-08-24',
        'end_time': '2023-08-25',
        'page': '1',
        'limit': '100'}
        files=[

        ]
        headers = {}
        response = requests.request("POST", url, headers=headers, data=payload, files=files)
        result = response.text #原始字符串格式
        result_to_json = json.loads(result) #字符串转json
        
        #========这是一个列表集合, 先把列表轮寻一遍, 提取邮件id
        #如果失败, 则再请求下一个列表
        email_list = result_to_json['data']['list'] #取值
        # print("=========所有的邮件在这里: ", email_list)
        for email in email_list:
            if email["to"] == email_to_be_activate:
                if email["from"] == email_from:
                    if email["subject"]== email_subject:
                        result_email_id = email["id"]
                        print("===========待激活的邮件id是:", result_email_id)
                        not_find_yet = False #防止死循环
                        return result_email_id

        try_times += 1
        time_sleep(5,f"尝试{try_times}次, 最多100次. 是不是参数 limit 太少了? ")
        if try_times == 100:
            print("找邮件重试了10分钟,还是失败")
            not_find_yet = False #防止死循环
            result_email_id = False
            return result_email_id


#找到邮件 id 后, 开始提取链接
def cuiqiu_extract_alchemy_link_from_email_id(email_id):
    # # box_id 通过 v1/box/list 获取邮箱列表接口获取
    url = "https://domain-open-api.cuiqiu.com/v1/box/detail"
    payload={'mail_id': cuiqiu_mail_id,
    'token': cuiqiu_token,
    'box_id': email_id}
    files=[

    ]
    headers = {}

    response = requests.request("POST", url, headers=headers, data=payload, files=files)
    result = response.text
    # print("====request邮件内容初始格式", result)

    to_json = json.loads(result)
    print("转化为json后, 如果有html 则可以xml解析, 很关键!!")
    # print("======转为json后", to_json) #这个决定了能不能用xml解析, 如果有html 则可以解析

    html_body = to_json['data']['content']['body']
    # print("===提取html_body, 输入到etree", html_body)

    html = etree.HTML(html_body)
    # print("=====输入到etree, 解析到的邮件html详情是:", type(html), html)
    
    activate_link = html.xpath("//a[text()='VERIFY EMAIL']/@href")
    print("=====提取到的链接是, 这是一个列表, 需要转为字符串", activate_link, type(activate_link))
    return str(activate_link[0])

#开始打开浏览器激活
def cuiqiu_browser_active_alchemy_link(browser, wait, activate_link):
    print("开始激活alchemy发来的邮件")
    dashboard_flag = False
    
    browser.get(activate_link)
    time_sleep(5, "等待 alchemy 加载")
    switch_tab_by_handle(browser, 1, 0)  # 切换到

    #=================激活帐号
    try:
        Verify_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Verify']")))
        time_sleep(2,"准备点击 Verify")
        browser.execute_script("arguments[0].click();", Verify_button)
        time_sleep(30,"=====已经点击 Verify")
    except:
        print("点击verify失败, 会不会有影响?")
 
    # ==================填写描述
    try:
        fill_in_alchemy_project_des(browser, wait)
    except:
        print("可能是不需要填写alchemy项目描述, 或哪里出错了")

    try:
        #如果能找到Alchemy的首页,说明已经进入了
        time_sleep(20, "尝试查看是不是已经进入 Alchemy 了")
        Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()[contains(.,'Dashboard')]]")))
        dashboard_flag = True
        print("=======已经登录了 Alchemy")
        return dashboard_flag
    except:
        browser.quit()
        print("======没有找到dashboard=====")
    


##===========切换IP相关
url_dashboard = "http://127.0.0.1:9090/ui/#/proxies"
url_google = "https://www.google.com"

##==========小狐狸相关
metamask_pw = "12345678" #小狐狸的密码
metamask_home = "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html"

##========== zk
ZK_balance_url = "https://wallet.zksync.io/account" #查询zk余额用

# 要导入的网络的一些信息
Op_network_name = 'Optimism'
Op_network_RPC_URL = 'https://mainnet.optimism.io'
Op_network_chain_id = 10
Op_network_token = 'ETH'
Op_network_browers = 'https://optimistic.etherscan.io'
Op_list=[Op_network_name, Op_network_RPC_URL, Op_network_chain_id, Op_network_token, Op_network_browers]

Arb_network_name = 'Arbitrum One'
Arb_network_RPC_URL = 'https://arb1.arbitrum.io/rpc'
#第三方RPC：https://arbitrum.getblock.io/mainnet/?api_key=15c8bb53-1257-47ba-a662-1133a9a2af1b
#https://rpc.ankr.com/arbitrum

Arb_network_chain_id = 42161
Arb_network_token = 'ETH'
Arb_network_browers = 'https://arbiscan.io'
Arb_list=[Arb_network_name, Arb_network_RPC_URL, Arb_network_chain_id, Arb_network_token, Arb_network_browers]

##### L1、L2转账用 url
orbiter_url =  "https://www.orbiter.finance/"
lifi_url = "https://transferto.xyz/swap"
##########

# =========== 官方桥
lifi_url = "https://transferto.xyz/swap"
ARB_url = "https://bridge.arbitrum.io/"
OP_url = "https://app.optimism.io/bridge"
ZK_url = "https://wallet.zksync.io/transaction/deposit"
bungee_url = "https://www.bungee.exchange/"
hop_url = "https://app.hop.exchange/#/send?token=ETH"
#######

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ 通用函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#txt索引从第一行开始
class DO_TXT:
    def __init__(self, input_path, input_x_line):
        self.path = input_path
        self.line = input_x_line
    def read_x_line(self):
        result = linecache.getline(self.path, self.line)
        return result
    def delet_x_line(self):
        file = open(self.path)
        lines = file.readlines()
        del lines[self.line-1:self.line]
        file.close()
        file_new = open(self.path, "w")
        file_new.writelines(lines)
        file_new.close()


# 保存剪切板数据
def save_clipper_data(filename):
    time.sleep(4)  # 给点延时，剪切板正在写入
    code = pyperclip.paste()
    print(f"我是复制的内容：{code}")
    with open(f'./{filename}.txt', 'a', encoding='utf-8') as file:
        file.write(f'{code}\n')
        file.close()

#保存数据到txt
def save_data_to_txt(data, filename):
    time.sleep(4)  # 给点延时，剪切板正在写入
    print(f"要保存的内容是：{data}")
    with open(f'./{filename}.txt', 'a', encoding='utf-8') as file:
        file.write(f'{data}\n')
        file.close()


#用来产生密码，特殊符号的长度是：总长度扣除正常的字符长度
def mkpasswd(length=12, digits=4, upper=3, lower=2):
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    salt = '!@#$%^&*()><?'
    password = list(
    chain(
    (choice(uppercase) for _ in range(upper)),
    (choice(lowercase) for _ in range(lower)),
    (choice(string.digits) for _ in range(digits)),
    (choice(salt) for _ in range((length - digits - upper - lower)))
    )
    )
    return "".join(sample(password, len(password)))


#可视化等待。第一个参数是要等待的时间；第二个参数是注释
def time_sleep(data, info="等待"):
    for x in range(data, -1, -1):
        mystr = f'{info}：' + str(x) + "s"
        print("\r", mystr, end="", flush=True)
        time.sleep(1)
    print("\r")

#构建浏览器
def my_chrome(chrome_user_num):
    # 获取驱动的路径
    driver_path = os.path.abspath('..') + "\chromedriver.exe"  # driver版本要和Chrome对应
    # print(driver_path)

    TIME_OUT = 40  # 设置显示等待的超时时间，尽量设置的长一点，考虑到网络可能缓慢
    option = ChromeOptions()
    # option.add_argument('--headless')#是否开启无头模式
    # option.add_argument('--disable-gpu')#屏蔽浏览器引擎

    option.add_argument("--disable-blink-features=AutomationControlled")  # 禁用启用Blink运行时的功能
    option.add_experimental_option('excludeSwitches', ['enable-automation'])  # 防止被网站识别
    option.add_experimental_option('useAutomationExtension', False)

    # 添加谷歌用户路径
    user_data_path = f'--user-data-dir=D:\\auto_video_projects\\xigua_video\\adidas\\{chrome_user_num}'
    # user_data_path = f'--user-data-dir=F:\\Chrome_5'
    # user_data_path = f'--user-data-dir=C:\\software\\Chrome_1_to_200'  #虚拟机里的chrome

    print("==========谷歌用户路径是：", user_data_path)
    option.add_argument(user_data_path)  # 用指定的浏览器进行测试

    ##=========== 准备浏览器
    browser = webdriver.Chrome(driver_path, options=option)
    # browser.minimize_window()
    browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                            {"source": "Object.defineProperty(navigator, 'webdriver', { get: () => undefined })"}
                            )  # 反爬
    wait = WebDriverWait(browser, TIME_OUT)  # wait是一个类
    browser.maximize_window()
    # browser.minimize_window()
    return wait, browser

#构建浏览器
def my_mac_chrome(time_out=30):
    # 获取驱动的路径
    driver_path = "/usr/local/bin/chromedriver" # chromedriver完整路径，path是重点。如果不行，试试chromedriver.exe
    # print(driver_path)

    TIME_OUT = time_out  # 设置显示等待的超时时间，尽量设置的长一点，考虑到网络可能缓慢
    option = ChromeOptions()
    # option.add_argument('--headless')#是否开启无头模式
    # option.add_argument('--disable-gpu')#屏蔽浏览器引擎
    option.add_argument("--disable-blink-features=AutomationControlled")  # 禁用启用Blink运行时的功能
    option.add_experimental_option('excludeSwitches', ['enable-automation'])  # 防止被网站识别
    option.add_experimental_option('useAutomationExtension', False)

    # Chrome path浏览器路径
    option.binary_location = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'

    # 添加谷歌用户路径
    user_data_path = '--user-data-dir=/Users/spencer/Library/Application Support/Google/Chrome'
    print("==========谷歌用户路径是：", user_data_path)
    option.add_argument(user_data_path)  # 用指定的浏览器进行测试

    ##=========== 准备浏览器
    browser = webdriver.Chrome(driver_path, options=option)
    # browser.minimize_window()
    browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                            {"source": "Object.defineProperty(navigator, 'webdriver', { get: () => undefined })"}
                            )  # 反爬
    wait = WebDriverWait(browser, TIME_OUT)  # wait是一个类
    # browser.maximize_window()
    # browser.minimize_window()
    return wait, browser

#构建浏览器
def my_linux_chrome(time_out=30):
    # 获取驱动的路径
    # driver_path = "/usr/local/bin/chromedriver" # chromedriver完整路径，path是重点。如果不行，试试chromedriver.exe
    # print(driver_path)

    TIME_OUT =  time_out # 设置显示等待的超时时间，尽量设置的长一点，考虑到网络可能缓慢
    option = ChromeOptions()
    # option.add_argument('--headless')#是否开启无头模式
    # option.add_argument('--disable-gpu')#屏蔽浏览器引擎
    option.add_argument('--no-sandbox')
    option.add_argument('--disable-dev-shm-usage')
    # option.add_argument('--disable-gpu')
    option.add_argument("--disable-blink-features=AutomationControlled")  # 禁用启用Blink运行时的功能
    option.add_experimental_option('excludeSwitches', ['enable-automation'])  # 防止被网站识别
    option.add_experimental_option('useAutomationExtension', False)

    # Chrome path浏览器路径
    option.binary_location = '/usr/lib/chromium-browser/chromium-browser'

    # 添加谷歌用户路径
    user_data_path = '--user-data-dir=/home/parallels/.config/chromium/Default'
    print("==========谷歌用户路径是：", user_data_path)
    option.add_argument(user_data_path)  # 用指定的浏览器进行测试

    ##=========== 准备浏览器
    browser = webdriver.Chrome(options=option)
    # browser.minimize_window()
    browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                            {"source": "Object.defineProperty(navigator, 'webdriver', { get: () => undefined })"}
                            )  # 反爬
    wait = WebDriverWait(browser, TIME_OUT)  # wait是一个类
    # browser.maximize_window()
    # browser.minimize_window()
    return wait, browser



#还是容易出现excel文件损坏的现象，加一个close
#特色功能：用英文字母就可以找到列；单元格内容追加而不是覆盖
class Do_Excel:
    def __init__(self,filename,sheetname='SheetJS'):  #这里要改为实际表单名字
        self.filename=filename
        self.sheetname=sheetname
    def write(self,i,j,my_value):
        real_column = ord(j)-64
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]
        #记录之前单元格的内容
        cell_befoe_content = sh.cell(i,real_column).value
        if cell_befoe_content == None: #如果是None，即无内容，则要写的字符串就是输入的字符串
            # final_value =  my_value
            final_value = '·' + my_value
        #最终要写到单元格里的内容
        else:
            final_value = cell_befoe_content + "\n" + '·' + my_value
            # final_value = cell_befoe_content + "\n"  + my_value
        print("excel正在写入数据，请勿操作，否则将损坏文件！")
        sh.cell(i,real_column).value = final_value
        wb.save(self.filename)
        wb.close()  #防止出错
        time_sleep(2, "excel写入数据成功")

    def plain_write(self,i,j,my_value):
        real_column = ord(j)-64
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]

        print("excel正在写入数据，请勿操作，否则将损坏文件！")
        sh.cell(i,real_column).value = my_value
        wb.save(self.filename)
        wb.close()  #防止出错
        time_sleep(2, "excel写入数据成功")

    def read(self, i, j):
        real_column = ord(j) - 64
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]
        print("excel正在读取数据，请勿操作，否则将损坏文件！")
        a = sh.cell(i, real_column).value
        wb.close()  # 防止出错
        time_sleep(1, "excel读取数据成功")
        return a


#自己写一个，加点延时
def Write_Excel(i,j,your_value):
    #加载文件
    wb = load_workbook("../eth1000_OP_操作后.xlsx")
    ws = wb.active
    ws.cell(i,j, value =  your_value)
    time.sleep(1)
    wb.save("eth1000_OP_操作后.xlsx")
    wb.close()

#将字典型字符串转为字典，返回的是字典
def dict_type_string_to_dict(str_info):
    dict_info = ast.literal_eval(str_info)
    return dict_info


#用于找到ETH最大值的那个生态，输入字典，返回字典中，值最大的那个键
def find_max_ETH_chain(dict_info):
    a = zip(dict_info.values(), dict_info.keys())
    b = max(a)  # 返回元组
    # print(b)
    return b  # 第一个元素是值，第二个元素是键


#读取excel表格
def read_excel():
    path = "../"  # 也可采用 r
    for root,dirs,files in os.walk(path):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx' or os.path.splitext(file)[1] == '.xls':
            # 使用join函数将文件名称和文件所在根目录连接起来
                print("excel文件名是:", os.path.join(root, file))
                excel_path = os.path.join(root, file)
                frame = pd.read_excel(excel_path, header=None)  # 直接使用 read_excel() 方法读取
                print(f"总共有{frame.shape[0]}行")
                return frame


def delete_cookie(browser):
    print("进入delete_cookie()，清理缓存、cookie")
    # 清除缓存提示框
    clean_url = "https://www.baidu.com"
    new_tab(browser, clean_url)
    # 2S 等待时间
    time_sleep(2)
    switch_tab_by_handle(browser, 1, 0)  # 切换到该页面

    clean_url = "chrome://settings/clearBrowserData"
    browser.get(clean_url)
    time_sleep(2,"准备清缓存, 记得谷歌清缓存选择'all time'")
    clearButton = browser.execute_script(
        "return document.querySelector('settings-ui').shadowRoot.querySelector('settings-main').shadowRoot.querySelector('settings-basic-page').shadowRoot.querySelector('settings-section > settings-privacy-page').shadowRoot.querySelector('settings-clear-browsing-data-dialog').shadowRoot.querySelector('#clearBrowsingDataDialog').querySelector('#clearBrowsingDataConfirm')")
    clearButton.click()
    time_sleep(2, "清理完毕")
    # browser.close()

##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ 通用函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ 小狐狸的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
def fox_import_token(browser, wait, token_address, token_name, accurancy):
    try:
        # browser.refresh()
        # time_sleep(15, "刷新后等待")
        import_token_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[text()[contains(.,'导入代币')]]")))
        time_sleep(2,"准备点击导入代币")
        browser.execute_script("arguments[0].click();", import_token_button)
       
        try:
            custom_token_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='自定义代币']")))
            time_sleep(2,"准备点击自定义代币")
            browser.execute_script("arguments[0].click();", custom_token_button)
        except:
            print("可能是不需要点击自定义代币")

        address_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='custom-address']")))
        time_sleep(3,"准备输入地址")
        address_button.send_keys(token_address)

        token_name_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='custom-symbol']")))
        time_sleep(3,"准备输入代币符号, 全选")
        token_name_button.send_keys(Keys.CONTROL, "a")
        time_sleep(3,"准备输入代币符号, 发送")
        token_name_button.send_keys(token_name)

        accurancy_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='custom-decimals']")))
        time_sleep(3,"准备输入精度, 全选")
        accurancy_button.send_keys(Keys.CONTROL, "a")
        time_sleep(3,"准备输入精度, 发送")
        accurancy_button.send_keys(accurancy)

        confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'添加自定义代币')]]")))
        time_sleep(3,"准备点击----添加自定义代币")
        browser.execute_script("arguments[0].click();", confirm_button)

        confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'导入代币')]]")))
        time_sleep(3,"准备点击确认-----导入代币")
        browser.execute_script("arguments[0].click();", confirm_button)
        time_sleep(60, "等待添加完成")

        close_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[@class='currency-display-component__text']")))
        time_sleep(2,"准备查询余额")
        browser.execute_script("arguments[0].click();", close_button)
        
        fox_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='app-header__logo-container app-header__logo-container--clickable']")))
        time_sleep(2,"准备回到首页")
        browser.execute_script("arguments[0].click();", fox_button)
        time_sleep(10)
        #

        # balance_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[@class='currency-display-component__text']")))
        # time_sleep(2,"准备查询余额")
        # browser.execute_script("arguments[0].click();", balance_button)
        # print(f"余额是{balance_button.text}: 直接切换下一个帐号")
        # return balance_button.text
        #
    except:
        try:#可能输入地址后, 自己会自动跳出剩余的信息,导致出错,这时直接确认添加
            print("++++++++出错了, 可能输入地址后, 自己会自动跳出剩余的信息,导致出错,这时直接确认添加+++++++++++")
            confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'添加自定义代币')]]")))
            time_sleep(3,"准备点击----添加自定义代币")
            browser.execute_script("arguments[0].click();", confirm_button)

            confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'导入代币')]]")))
            time_sleep(3,"准备点击确认-----导入代币")
            browser.execute_script("arguments[0].click();", confirm_button)
            time_sleep(60, "等待添加完成")
            #
            fox_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='app-header__logo-container app-header__logo-container--clickable']")))
            time_sleep(2,"刚才直接点击了添加, 准备回到首页")
            browser.execute_script("arguments[0].click();", fox_button)
            time_sleep(10)
        except:  #再出错就只能回到首页了
            #
            fox_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='app-header__logo-container app-header__logo-container--clickable']")))
            time_sleep(2,"出错了, 可能是已经添加了, 准备回到首页")
            browser.execute_script("arguments[0].click();", fox_button)
            time_sleep(10)


#登陆小狐狸，直到登陆成功.
def login_metamask(browser, wait, metamask_pw, metamask_home, netname=None):
    print("我已进入login_metamask，开始登陆小狐狸")
    # new_tab(browser, metamask_home)
    browser.get(metamask_home)
    time_sleep(3)
    send_password = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='password']")))
    send_password.send_keys(metamask_pw)
    time.sleep(2)
    send_password.send_keys(Keys.ENTER)
    time_sleep(35,"正在打开小狐狸")
    for i in range(1,10):
        try:
            just_wait = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
            break
        except:
            # browser.refresh()
            time_sleep(5, "已经输入小狐狸登陆，还未进入主页，继续等待")
    if netname: #如果有参数，说明要切换网络
        for i in range(1,6): #尝试切换5次
            time_sleep(5, f"小狐狸准备切换网络，第{i}次")
            fox_change_network(browser, wait, netname)
            try:  #如果出现"切换网络"失败，则关闭提示
                time_sleep(30, "准备查看是否网络失败")
                # change_net_error 即出现"切换网络"失败的提醒
                change_net_error = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div[1]/div[2]/div/div/button[1]')))
                time_sleep(2, "确实找到了失败按钮，需要再次切换，先关闭提醒")

                # =====下面是关闭错误提示的代码。需要用鼠标模拟点击，因为有after
                # close_change_network = browser.find_element(By.CSS_SELECTOR, '#app-content > div > div.main-container-wrapper > div.loading-overlay > div.page-container__header-close')
                close_change_network = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#app-content > div > div.main-container-wrapper > div.loading-overlay > div.page-container__header-close')))
                time_sleep(2)
                ActionChains(browser).click(close_change_network).perform()  # 模拟鼠标点
                time_sleep(2,"已经关闭了错误提醒")
                if i == 5:
                    print(f"已经尝试切换了{i}次，还是失败，关闭浏览器")
                    browser.close()
                else:
                    print("已经关闭切换网络失败的提醒，准备再次切换")
                    continue

            except:
                print("√ √ √ 小狐狸可能没有遇到切换网络失败的问题")
                break

#先不考虑
def fox_commom_confirm(browser, wait):
    print("我已经进入fox_commom_confirm，这是小狐狸【万能】确认函数")
    browser.refresh()
    time_sleep(5)
    browser.refresh()
    time_sleep(6)
    try: #判断是不是全选账户的页面
        judge = comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='permissions-connect-header__subtitle']")))
        if judge.text == "选择账户":
            print("需要全选账号")
            count_acounts = wait.until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@class='choose-account-list__account']")))
            print("Accounts账户个数：", len(list(count_acounts)))
            ACC_NUM = len(list(count_acounts))
            select_all_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                       "//input[@class='check-box choose-account-list__header-check-box fa fa-minus-square check-box__indeterminate']")))
            browser.execute_script("arguments[0].click();", select_all_button)
            print('全选Account结束，点击下一步')

    except:
        print("不是全选账户的页面")

    # try:

#小狐狸确认交易
def fox_confirm_swap(browser, wait):
    print("我已经进入fox_confirm_swap，小狐狸确认交易")
    browser.refresh()#预防之前没刷新上
    time_sleep(10)
    # #为了保证能够找到 gas fee
    # gas_fee = wait.until(EC.element_to_be_clickable(
    #     (By.XPATH, "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
    # time_sleep(3)
    # loc = gas_fee.location_once_scrolled_into_view
    # time_sleep(3)
    # print("已经滚动鼠标滚轮")
    gas_fee = wait.until(EC.element_to_be_clickable(
        (By.XPATH,
         "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
    print("gas fee预估是：",gas_fee.text)
    gas = gas_fee.text
    if float(gas) <=0.005:#gas fee 比较低  千分之五内能接受
        try:
            # # 为了保证能够找到 gas fee
            # gas_fee = wait.until(EC.element_to_be_clickable(
            #     (By.XPATH,
            #      "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
            # time_sleep(3)
            # loc = gas_fee.location_once_scrolled_into_view
            # time_sleep(3)
            # print("gas fee 满足要求，已经滚动鼠标滚轮")

            ###找到确认按钮
            comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
            browser.execute_script("arguments[0].click();", comfirm_button)
            time_sleep(3)
            #判断是否点击成功，因为有时按钮灰色，点不到
            while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
                ###找到确认按钮
                print("循环判断小狐狸是否真的点击【确定交易】")
                comfirm_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
                browser.execute_script("arguments[0].click();", comfirm_button)
                time_sleep(1)
            print("已经点击【确认】")
            print(f"小狐狸交易成功，gas fee预估{gas}")
            return f"小狐狸交易成功，gas fee预估{gas}"
        except:
            # print(f"fox_confirm_swap，gas fee满足条件，但小狐狸点击【确认】失败，刷新{i}次试一试")
            # browser.refresh()
            # time_sleep(10)# 刷新一次，可能 gas fee 会比较低
            # if i == 3: #试了3次都不行，则拒绝交易
            #     # 为了保证能够找到 gas fee
            #     # gas_fee = wait.until(EC.element_to_be_clickable(
            #     #     (By.XPATH,
            #     #      "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
            #     # time_sleep(3)
            #     # loc = gas_fee.location_once_scrolled_into_view
            #     # time_sleep(3)
            #     # print("已经滚动鼠标滚轮")
            #     #拒绝交易
            cancel_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
            print("小狐狸交易失败，点击【拒绝交易】")
            return f"小狐狸交易失败，点击【拒绝交易】"

#小狐狸查询 gas fee确认交易
def fox_get_gas_fee_and_confirm_swap(browser, wait):
    print("我已经进入fox_get_gas_fee_and_confirm_swap，小狐狸确认交易")
    browser.refresh()#预防之前没刷新上
    time_sleep(10)
    #=====为了保证能够找到 gas fee
    gas_fee = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
    '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div/div/div/div/div[1]/div[1]/div/h6[2]'
    time_sleep(1)
    loc = gas_fee.location_once_scrolled_into_view
    time_sleep(1, "已经滚动鼠标滚轮")
    gas_fee = wait.until(EC.element_to_be_clickable(
        (By.XPATH,
         "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
    print("gas fee预估是：",gas_fee.text)
    gas = gas_fee.text

    if float(gas) <=0.003:#gas fee 比较低  千分之五内能接受
        try:

            ###找到确认按钮
            comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
            browser.execute_script("arguments[0].click();", comfirm_button)
            time_sleep(3)
            #判断是否点击成功，因为有时按钮灰色，点不到
            while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
                ###找到确认按钮
                print("循环判断小狐狸是否真的点击【确定交易】")
                comfirm_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
                browser.execute_script("arguments[0].click();", comfirm_button)
                time_sleep(1)
            detail = f"小狐狸已成功点击【确认】交易，gas fee预估{gas}"
            print(detail)
            return detail
        except:
            print("fox_get_gas_fee_and_confirm_swap(), 出错了，失败")

    # ========拒绝交易
    else:
        # gas_fee = wait.until(EC.element_to_be_clickable(
        #     (By.XPATH,
        #      "//div[@class='transaction-detail-item__detail-values']//div[@class='confirm-page-container-content__currency-container']/div[@class='currency-display-component']/span[@class='currency-display-component__text']")))
        # time_sleep(3)
        # loc = gas_fee.location_once_scrolled_into_view
        # time_sleep(3, "已经滚动鼠标滚轮")

        # 拒绝交易
        cancel_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", cancel_button)
        detail = f"小狐狸交易失败，点击【拒绝交易】，gas fee预估{gas}"
        print(detail)
        return detail


#小狐狸确认交易，因为L2与L2交易时，gas fee 的位置不一样
def fox_confirm_L2_swap(browser, wait):
    print("我已经进入fox_confirm_L2_swap，小狐狸确认交易")
    browser.refresh()#预防之前没刷新上
    time_sleep(10)
    c = 0#默认c为0，print防止出错
    #=======如果 gas fee 为 0 ，则拒绝交易
    try:
        gas_fee_button  = browser.find_element(By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div/div/div[1]/div[1]/div/h6[2]')
        print("小狐狸gas fee 为: ", gas_fee_button.text)
        # gas_fee_button = wait.until(EC.element_to_be_clickable(
        #     (By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div[2]/div[1]/div[2]/h6/div/div[2]/span[2]')))
        a = gas_fee_button.text
        b = a.split()
        c = b[0]
        print("提取后的gas fee 为: ", c)
        if c == "0" or float(c) >=0.005:
            cancel_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
            info = "考虑重新交易，小狐狸交易失败，点击【拒绝交易】，因为 gas fee是 0 或者大于 0.003 "
            print(info)
            return info
    except:
        print("没有找到gas fee")
    try:
        # 为了保证能够找到 确定按钮
        print(f"fox_confirm_L2_swap()小狐狸尝试去【确认】")
        # comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        # time_sleep(2)
        # loc = comfirm_button.location_once_scrolled_into_view
        # time_sleep(2)
        # print("已经滚动鼠标滚轮")

        ###找到确认按钮
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        browser.execute_script("arguments[0].click();", comfirm_button)
        time_sleep(2, "小狐狸已经点击确认")
        #判断是否点击成功，因为有时按钮灰色，点不到
        # while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
        #     ###找到确认按钮
        #     print("循环判断小狐狸是否真的点击【确定交易】")
        #     comfirm_button = wait.until(EC.element_to_be_clickable(
        #         (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        #     browser.execute_script("arguments[0].click();", comfirm_button)
        #     time_sleep(1)
        print("小狐狸交易成功，gas fee：", c)
        return " 小狐狸交易成功"
    except:
        #拒绝交易
        cancel_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", cancel_button)
        print("小狐狸交易失败，点击【拒绝交易】")
        return f"小狐狸交易失败，点击【拒绝交易】"



def fox_confirm_orb_swap(browser, wait):
    print("我已经进入fox_confirm_orb_swap，小狐狸确认交易")
    browser.refresh()#预防之前没刷新上
    time_sleep(10)
    #=======如果 gas fee 为 0 ，则拒绝交易
    try:
        gas_fee_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div[2]/div[1]/div[1]/div/h6[1]/div/div[2]/span[2]')))
        print("小狐狸 gas fee 为: ", gas_fee_button.text)
        # gas_fee_button = wait.until(EC.element_to_be_clickable(
        #     (By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div[2]/div[1]/div[2]/h6/div/div[2]/span[2]')))
        a = gas_fee_button.text
        # b = a.split()
        # c = b[0]
        print("提取后的gas fee 为: ", a)
        if a == "0" or float(a) >=0.005:
            cancel_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
            info = "考虑重新交易，小狐狸交易失败，点击【拒绝交易】，因为 gas fee是 0 或者大于 0.003 "
            print(info)
            return info
    except:
        print("没有找到gas fee")
    try:
        # 为了保证能够找到 确定按钮
        print("fox_confirm_orb_swap()小狐狸尝试去【确认】")
        # comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        # time_sleep(2)
        # loc = comfirm_button.location_once_scrolled_into_view
        # time_sleep(2)
        # print("已经滚动鼠标滚轮")

        ###找到确认按钮
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        browser.execute_script("arguments[0].click();", comfirm_button)
        time_sleep(2, "小狐狸已经点击确认")
        #判断是否点击成功，因为有时按钮灰色，点不到
        # while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
        #     ###找到确认按钮
        #     print("循环判断小狐狸是否真的点击【确定交易】")
        #     comfirm_button = wait.until(EC.element_to_be_clickable(
        #         (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        #     browser.execute_script("arguments[0].click();", comfirm_button)
        #     time_sleep(1)
        print("小狐狸交易成功，gas fee：", a)
        return " 小狐狸交易成功"
    except:
        #拒绝交易
        cancel_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", cancel_button)
        print("小狐狸交易失败，点击【拒绝交易】")
        return f"小狐狸交易失败，点击【拒绝交易】"


def fox_confirm_bungee_swap(browser, wait):
    print("我已经进入fox_confirm_bungee_swap，小狐狸确认交易")
    # browser.refresh()#预防之前没刷新上
    time_sleep(5)
    #=======如果 gas fee 为 0 ，则拒绝交易
    try:
        gas_fee_button  = browser.find_element(By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div[1]/div/h6[2]/div/div[2]/span[2]')
        print("小狐狸gas fee 为: ", gas_fee_button.text)
        # gas_fee_button = wait.until(EC.element_to_be_clickable(
        #     (By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div[2]/div[1]/div[2]/h6/div/div[2]/span[2]')))
        a = gas_fee_button.text
        # b = a.split()
        # c = b[0]
        print("提取后的gas fee 为: ", a)
        if a == "0" or float(a) >= 0.005:
        # if c == "0" or float(c) >=0.005:
            cancel_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
            info = "考虑重新交易，小狐狸交易失败，点击【拒绝交易】，因为 gas fee是 0 或者大于 0.003 "
            print(info)
            return info
    except:
        print("没有找到gas fee")
    try:
        # 为了保证能够找到 确定按钮
        print(f"fox_confirm_bungee_swap()小狐狸尝试去【确认】")
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        time_sleep(2)
        loc = comfirm_button.location_once_scrolled_into_view
        time_sleep(2)
        print("已经滚动鼠标滚轮")

        ###找到确认按钮
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        browser.execute_script("arguments[0].click();", comfirm_button)
        time_sleep(2, "小狐狸已经点击确认")
        #判断是否点击成功，因为有时按钮灰色，点不到
        # while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
        #     ###找到确认按钮
        #     print("循环判断小狐狸是否真的点击【确定交易】")
        #     comfirm_button = wait.until(EC.element_to_be_clickable(
        #         (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        #     browser.execute_script("arguments[0].click();", comfirm_button)
        #     time_sleep(1)
        print("小狐狸交易成功，gas fee：", a)
        return " 小狐狸交易成功"
    except:
        #拒绝交易
        cancel_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", cancel_button)
        print("小狐狸交易失败，点击【拒绝交易】，可能是不足以支付 gas fee:", a)
        return f"小狐狸交易失败，点击【拒绝交易】"


def fox_confirm_hop_swap(browser, wait):
    print("我已经进入fox_confirm_hop_swap，小狐狸确认交易")
    # browser.refresh()#预防之前没刷新上
    # time_sleep(10)
    #=======如果 gas fee 为 0 ，则拒绝交易
    try:
        gas_fee_button  = browser.find_element(By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div[1]/div/h6[2]/div/div[2]/span[2]')
        print("小狐狸gas fee 为: ", gas_fee_button.text)
        # gas_fee_button = wait.until(EC.element_to_be_clickable(
        #     (By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[2]/div/div[2]/div[1]/div[2]/h6/div/div[2]/span[2]')))
        a = gas_fee_button.text
        b = a.split()
        c = b[0]
        print("提取后的gas fee 为: ", c)
        if c == "0" or float(c) >= 0.005:
        # if a == "0" or float(a) >=0.005:
            cancel_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
            info = "考虑重新交易，小狐狸交易失败，点击【拒绝交易】，因为 gas fee是 0 或者大于 0.003 "
            print(info)
            return info
    except:
        print("没有找到gas fee")
        return "失败，没有找到gas fee"
    try:
        # 为了保证能够找到 确定按钮
        print(f"gas fee 满足要求，fox_confirm_bungee_swap()小狐狸尝试去【确认】")
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        time_sleep(2)
        loc = comfirm_button.location_once_scrolled_into_view
        time_sleep(2)
        print("已经滚动鼠标滚轮")

        ###找到确认按钮
        comfirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        browser.execute_script("arguments[0].click();", comfirm_button)
        time_sleep(2, "小狐狸已经点击确认")
        # #判断是否点击成功，因为有时按钮灰色，点不到
        # while not wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='currency-display-component eth-overview__primary-balance']/span[@class='currency-display-component__text']"))):
        #     ###找到确认按钮
        #     print("循环判断小狐狸是否真的点击【确定交易】")
        #     comfirm_button = wait.until(EC.element_to_be_clickable(
        #         (By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        #     browser.execute_script("arguments[0].click();", comfirm_button)
        #     time_sleep(1)
        print("小狐狸交易成功, fas fee:", c)
        return " 小狐狸交易成功"
    except:
        #拒绝交易
        cancel_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", cancel_button)
        print("小狐狸交易失败，点击【拒绝交易】可能是点击不动，不足gas fee：", c)
        return f"小狐狸交易失败，点击【拒绝交易】"

#绑定所有账号，从小狐狸端【首次连接新网络时】
def fox_confirm_connect_account(browser, wait):
    print('进入fox_confirm_connect_account()，【小狐狸】确认连接Account')
    ACC_NUM = 0  #设置总账号数量
    #先刷新网页
    for i in range(1,5):
        try: #等待全选账户按钮出现，但有时没有
            just_wait_select_all_button = browser.find_element(By.XPATH,
                                                               "//input[@class='check-box choose-account-list__header-check-box fa fa-minus-square check-box__indeterminate']")
            print("小狐狸上找到了【全部选择账号】")
            break

        except:
            browser.refresh()
            time_sleep(2)
    #有时网站不需要小狐狸确认连接，因为之前连接过了
    try:
        # select_single_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='choose-account-list__account'][2]")))
        count_acounts = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='choose-account-list__account']")))
        print("Accounts账户个数：", len(list(count_acounts)))
        ACC_NUM = len(list(count_acounts))
        # select_single_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='choose-account-list__account'][2]")))
        #点击“全选”
        select_all_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='check-box choose-account-list__header-check-box fa fa-minus-square check-box__indeterminate']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", select_all_button)
        print('全选Account结束，点击下一步')
    except:
        print("fox_confirm_connect_account(), 小狐狸选择全选Account出错，是否影响？")

    try:
        next_step = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        browser.execute_script("arguments[0].click();", next_step)
        print('下一步点击结束，现在要点击“连接')
    except:
        print("fox_confirm_connect_account(), 小狐狸点击下一步出错，是否影响？")

    time_sleep(2)
    try:
        next_step = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        browser.execute_script("arguments[0].click();", next_step)
        print('点击“连接”结束')
    except:
        print("fox_confirm_connect_account(), 小狐狸可能之前连接过该网站了")

    return ACC_NUM  # 返回统计的Account账户个数

#确认【签名】，从小狐狸端
def fox_confirm_sign(browser, wait):
    print('fox_confirm_sign，【从小狐狸端】确认签名')
    browser.refresh()
    time_sleep(10, "刷新后等待")
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary btn--large request-signature__footer__sign-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("fox_confirm_sign, 小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_sign()小狐狸点击【签名】失败了，是否影响？")
        return "小狐狸签名失败"

#小狐狸【确认】showme
def fox_confirm_OP(browser, wait):
    print('fox_confirm_OP，【从小狐狸端】确认签名，showme、clipper、zipswap、matcha通用')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_OP()小狐狸点击【签名】失败了，是否影响？")
        #尝试去拒绝
        try:
            cancel_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
        except:
            print("小狐狸尝试拒绝失败，是否影响？")
        return "fox_confirm_OP()小狐狸点击【签名】失败了，是否影响？"


#小狐狸【确认】允许使用代币
def fox_confirm_allow_token(browser, wait):
    print('fox_confirm_allow_token, 多种swap 网站通用')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_allow_token 小狐狸点击【签名】失败了，是否影响？")
        #尝试去拒绝
        try:
            cancel_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@class='button btn--rounded btn-secondary page-container__footer-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
        except:
            print("小狐狸尝试拒绝失败，是否影响？")
        return "fox_confirm_allow_token 小狐狸点击【签名】失败了，是否影响？"

#小狐狸【确认】允许Bentobox
def fox_confirm_allow_bentobox(browser, wait):
    print('fox_confirm_allow_bentobox, 多种swap 网站通用')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_allow_bentobox 小狐狸点击【签名】失败了，是否影响？")
        #尝试去拒绝
        try:
            cancel_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@class='button btn--rounded btn-secondary']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", cancel_button)
        except:
            print("小狐狸尝试拒绝失败，是否影响？")
        return "fox_confirm_allow_token 小狐狸点击【签名】失败了，是否影响？"

#小狐狸【确认】showme
def fox_confirm_gmx_referal(browser, wait):
    print('fox_confirm_gmx_referal，【从小狐狸端】确认签名gmx 推荐码')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[3]/footer/button[2]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_gmx_referal 小狐狸点击【签名】失败了，是否影响？")
        return "fox_confirm_gmx_referal 小狐狸点击【签名】失败了，是否影响？"



#小狐狸【确认】showme
def fox_confirm_ARB(browser, wait):
    print('fox_confirm_ARB，【从小狐狸端】确认签名，showme、clipper、zipswap、matcha通用')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary page-container__footer-button']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_ARB()小狐狸点击【签名】失败了，是否影响？")
        return "fox_confirm_ARB()小狐狸点击【签名】失败了，是否影响？"

#小狐狸【确认】galaxy
def fox_confirm_galaxy(browser, wait):
    print('fox_confirm_galaxy，【从小狐狸端】确认galaxy, NFT')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        confirm_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[3]/footer/button[2]')))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", confirm_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        try:#拒绝
            cancle_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[3]/footer/button[1]')))
    
            time_sleep(1)
            browser.execute_script("arguments[0].click();", cancle_button)
            print("fox失败, 已经点击拒绝交易")
            return "fox失败, 已经点击拒绝交易"
        except:
            print("小狐狸拒绝前面失败")
            return "fox失败"



#绑定网络，从小狐狸端：批准——>切换网络
def fox_confirm_connect_network(browser, wait):
    print('fox_confirm_connect_network，【从小狐狸端】确认连接网络')
    #先刷新网页
    for i in range(1,5):
        try:
            just_wait_confirm_button = browser.find_element(By.XPATH, "//button[@class='button btn--rounded btn-primary']")
            break
        except:
            browser.refresh()
            time_sleep(2)

    for i in range(1,5):
        try:
            # 签名
            sign_button = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary btn--large request-signature__footer__sign-button']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", sign_button)
            break
        except:
            print("fox_confirm_connect_network()小狐狸点击【签名】失败了，是否影响？")
            browser.refresh()
            time_sleep(2)

    try:
        #批准，有时会有这个步骤
        confirm_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", confirm_button)
    except:
        print("fox_confirm_connect_network()小狐狸点击【批准】失败了，是否影响？")

    try:
        # 切换网络
        switch_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", switch_button)
    except:
        print("fox_confirm_connect_network()小狐狸点击【切换网络】失败了，是否影响？")

#小狐狸去切换网络
def fox_change_network(browser, wait, keywords):
    time_sleep(3, f"我已经进入fox_change_network，小狐狸去切换网络{keywords}")
    for i in range(1, 5):
        try:  # 找网络下拉的图标
            list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                 "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
            # browser.execute_script("arguments[0].click();", list_button)
            break
        except:
            browser.refresh()
            time_sleep(2)
    # 找到下拉图标
    list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                         "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
    browser.execute_script("arguments[0].click();", list_button)

    # 点击想要的网络
    # s = f"//span[@class='network-name-item' and contains(text(), '{keywords}')]"
    # print(s)
    desired_net = wait.until(
        EC.element_to_be_clickable((By.XPATH, f"//span[@class='network-name-item' and contains(text(), '{keywords}')]")))
    time.sleep(2)
    browser.execute_script("arguments[0].click();", desired_net)
#     pass

#小狐狸本地添加网络
def fox_add_network(browser, wait, keywords):
    print('进入fox_add_network()，【小狐狸】本地添加网络')
    browser.refresh()
    time_sleep(5)
    for i in range(1, 5):
        try: #找网络下拉的图标
            list_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
            # browser.execute_script("arguments[0].click();", list_button)
            break
        except:
            browser.refresh()
            time_sleep(2)
    #找到下拉图标
    list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                         "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
    browser.execute_script("arguments[0].click();", list_button)

    #点击“添加网络”
    add_net_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-secondary']")))
    browser.execute_script("arguments[0].click();", add_net_button)
    if keywords == 'Optimism':
        print("准备添加Optimism网络")
        for i,j in zip(range(1,6), Op_list):
            input_box = wait.until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='networks-tab__add-network-form-body']/div[{i}]//input")))
            input_box.send_keys(j)
        #保存按钮
        save_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        browser.execute_script("arguments[0].click();", save_button)

    if keywords == 'Arbitrum One':
        print("准备添加Arbitrum One网络")
        for i, j in zip(range(1, 6), Arb_list): #需要传一个列表
            input_box = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//div[@class='networks-tab__add-network-form-body']/div[{i}]//input")))
            input_box.send_keys(j)
            time_sleep(2)
        # 保存按钮
        save_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary']")))
        browser.execute_script("arguments[0].click();", save_button)

    # search.send_keys(keywords)
    # search.send_keys(Keys.ENTER)
    #

#小狐狸本地导入私钥
def fox_import_private_key(browser, wait, private_key):
    print("fox_import_private_key()，小狐狸导入私钥")
    # 先刷新网页
    for i in range(1, 5):
        try:  # 找头像
            icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
            break
        except:
            browser.refresh()
            time_sleep(2)

    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", icon_button)
    print("已经点击头像")

    import_account_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='account-menu__item__text' and text() ='导入账户']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", import_account_button)
    print("已经点击【导入账户】")

    input_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='private-key-box']")))
    time_sleep(2)
    input_box.send_keys(private_key)
    print("已经粘贴私钥")

    confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='button btn--rounded btn-primary btn--large new-account-create-form__button']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", confirm_button)
    print("已经点击【导入】")

    #等待导入完成。完成后会自动回到首页
    try:
        time_sleep(2)
        just_wait_icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
    except:
        print("可能导入失败")

#获取小狐狸上的帐号. 因为有时只有密钥, 没有帐号
def fox_get_account(browser, wait):
    print("开始获取小狐狸帐号, 因为有时只有密钥")
    ##=========== 点击账户详情
    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@title='账户选项']")))
    time_sleep(1, "先点击三个点")
    browser.execute_script("arguments[0].click();", icon_button)

    account_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='账户详情']")))
    time_sleep(1, "再点击账户详情")
    browser.execute_script("arguments[0].click();", account_detail_button)

    ##=========== 点击提取account
    account_text_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='qr-code__address-container']/div[1]")))
    account_text = account_text_button.text
    print("======找到的账户是:",account_text)

    ##=========== 点击完成
    close_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='account-modal__close']")))
    time_sleep(1, "点击关闭")
    browser.execute_script("arguments[0].click();", close_button)
    
    return account_text

#小狐狸本地添加代币
def fox_add_token(browser, wait, keywords):
    print("进入fox_add_token()，小狐狸添加代币")
    # 先刷新网页
    for i in range(1, 5):
        try:  # 找导入代币
            import_token_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@class='button btn-link import-token-link__link' and text()='导入代币']")))
            browser.execute_script("arguments[0].click();", import_token_button)
            time_sleep(5)
            break
        except:
            browser.refresh()
            time_sleep(5)
    #点击“代币合约地址”
    token_address_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='custom-address']")))
    token_address_button.send_keys(keywords)

    print("我准备点“添加自定义代币")
    time_sleep(5)
    # 点击“添加自定义代币”
    add_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text() = '添加自定义代币']")))
    browser.execute_script("arguments[0].click();", add_button)

    print("我准备点击“导入代币”")
    time_sleep(5)
    # 点击“导入代币”
    import_confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='导入代币']")))
    browser.execute_script("arguments[0].click();", import_confirm_button)
    time_sleep(5)
    #点击返回
    back_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='asset-breadcrumb']")))
    browser.execute_script("arguments[0].click();", back_button)
    time_sleep(4)
    #
    #############################

#查询、返回小狐狸代币的具体值。前提是调用 get_fox_network_token_balance
#输入两个参数：all_networks=所有的网络（列表），all_token_and_balance（列表，该列表里有字典，字典键值对是网络:币值）
def return_fox_net_token_balance(netname, token, all_networks, all_token_and_balance):
    print("进入return_fox_net_token_balance(), 查询、返回小狐狸代币的具体值")
    clean_all_networks = []  # 将来用来过滤 all_networks
    for i in all_networks:
        matches = re.search('[a-zA-Z]+', i)
        clean_all_networks.append(matches.group())
    try:
        index_num = clean_all_networks.index(netname)
        if index_num >=0: #如果有找到
            token_and_balance = all_token_and_balance[index_num]
            print(f"{netname}上的{token}有：",token_and_balance[token])
            return token_and_balance[token]
    except:
        print(f"{netname}上没有{token}，考虑添加")
        return None

#获取小狐狸有哪些网络、代币、余额，是去网页端爬数据
def get_fox_network_token_balance(browser, wait):
    network_names = []  #到时把网络装在列表里，
    dic_token_balance = [] #以网络名为字典名，字典里的键值对是代币和数额。最后把字典放在这个列表里
    print("进入get_fox_network_token_balance()，获取小狐狸有哪些网络、代币、余额")
    browser.refresh()
    time_sleep(5)
    for i in range(1, 5):
        try: #找网络下拉的图标
            list_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
            time.sleep(2)
            browser.execute_script("arguments[0].click();", list_button)
            break
        except:
            browser.refresh()
            time_sleep(2)
    #点击网络列表图标
    # list_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
    # browser.execute_script("arguments[0].click();", list_button)

    networks = wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//span[@class='network-name-item']")))
    networks_num = len(list(networks))
    print("网络个数：", networks_num)

    #遍历网络，并且在遍历的过程中，读取网络的代币和余额
    for i in range(0,networks_num):
        browser.refresh()
        time_sleep(5)
        #点击下拉图标
        list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                             "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
        browser.execute_script("arguments[0].click();", list_button)
        #获取全部网络，点击第 i 个网络
        networks = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//span[@class='network-name-item']")))
        network_names.append(networks[i].text)  # 把先网络名加到列表里
        print("==网络名是：", networks[i].text)

        #切换到该网络
        browser.execute_script("arguments[0].click();", networks[i])
        time_sleep(5)
        #获取该网络下的所有代币和余额
        tokens = wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//button[@class='asset-list-item__token-button']")))
        tokens_num = len(list(tokens))
        #解析代币和余额
        a = f"network_{i}"
        a = dict()  # 以网络名，新建一个字典，用于存下面的代币和金额
        for j in range(0,tokens_num):
            token_balance = tokens[j].get_attribute('title')
            print("代币和余额是：", token_balance)
            balance, token = token_balance.split()
            a[token] = balance
        dic_token_balance.append(a)
        #找代币# '//span[@class='asset-list-item__token-symbol']'
        # //button[@class='asset-list-item__token-button']

    # print("get_fox_network_token_balance, 网络名分别是：",network_names)
    #切换不同网络，读取该网络下的余额
    # for i in range(0,a)

    return network_names, dic_token_balance

#获取小狐狸有哪些网络，大概率用不上了
def get_fox_network(browser, wait):
    network_names=[] #创建一个空列表，用于存网络名
    print("进入get_fox_network()，获取小狐狸有哪些网络")
    # 先刷新网页
    for i in range(1, 5):
        try: #找网络下拉的图标
            list_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
            browser.execute_script("arguments[0].click();", list_button)
            break
        except:
            browser.refresh()
            time_sleep(2)
    #点击网络列表图标
    # list_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")))
    # browser.execute_script("arguments[0].click();", list_button)

    networks = wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//span[@class='network-name-item']")))
    a = len(list(networks))
    print("网络个数：", a)
    for i in range(0,a):
        network_names.append(networks[i].text)

    print("网络名分别是：",network_names)
    return network_names

#获取小狐狸在OP上的ETH有多少————》扩展成获取所有余额，大概率用不上了
def get_fox_balance(browser, wait):
    balance={} #创建一个空字典，用于存余额
    print("进入get_fox_balance()，获取小狐狸的资产分布情况")
    # 先刷新网页
    for i in range(1, 5):
        try: #找余额的图标
            just_wait_eth_button = browser.find_element(By.XPATH,"//span[@class='currency-display-component__text']")
            break
        except:
            browser.refresh()
            time_sleep(2)
    #点击网络列表图标
    list_button = browser.find_element(By.XPATH,"//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")
    browser.execute_script("arguments[0].click();", list_button)
    #查看有多少个网络，注意用 find_elements
    network_lists = browser.find_elements(By.XPATH,"//div[@class='network-dropdown-list']/li")
    a = len(list(network_lists))
    print(f"总共有{a}个网络")
    browser.execute_script("arguments[0].click();", list_button)
    for i in range(0, a):
        # 点击网络列表图标
        list_button = browser.find_element(By.XPATH,
                                           "//div[@class='network-display network-display--clickable chip chip--with-left-icon chip--with-right-icon chip--border-color-border-default chip--background-color-undefined chip--max-content']")
        browser.execute_script("arguments[0].click();", list_button)
        #查看全部的网络
        network_lists = browser.find_elements(By.XPATH, "//div[@class='network-dropdown-list']/li")
        # 切换想要的网络
        browser.execute_script("arguments[0].click();", network_lists[i])
        browser.refresh()
        time_sleep(5)
        #查询余额
        balance_button = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//span[@class='currency-display-component__text']")))
        #查询网络名称
        net_name = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//span[@class='box box--margin-top-1 box--margin-right-0 box--margin-bottom-1 box--margin-left-0 box--flex-direction-row typography chip__label typography--h7 typography--weight-normal typography--style-normal typography--color-text-alternative']")))
        # print(f"网络{net_name.text}的余额是{balance_button.text}")
        balance[net_name.text] = balance_button.text
    print(balance)
    return balance

#获取小狐狸在OP上的ETH有多少————》扩展成获取所有余额，大概率用不上了
def get_fox_Ethereum_ETH_balance(browser):
    balance={} #创建一个空字典，用于存余额
    print("进入get_fox_Ethereum_ETH_balance，获取小狐狸的资产分布情况")
    # 先刷新网页
    for i in range(1, 5):
        try: #找余额的图标
            just_wait_eth_balance_button = browser.find_element(By.XPATH,"//span[@class='currency-display-component__text']")
            time_sleep(2)
            break
        except:
            browser.refresh()
            time_sleep(5)
    eth_balance_button = browser.find_element(By.XPATH, "//span[@class='currency-display-component__text']")
    time_sleep(5)
    print("主网上的ETH余额是：", eth_balance_button.text)
    return eth_balance_button.text

#获取小狐狸有多少个Account
def get_fox_accounts(browser, wait):
    print("get_fox_accounts()，获取小狐狸有多少个Account")
    #先刷新
    browser.refresh()
    time_sleep(5)

    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
    browser.execute_script("arguments[0].click();", icon_button)
    accounts = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='account-menu__account account-menu__item--clickable']")))
    print(f"总共有{len(list(accounts))}个账户Account")
    ACC_NUM = len(list(accounts))

    return ACC_NUM #返回全部的账户

#小狐狸切换账号，i表示第i个账号
def fox_change_account(browser, wait, i):
    try:
        print("进入fox_change_account()，小狐狸切换账号")
        #先刷新
        # browser.refresh()
        # time_sleep(5, "刷新网页后等待")

        #先点击头像，展开账户
        icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='identicon__address-wrapper']")))
        time_sleep(2, "先点击头像，展开账户")
        browser.execute_script("arguments[0].click();", icon_button)
        print("已经点击展开账户")
        #索引第 i 个账号
        # account = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='account-menu__account account-menu__item--clickable'][4]")))
        account = wait.until(EC.element_to_be_clickable((By.XPATH, f"//button[@class='account-menu__account account-menu__item--clickable'][{i}]")))
        time_sleep(1)
        browser.execute_script("arguments[0].click();", account)
        time_sleep(5, "点击切换账号完成")
    except:
        print("换号失败, 可能是之前小狐狸没有拒绝")
        #可能是之前小狐狸没有拒绝
        try:#拒绝
            cancle_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]//button[text()="拒绝"]')))
            time_sleep(1)
            browser.execute_script("arguments[0].click();", cancle_button)
            print("fox失败, 已经点击拒绝交易")
            return "fox失败, 已经点击拒绝交易"
        except:
            print("小狐狸拒绝前面失败")
            return "fox失败"

def fox_get_rinkeby_balance(browser, wait):
    # time_sleep(3600, "waiting..........")
    # time_sleep(3600, "waiting..........")
    # //div[@class='tabs__content']/div[2]//span[@class='asset-list-item__token-value']
    # all_balance = browser.ele("//button[@class='asset-list-item__token-button']")
    try:
        token_1_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tabs__content']/div[1]//span[@class='asset-list-item__token-symbol']")))
        token_1_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tabs__content']/div[1]//span[@class='asset-list-item__token-value']")))
        token_1_balance = token_1_balance_button.text
        print("=========== 1 获取到的名字是: ", token_1_name.text)
        print("=========== 1 获取到的余额是: ", token_1_balance)
    except:
        print("可能没有Rinkeby ETH")
        token_1_balance = 0
        
    try:
        token_2_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tabs__content']/div[2]//span[@class='asset-list-item__token-symbol']")))
        token_2_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tabs__content']/div[2]//span[@class='asset-list-item__token-value']")))
        token_2_balance = token_2_balance_button.text
        print("=========== 2 获取到的名字是: ", token_2_name.text)
        print("=========== 2 获取到的余额是: ", token_2_balance)
    except:
        print("可能没有 link 代币")
        token_2_balance = 0


    return token_1_balance, token_2_balance
    # balance = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='asset-list-item__token-button']")))
    # all_balance = browser.find_elements(By.XPATH, "//button[@class='asset-list-item__token-button']")
    # for balance in all_balance:
    # text = balance.get_attribute('title')
    # print("===========获取到的余额是: ", text)
    #//button[@class='asset-list-item__token-button'] 

#导出小狐狸的帐号, 密钥
def fox_account_key_detail(browser, wait):
    ##=========== 点击账户详情
    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@title='账户选项']")))
    time_sleep(1, "先点击三个点")
    browser.execute_script("arguments[0].click();", icon_button)

    account_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='账户详情']")))
    time_sleep(1, "再点击账户详情")
    browser.execute_script("arguments[0].click();", account_detail_button)

    ##=========== 点击提取account
    account_text_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='qr-code__address-container']/div[1]")))
    account_text = account_text_button.text
    print("账户详情是: ", account_text)

    ##=========== 点击导出私钥匙
    key_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='导出私钥']")))
    time_sleep(1, "点击导出私钥匙")
    browser.execute_script("arguments[0].click();", key_detail_button)
    
    #输入密码
    send_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='password']")))
    time_sleep(1, "输入密码")
    send_detail_button.send_keys("12345678")

    confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='确认']")))
    time_sleep(1, "点击导出私钥匙")
    browser.execute_script("arguments[0].click();", confirm_button)
    try:
        #获取text
        key_text_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='export-private-key-modal__private-key-display']")))
        key_text = key_text_button.text
        print("私钥是: ", key_text)
    except:
        time_sleep(3, "出错")

    ##=========== 点击完成
    key_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='完成']")))
    time_sleep(1, "点击完成")
    browser.execute_script("arguments[0].click();", key_detail_button)
    return account_text, key_text
##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ 小狐狸的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ zipswap的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#判断 zipswap 是否要连接钱包
def zipswap_whether_connect_wallet(browser, wait):
    print("我已进入zipswap_whether_connect_wallet，判断zipswap是否要连接钱包")
    try:
        zipswap_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='sc-jwdaid-5 gvwkgl']")))
        print("zipswap按钮是", zipswap_connect_wallet.text)
        return zipswap_connect_wallet.text
    except:
        print("没有找到 zipswap 上的 connect wallet 按钮")
        return None

# zipswap 准备连接小狐狸钱包
def zipswap_connect_wallet(browser, wait):
    #先等全部元素加载出来，
    print("我已进入zipswap_connect_wallet，zipswap准备连接小狐狸钱包")
    for i in range(0, 5): #最多刷新5次
        print(f"刷新第{i}次（最多5次）预防 pika 加载不完全")
        try:
            zipswap_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='sc-jwdaid-5 gvwkgl']")))
            break
        except:
            browser.refresh()
            time_sleep(5)
            continue
    try:
        zipswap_connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='sc-jwdaid-5 gvwkgl']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", zipswap_connect_wallet)
        time_sleep(2)
    except:
        print("zipswap_connect_wallet()中，点击connect wallet失败了")

    try:#选择小狐狸
        metamask = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='connect-METAMASK']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", metamask)
        time_sleep(2)
    except:
        print("zipswap_connect_wallet()中，点击connect wallet后，再选择小狐狸失败了，是否影响？")


def allow_zipswap_use_USDC(browser, wait):
    print("我已进入allow_zipswap_use_USDC，允许zipswap 访问USDC")
    try:
        zipswap_connect_USDC = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@style = 'display: flex; align-items: center;']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", zipswap_connect_USDC)
        time_sleep(2)
    except:
        print("allow_zipswap_use_USDC()中，点击connect USDC失败了")

#判断 zipswap 准备转账
def zipswap_prepare_transfer(browser, wait, L2_ETH_value, from_token, to_token):
    print("我已进入zipswap_prepare_transfer，zipswap准备选择代币、开始转账")

    if from_token == "USDC":
        print("应该交换一下from 、to 的位置")
        # =======选择 from, 列表
        try:
            select_a_token = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//div[@id='swap-currency-input']//span[@class='sc-1w2qsjp-7 jA-DiSO']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", select_a_token)

            # =======选择 from 要转到的代币
            search_a_token = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
            time_sleep(2)
            search_a_token.send_keys(from_token)
            time_sleep(5)
            search_a_token.send_keys(Keys.ENTER)
            time_sleep(2)

            # =======选择 to, “select a token"要转到的代币
            select_a_token = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//span[@class='sc-1w2qsjp-9 cMXIzw token-symbol-container' and text()='Select a token']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", select_a_token)

            # =======选择 to 要转到的代币
            search_a_token = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
            time_sleep(2)
            search_a_token.send_keys(to_token)
            time_sleep(5)
            search_a_token.send_keys(Keys.ENTER)
            time_sleep(2)
        except:
            print("可能已经交换了")
     # =======输入金额
    input_amount = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//div[@id='swap-currency-input']//input")))
    time_sleep(3,"找金额输入框")
    #=======随机金额
    point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
    try_times = 1
    while float(L2_ETH_value) - input_value < 0.005: #如果会使余额小于 0.005
        point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.3, L2_ETH_value * 0.7), point)
        try_times = try_times + 1
        if try_times == 100:
            print("余额不足，直接报错")
            return 0  #说明余额不做，直接报错即可

    print(f"本次zipswap从 {from_token} 随机转到 {to_token} 的金额是：{input_value}，将来预估余额是：{float(L2_ETH_value) - input_value}")
    input_amount.send_keys(str(input_value))

    #======= 可能要授权使用USDC
    if from_token != "ETH":
        try:
            allow_zipswap_use_USDC(browser, wait)
            switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
            fox_info = fox_confirm_OP(browser, wait)
            print("尝试允许zipswap 访问 USDC，结果：", fox_info)
            switch_tab_by_handle(browser, 2, 0)  #
        except:
            switch_tab_by_handle(browser, 2, 0)  #
            print("可能之前授权过了")

    # ==== 提交订单 Swap
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='swap-button']")))
    time_sleep(8,"等待提交订单")
    browser.execute_script("arguments[0].click();", swap_button)

    # ======comfirm swap
    try:
        confirm_swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='confirm-swap-or-send']")))
        time_sleep(3,"comfirm swap")
        browser.execute_script("arguments[0].click();", confirm_swap_button)
        return f"OP zipswap 任务完成，本次 zipswap 从 {from_token} 转到 {to_token} 的金额是：{input_value}；"
    except:
        print("【失败】OP zipswap")
        return "【失败】OP zipswap"

# 实时获取 op 上eth 的金额多少
def get_OP_ETH_and_select_from_to_token_by_zipswap(browser, wait, from_token, to_token):
    print("我已进入get_OP_ETH_and_select_from_to_token_by_zipswap，实时获取 OP 上 最大代币 的金额多少，确定用什么代币")

    # =======选择 from, 列表
    select_a_token = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//div[@id='swap-currency-input']//span[@class='sc-1w2qsjp-7 jA-DiSO']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", select_a_token)

    # =======选择 from 要转到的代币
    search_a_token = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(2)
    search_a_token.send_keys(from_token)
    time_sleep(5)
    search_a_token.send_keys(Keys.ENTER)
    time_sleep(2)

    # =======选择 to, “select a token"要转到的代币
    select_a_token = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//span[@class='sc-1w2qsjp-9 cMXIzw token-symbol-container' and text()='Select a token']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", select_a_token)

    # =======选择 to 要转到的代币
    search_a_token = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(2)
    search_a_token.send_keys(to_token)
    time_sleep(5)
    search_a_token.send_keys(Keys.ENTER)
    time_sleep(2)

    # === 查找最大金额
    try:
        max_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="swap-currency-input"]/div/div[2]/div/div[1]/div')))

        print("zipswap找到的按钮是", max_balance_button.text) #返回的是 "balance 0.0332 ETH"
        max_a = max_balance_button.text
        max_b = max_a.split()
        max_c = max_b[1]

        print(f"分割后，最终找到的 ETH 金额是：{max_c}")

        return float(max_c)
    except:
        print("get_OP_ETH_by_zipswap，没有找到 OP 上代币的金额")

##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ zipswap的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ clipper的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

# 实时获取 op 上eth 的金额多少
def get_OP_ETH_and_select_from_to_token_by_clipper(browser, wait, from_token, to_token):
    print("我已进入get_OP_ETH_and_select_from_to_token_by_clipper，实时获取 OP 上 最大代币 的金额多少")

    # ==== 一, 找 from
    select_a_token = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//div[@id="rfq-container"]/div/div[2]/div[1]//div[@class="token"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", select_a_token)
   
    ###======新方法,用搜索法

    # from 输入框
    from_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Name, symbol or address']")))
    time_sleep(3)
    from_button.send_keys(from_token)
    
    time_sleep(3,"再确认选择该代币, 第一个")
    first_token_xpath = "//div[@id='tokenItemsDiv']/div[1]/div"
    select_first_token_button = wait.until(EC.element_to_be_clickable((By.XPATH, first_token_xpath)))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", select_first_token_button)
    time_sleep(6,"已经选择 from 代币, 等待响应")

    # ==== 二, 找 to ，下拉框
    to_token_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='token']/span[text()='Select token']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", to_token_list)

    #to 输入框
    to_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Name, symbol or address']")))
    time_sleep(3)
    to_button.send_keys(to_token)
    
    time_sleep(3,"再确认选择该代币, 第一个")
    first_token_xpath = "//div[@id='tokenItemsDiv']/div[1]/div"
    select_first_token_button = wait.until(EC.element_to_be_clickable((By.XPATH, first_token_xpath)))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", select_first_token_button)
    time_sleep(6,"已经选择 to 代币, 等待响应")

    #尝试寻找余额
    try:
        max_balance_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="rfq-container"]/div/div[2]/div[1]/div[1]/small')))
        print("clipper找到的按钮是", max_balance_button.text)  # 返回的是 "Balance 0.005505 ETH"
        max_a = max_balance_button.text
        max_b = max_a.split()
        max_c = max_b[1]

        print(f"分割后，最终找到的金额是：{max_c}")
    except:
        print(f"get_OP_ETH_and_select_from_to_token_by_clipper，没有找到 OP 上 {from_token} 的金额")
    # 下面是 to 的金额
    # try:
    #     USDC_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="rfq-container"]/div/div[2]/div[3]/div[1]/small')))
    #     print("clipper找到的按钮是", USDC_balance_button.text)  # 返回的是 "Balance 0.005505 ETH"
    #     USDC_a = USDC_balance_button.text
    #     USDC_b = USDC_a.split()
    #     USDC_c = USDC_b[1]
    #     print(f"分割后，最终找到的金额是：{USDC_c}")
    # except:
    #     print(f"get_OP_ETH_and_select_from_to_token_by_clipper，没有找到 OP 上 {to_token} 的金额")
    return float(max_c)

#判断 clipper 是否要连接钱包
def clipper_whether_connect_wallet(browser, wait):
    print("我已进入clipper_whether_connect_wallet，判断clipper是否要连接钱包")
    try:
        clipper_connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='__next']/div[1]/div/header/div[2]/div/button[1]/span[1]")))
        print("clipper找到的按钮是", clipper_connect_wallet.text)
        return clipper_connect_wallet.text
    except:
        print("没有找到 clipper 上的 connect wallet 按钮")
        return None

# clipper 准备连接小狐狸钱包
def clipper_connect_wallet(browser, wait):
    #先等全部元素加载出来，
    print("我已进入clipper_connect_wallet，clipper准备连接小狐狸钱包")
    for i in range(0, 5): #最多刷新5次
        print(f"刷新第{i}次（最多5次）预防 clipper 加载不完全")
        try:
            clipper_connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='__next']/div[1]/div/header/div[2]/div/button[1]/span[1]")))
            break
        except:
            browser.refresh()
            time_sleep(5, f"clipper_connect_wallet，clipper 加载网页不完全，第{i}次刷新")
            continue
    try:
        clipper_connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='__next']/div[1]/div/header/div[2]/div/button[1]/span[1]")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", clipper_connect_wallet)
        time_sleep(5)
    except:
        print("clipper_connect_wallet()中，点击connect wallet失败了")

    try:#选择 OP 网络
        time_sleep(2, "准备找 OP 网络")
        OP_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='MuiTypography-root MuiFormControlLabel-label MuiTypography-body1' and text() = 'Optimism Mainnet']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", OP_button)
    except:
        print("clipper_connect_wallet()中，选择 OP 网络失败了，是否影响？")

    try:#选择小狐狸
        time_sleep(2, "准备找小狐狸")
        metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='connect-METAMASK']/span[2]")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", metamask_button)
    except:
        print("clipper_connect_wallet()中，点击connect wallet后，再选择小狐狸失败了，是否影响？")


def allow_clipper_use_USDC(browser, wait):
    print("我已进入allow_clipper_use_USDC，允许zipswap 访问USDC")
    try:
        clipper_connect_USDC = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="rfq-container"]/div/div[4]/div')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", clipper_connect_USDC)
        time_sleep(8)
        return False
    except:
        print("allow_clipper_use_USDC()中，点击connect USDC失败了，是否影响？")
        return True

#判断 clipper 准备转账
def clipper_prepare_transfer(browser, wait, L2_max_value, from_token, to_token):
    print("我已进入clipper_prepare_transfer，clipper 选择代币、转账")
    
    #==============输入交易金额
    time_sleep(5,"准备输入金额")
    input_amount = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='rfq-container']/div/div[2]/div[1]/div[2]/input")))
    time_sleep(2)

    # 随机金额
    point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_max_value * 0.7, L2_max_value * 0.8), point)
    try_times = 0
    while float(L2_max_value) - input_value < 0.005:
        point = random.randint(3, 4)  # 最起码保留3位小数
        input_value = round(random.uniform(L2_max_value * 0.3, L2_max_value * 0.7), point)
        try_times = try_times + 1
        if try_times == 100:
            print("余额不足，直接报错")
            return 0 #说明余额不足，直接报错即可
    print(f"本次从 {from_token} 转到 {to_token} 的随机金额是{input_value}，将来预估余额是：{float(L2_max_value) - input_value}")
    input_amount.send_keys(str(input_value))
    time_sleep(10,"已经输入金额")

    # 如果from token 是 USDC，则允许网站访问 USDC
    if from_token != "ETH":
        try:
            error_status = allow_clipper_use_USDC(browser, wait)
            if error_status == False: #如果没有出错，即确实需要授权
                switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
                fox_info = fox_confirm_OP(browser, wait)
                time_sleep(10)
                print(f"尝试允许clipper 访问 {from_token}，结果：", fox_info)
                switch_tab_by_handle(browser, 2, 0)  # 切换到第0个标签页：小狐狸
        except:
            print("可能之前授权了 clipper 访问 USDC")
            switch_tab_by_handle(browser, 2, 0)  # 切换到第

    #===== 可能提醒是不是还在线 Are you still aboard?
    try:
        continue_button = browser.find_element(By.XPATH, "//span[@class='MuiButton-label' and text()='Continue']")
        time_sleep(2)
        browser.execute_script("arguments[0].click();", continue_button)
    except:
        print("clipper 还在线")

    # # 提交订单
    try:
        print("准备提交订单")
        accept_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='rfq-container']/div/div[4]/button/span[1]")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", accept_button)
        time_sleep(8,"已经点击提交订单")
        return f"OP clipper 提交订单成功，本次从 {from_token} 转到 {to_token} 的随机金额是{input_value}；"
    except:
        print(f"提交订单失败")
        return f"【失败】OP clipper 提交订单失败；"
    # #comfirm swap
    # confirm_swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='confirm-swap-or-send']")))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_swap_button)

##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ clipper的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ showme的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#判断showme是否要连接钱包
def showme_whether_connect_wallet(browser, wait):
    print("我已进入showme_whether_connect_wallet，判断showme是否要连接钱包")
    try:
        pika_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='KastelovSemiBold']")))
        print("showme要连接钱包，因为按钮是", pika_connect_wallet.text)
        return pika_connect_wallet.text
    except:
        print("没有找到 showme 上的 connect wallet 按钮")
        return None

#showme准备连接小狐狸钱包
def showme_connect_wallet(browser, wait):
    #先等全部元素加载出来，
    print("我已进入showme_connect_wallet，showme准备连接小狐狸钱包")
    for i in range(0, 5): #最多刷新5次
        print(f"刷新第{i}次（最多5次）预防 pika 加载不完全")
        try:
            just_wait_select_coin = browser.find_element(By.XPATH, "//span[@class='KastelovSemiBold' and text()='Connect Wallet']")
            break
        except:
            browser.refresh()
            time_sleep(5)
            continue
    try:
        connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='KastelovSemiBold' and text()='Connect Wallet']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", connect_wallet)
        time_sleep(2)
    except:
        print("showme_connect_wallet()中，点击connect wallet失败了")

    try:
        metamask = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='style_button__1PX2t']//span")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", metamask)
    except:
        print("pika_connect_wallet()中，点击connect wallet后，再选择小狐狸失败了，是否影响？")

#showme准备领取NFT
def show_claim_NFT(browser, wait):
    #先等全部元素加载出来，
    switch_tab_by_handle(browser, 2, 0)  # 切换到showme
    time_sleep(5, "我已进入show_claim_NFT，showme领取NFT")
    for i in range(0, 5): #最多刷新5次
        print(f"刷新第{i}次（最多5次）预防 showme 加载不完全")
        try:
            join_club = browser.find_element(By.XPATH, "//span[text()='Join Club']")
            break
        except:
            browser.refresh()
            time_sleep(5)
            continue

    try:
        join_club = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Join Club']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", join_club)
        time_sleep(5,"已经点击join club")
    except:
        print("showme_connect_wallet()中，点击join club失败了")

    try:
        claim_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Claim']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", claim_button)
        time_sleep(8, "已经点击 Claim ")
    except:
        print("pika_connect_wallet()中，点击claim 按钮button")

##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ showme的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ Pika的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
#通过 Pika，实时查看 USDC的金额
def get_OP_USDC_by_pika(wait):
    print("我已进入get_OP_USDC_by_pika，实时获取 OP 上 ETH 的金额多少")
    try:
        balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='availabl-balance svelte-1ldo7z']/a")))
        print("pika 上找到的按钮是", balance_button.text) #返回的是 "0.0332 ETH"
        a = balance_button.text
        b = a.split()
        c = b[0]
        print("分割后，最终找到的ETH金额是:", c)
        return float(c)
    except:
        print("get_OP_USDC_by_pika，没有找到 OP 上ETH 的金额")

#判断pika是否要连接钱包
def pika_whether_connect_wallet(browser, wait):
    print("我已进入pika_whether_connect_wallet，判断pika是否要连接钱包")
    try:
        pika_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, "//header[@class='big svelte-1nltbvo']//button[@class='submit-btn svelte-6hhyej']")))
        print("pika要连接钱包，因为按钮是", pika_connect_wallet.text)
        return pika_connect_wallet.text
    except:
        print("没有找到pika 上的 connect wallet 按钮")
        return None

#pika准备连接小狐狸钱包
def pika_connect_wallet(browser, wait):
    #先等全部元素加载出来，有时pika上的元素加载不完全
    print("我已进入pika_connect_wallet，pika准备连接小狐狸钱包")
    for i in range(0, 5): #最多刷新5次
        print(f"刷新第{i}次（最多5次）预防 pika 加载不完全")
        try:
            just_wait_select_coin = browser.find_element(By.XPATH, "//div[@class='product-wrap product-input svelte-13vjqxo']")
            break
        except:
            browser.refresh()
            time_sleep(5)
            continue
    try:
        connect_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, "//header[@class='big svelte-1nltbvo']//div[@class='wallet svelte-6hhyej']/button[@class='submit-btn svelte-6hhyej']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", connect_wallet)
        time_sleep(2)
    except:
        print("pika_connect_wallet()中，点击connect wallet失败了")

    try:
        metamask = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='buttons svelte-18sopm6']/button[@id='connect-MetaMask']")))
        browser.execute_script("arguments[0].click();", metamask)
    except:
        print("pika_connect_wallet()中，点击connect wallet后，再选择小狐狸失败了，是否影响？")

#pika连接网络，
def pika_connect_network(browser, wait):
    #先等全部元素加载出来
    print("我已进入pika_connect_network，pika连接网络")
    for i in range(1, 10): #最多刷新九次
        print(f"刷新第{i}次")
        try:
            print("我进入try")
            just_wait_select_coin = browser.find_element(By.XPATH, "//div[@class='product-wrap product-input svelte-13vjqxo']")
            # select_coin = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='product-wrap product-input svelte-13vjqxo']")))
            # browser.execute_script("arguments[0].click();", select_coin)
            break
        except:
            browser.refresh()
            time_sleep(1)
            continue

    print("开始连接网络")
    connect_net = wait.until(EC.element_to_be_clickable((By.XPATH, "//header[@class='big svelte-1nltbvo']//div[@class='wallet svelte-6hhyej']/button[@class='submit-btn svelte-6hhyej']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", connect_net)
    time_sleep(2)

# pika输入margin
def pika_input_margin(wait, USDC_Balance):
    print("我已进入pika_input_margin，准备输入随机交易金额")
    input_margin = wait.until(EC.element_to_be_clickable((By.XPATH, "//main/div[8]/div[2]/div/div[3]/div[5]/input")))
    time_sleep(2,"已经找到输入框")

    #设置随机交易金额
    point = random.randint(0, 3)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    input_value = round(random.uniform(USDC_Balance * 0.2, USDC_Balance * 0.4), point)  # 随机金额
    while input_value > float(USDC_Balance):
        point = random.randint(0, 1)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
        input_value = round(random.uniform(USDC_Balance * 0.2, USDC_Balance * 0.4), point)  # 随机金额
    time_sleep(3,f"随机转账的USDC金额是{input_value}")
    input_margin.send_keys(str(input_value))

    #设置随机杠杆
    # input_leverage = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='range svelte-13vjqxo']/input")))
    # time_sleep(2)
    # point = random.randint(0, 2)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    # leverage = round(random.uniform(1, all_coin), point)  # 随机金额
    # print("随机转账的USDC金额是：", leverage)
    # if leverage *
    #
    # input_margin.send_keys(input_value)
    # browser.execute_script("arguments[0].click();", input_margin)

# 首次交易pika，可能出现approve
def pika_try_approve_or_submit(browser, wait):
    print("我已进入pika_try_approve，pika首次交易可能有approve")
    # 点击“Approve”，有时候有
    approve = 0
    approve_or_try_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//body/div[3]/main/div[8]/div[2]/div/div[3]/button")))
    time_sleep(2)

    if approve_or_try_button.text == 'Approve': #说明需要approve，后续小狐狸需要同意 approve
        print("pika有approve")
        browser.execute_script("arguments[0].click();", approve_or_try_button)
        approve = 1
        return approve
    else:
        # browser.execute_script("arguments[0].click();", approve_or_try_button)
        print("pika没有approve，直接交易")
        time_sleep(2)
        approve = 0
        return approve

# pika确认交易
def pika_confirm(browser, wait):
    print("我已进入pika_confirm，pika确认交易")
    #先点submit
    sutmit_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//body/div[3]/main/div[8]/div[2]/div/div[3]/button")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", sutmit_button)
    time_sleep(2)

    # 点击“confirm”
    confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='svelte-jz1hsz']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", confirm_button)
    time_sleep(2)


#pika取消交易
def pika_close_transaction(browser, wait):
    print("我已进入pika_close_transaction，pika取消交易")
    #点击close
    close_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='big svelte-17ozs3n']//td[@class='btn_wrap svelte-dccq2l']/button[@class='close svelte-dccq2l']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", close_button)
    time_sleep(2)
    #选择100%
    amount_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='cal-btn svelte-12umri2' and text()='100%']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", amount_button)
    time_sleep(2)
    #点击 close_position_button
    close_position_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='svelte-jz1hsz' and text()='Close Position']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", close_position_button)

##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ Pika的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

#这是谁家的？
def choose_coin(browser, wait):
    # 选择币种
    select_coin = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='product-wrap product-input svelte-13vjqxo']")))
    browser.execute_script("arguments[0].click();", select_coin)
    print('选择货币结束')

def verify_next(browser, wait):
    print("开始点击verify_next")
    verify_next = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[2]/div/div/div[2]/form/div[5]/div[1]/button/span')))
    browser.execute_script("arguments[0].click();", verify_next)
    print('verify_next点击结束')

def email_submit(browser, wait, email):
    print("点击邮箱输入框")
    email_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='email1']")))
    email_input.send_keys(email)
    print('email输入结束，等待会')
    time_sleep(2)


    print("开始点击提交按钮")
    submit = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/button')))
    browser.execute_script("arguments[0].click();", submit)
    print('点击提交结束，等待会')
    time_sleep(3)

def just_wait_captcha(browser):
    TIME_OUT1 = 300  # 最长等待5分钟
    wait_1 = WebDriverWait(browser, TIME_OUT1)  # wait是一个类
    print("开始点击2captcha")
    captcha_button = wait_1.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="captcha-solver-info"]')))
    browser.execute_script("arguments[0].click();", captcha_button)
    print("已经点击2captcha，等待验证")

    #等待验证码成功后的某个特征
    # just_wait = wait_1.until(EC.invisibility_of_element((By.XPATH, '//*[@id="input_4"]')))
    continue_button = wait_1.until(EC.element_to_be_clickable((By.XPATH, '//button[@type="submit"]')))
    browser.execute_script("arguments[0].click();", continue_button)
    print("验证成功")
    time_sleep(100)

# 切换标签页
# 第一个数字参数 handle_index 表示句柄索引，第二个数字参数 refresh 表示切换过去后是否刷新页面，
def switch_tab_by_handle(browser, handle_index, refresh):
    handles = browser.window_handles
    browser.switch_to.window(handles[handle_index])
    if refresh == 1:
        time_sleep(2)
        browser.refresh()
        time_sleep(5)

# 新建标签页
def new_tab(browser, url):
    print(f"我已进入new_tab，正在打开{url}")
    new_window = 'window.open("{}")'.format(url)  # js函数，此方法适用于所有的浏览器
    browser.execute_script(new_window)
    time_sleep(3)

def open_clash_dashboard(browser, wait, url_dashboard):
    print("我已经进入open_clash_dashboard")
    browser.get(url_dashboard)
    for i in range(1,5): #最多刷新九次
        print(f"刷新第{i}次")
        try:
            global_expand = browser.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/div/div[1]/div/ul/li[1]/div/div[2]/div/span')
            # global_expand = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='tags-expand']")))
            break
        except:
            browser.refresh()
            time_sleep(3, f"clash_web还未加载出来，刷新第{i}次")

    print("全局代理，展开节点")

    global_expand = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[1]/div/ul/li[1]/div/div[2]/div/span')))
    time.sleep(2)
    browser.execute_script("arguments[0].click();", global_expand)
    #先点击测速，否则节点会出现红色erro
    speed_test = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='proxies-speed-test']")))
    time.sleep(1)
    browser.execute_script("arguments[0].click();", speed_test)
    time_sleep(10, "先进行测速处理，防止clash节点出错")
    IPs = wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[1]/div/ul/li[1]/div/div[2]/div/ul/li[@class="cursor-pointer"]')))
    print("节点个数（需要去掉一些）：", len(list(IPs)))
    return 1

def ip_switcher(browser, wait, url_google):
    print("我已进入ip_switcher")
    IPs = wait.until(
        EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[1]/div/ul/li[1]/div/div[2]/div/ul/li[@class="cursor-pointer"]')))
    for i in range(3, len(list(IPs)) - 26):
        print(f"进入ip_switcher的for循环,第{i}次")
        switch_tab_by_handle(browser, 0, 0)  # 先切换到 clash 窗口
        ip_num = random.randint(3, len(list(IPs)) - 26)  # 随机取一个 ip 序列
        browser.execute_script("arguments[0].click();", IPs[ip_num])
        # 如果能打开谷歌，则说明如果响应正常，说明翻墙成功
        new_tab(browser, url_google)
        switch_tab_by_handle(browser, 1, 0)  # 切换到谷歌页面
        try:
            google_search = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[3]/center/input[1]')))
            print(f"我已打开谷歌，说明已挂上代理，现在用的是随机ip_{ip_num}")
            print("开始执行任务")
            # browser.close() #关闭该标签页
            break
        except:
            print("谷歌打不开，换下一个代理")
            browser.close()  # 关闭该标签页
            continue


def random_select_clash_ip(browser, wait):
    print("我已进入random_select_clash_ip,随机选择IP")
    #这是一个列表
    IPs = wait.until(
        EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="root"]/div/div[2]/div/div[1]/div/ul/li[1]/div/div[2]/div/ul/li[@class="cursor-pointer"]')))
    ip_num = random.randint(3, len(list(IPs)) - 26)  # 随机取一个 ip 序列
    browser.execute_script("arguments[0].click();", IPs[ip_num])
    time_sleep(5, "已经切换了随机节点")
    

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ zksync的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
#等待zk 网页加载完成
def wait_zk_tab(browser,wait):
    time_sleep(3, "进入wait_zk_tab(), 等待zk 网页加载完成")
    for i in range(1,5):
        try:
            just_wait_Wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tileName' and text()='Ethereum Wallet']")))
            break
        except:
            browser.refresh()
            time_sleep(10,"wait_zk_tab(), 网页加载未完成，刷新")

#获取 zk 的q金额，返回的是字典，包括所有代币
def get_zk_all_balance(wait):
    time_sleep(3, "进入get_zk_all_balance(), 准备去获取 zk 余额")
    all_balance = {}
    try:
        #先找到所有代币
        all_tokens = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='balancesList']/a")))
        token_num = len(list(all_tokens))
        print("zk上的所有代币数量是：", token_num)
        for i in range(1,token_num+1):
            token_button = wait.until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='balancesList']/a[{i}]//div[@class='tokenSymbol']")))
            balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='balancesList']/a[{i}]//div[@class='total']")))
            #提取代币名称和对应的金额
            token_name = token_button.text
            amount = balance_button.text
            deal_amount = amount.split('   ')  # 用3个空格分割str字符串，并保存到列表
            token_value = deal_amount[1]  # 取列表最后一个值，即 0.054992745993，字符串
            all_balance[token_name] = token_value
            print("zksync 上找到的余额是：", token_value)
        return all_balance
    except:
        print("zksync 上没有找到余额")
        return 0

#获取 zk 的金额，返回的是字典，只读取 ETH
def get_zk_eth_balance(wait):
    time_sleep(3, "进入get_zk_all_balance(), 准备去获取 zk 余额")
    try:
        #=======只找eth
        balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='balancesList']/a[1]//div[@class='total']")))
        #=======提取代币名称和对应的金额
        amount = balance_button.text
        deal_amount = amount.split('   ')  # 用3个空格分割str字符串，并保存到列表
        token_value = deal_amount[1]  # 取列表最后一个值，即 0.054992745993，字符串
        print("zksync 上找到的余额是：", token_value)
        return token_value
    except:
        print("zksync 上没有找到余额")
        return 0


#进入zk官方桥，连接钱包
def zk_connect_wallet(browser, wait):
    print("进入zk_connect_wallet()，判断是不是要连接钱包？")
    try:
        #点击钱包
        Wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='tileName' and text()='Ethereum Wallet']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", Wallet_button)
        #选择小狐狸
        MetaMask_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[@class='svelte-1799bj2' and text()='MetaMask']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", MetaMask_button)
        time_sleep(8,"已经链接小狐狸")
    except:
        print("zk_connect_wallet(),没有找到小狐狸，是否有影响？")
    time_sleep(5)

# zk 准备从L1转到L2
def zk_prepare_transfer(browser, wait,L1_ETH_value, L1_ETH_save_min):
    print("进入zk_prepare_transfer()，zk 准备从L1转到L2")
    try:
        # 转账随机值
        point = random.randint(2, 4) #最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min), point)
        while input_value > float(L1_ETH_value):
            point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min),
                                point)

        print("本次 zk 从L1转到L2的ETH金额是：", input_value)
        #点击 top up
        top_up_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-cy='account_deposit_button']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", top_up_button)

        # 选择 zksync
        zksync_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@data-cy='deposit_provider_zksync']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", zksync_button)
        time_sleep(5)
        #选择token

        #输入金额
        deposit_value_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@data-cy='amount_block_token_input']")))
        time_sleep(2)
        deposit_value_button.send_keys(str(input_value))
        time_sleep(3)

        #确认交易
        confirm_topup_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-cy='commit_transaction_button']//div[text()='Top up']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", confirm_topup_button)
    except:
        print("zk_prepare_transfer()，输入随机金额、提交订单失败")
    return input_value
    #可能弹出警告
    # try:
    #     warn_button = wait.until(
    #         EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and text()='MOVE FUNDS TO ARBITRUM']")))
    #     time_sleep(2)
    #     browser.execute_script("arguments[0].click();", warn_button)
    # except:
    #     print("ARB_prepare_transfer(),没有找到警告弹窗，是否有影响？")
    # time_sleep(5)

#等待 zk 转账到账
def wait_zk_complete(wait):
    print("进入wait_zk_complete()，等待zk 转账完成")
    for i in range(1,5):
        try:
            wait_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-cy='success_block_ok_button']")))
            break
        except:
            time_sleep(30)

#zk任务：zigzag，点击start tradding
def zk_zigzag_start_tradding(browser, wait):
    print("进入zk_zigzag_start_tradding()，点击start tradding")
    try:
        #点击Start_Tradding
        Start_Tradding = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/header/div/div/nav/ul[2]/li[5]/a")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", Start_Tradding)

    except:
        print("zk_zigzag_start_tradding,没有找到start tradding，是否有影响？")
    time_sleep(10)

#zk任务：zigzag，连接小狐狸
def zk_zigzag_connect_metamask(browser, wait):
    print("进入zk_zigzag_connect_metamask()，选择连接小狐狸")
    time_sleep(5)
    try:
        # 点击Connect Wallet，会自动跳都小狐狸
        connect_wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div[2]/button')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", connect_wallet_button)
        time_sleep(8, "zk_zigzag_connect_metamask， 已经点击了小狐狸按钮，等待下一步")
        #点击小狐狸
        MetaMask_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="WEB3_CONNECT_MODAL_ID"]/div/div/div[2]/div[1]/div/div[2]')))
        time_sleep(2, "准备点击小狐狸")
        browser.execute_script("arguments[0].click();", MetaMask_button)
    except:
        print("zk_zigzag_connect_wallet(),没有找到小狐狸，是否有影响？")
    time_sleep(10, "zk_zigzag_connect_metamask 已经点击了小狐狸，准备下一步")

#如果按钮是“connect you wallet”
def zk_whether_connect_wallet(wait):
    print("进入zk_whether_connect_wallet()，判断按钮是不是要连接钱包？")
    try:
        # 寻找 Connect Wallet
        # connect_wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='spf_btn']/button[@class='zig_btn bg_btn']/span[@style='opacity: 1;']")))
        connect_wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div[2]/button')))
        print("找到的文本是：", connect_wallet_button.text)
        return connect_wallet_button.text
    except:
        print("zk_whether_connect_wallet()出错，可能是不要连接钱包，因为没有找到按钮")

def ZK_zigzag_choose_token(browser, wait, token):
    print("进入ZK_zigzag_choose_token()，选择何种代币")

    #先点击 xiala
    down_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div/header/div/div[1]/div[1]/div/div[1]/button')))
    time_sleep(3,"先点击下拉")
    browser.execute_script("arguments[0].click();", down_button)
    time_sleep(5)
    #先点击 ALL
    all_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div/header/div/div[1]/div[1]/div/div[2]/div[3]/div[1]/div[1]/div')))
    time_sleep(3,"点击ALL")
    browser.execute_script("arguments[0].click();", all_button)
    time_sleep(5,"准备选择token")
    if token == "DAI":
        token_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='ETH/DAI']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", token_button)
        #收起列表
        # all_button = wait.until(
        #     EC.element_to_be_clickable(
        #         (By.XPATH, '//*[@id="root"]/div[2]/div/article/header/div/div/div[1]/div/div/button')))
        # time_sleep(8,"选择了DAI")
        # browser.execute_script("arguments[0].click();", all_button)
        down_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div/header/div/div[1]/div[1]/div/div[1]/button')))
        time_sleep(3,"收起列表")
        
        browser.execute_script("arguments[0].click();", down_button)
    elif token == "USDC":
        token_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='ETH/USDC']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", token_button)

        # 收起列表
        down_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div/header/div/div[1]/div[1]/div/div[1]/button')))
        time_sleep(3,"收起列表")
    elif token == "USDT":
        token_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='ETH/USDT']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", token_button)

        # 收起列表
        down_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div/header/div/div[1]/div[1]/div/div[1]/button')))
        time_sleep(3,"收起列表")
    else:
        print("ZK_zigzag_choose_token 代币输入错误，请重新输入")
    time_sleep(8, "ZK_zigzag_choose_token 运行结束")

#zk任务：获取zigzag上的ETH余额
def get_ZK_zigzag_balance(browser, wait):
    print("进入get_ZK_zigzag_balance()，获取zigzag上的ETH余额")
    # browser.refresh()
    # time_sleep(5)
    #点击图标"
    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='StyledButton-sc-19l6nhk-0 cHQUGM IconButton-sc-1wt1b69-0 AccountButton__IconButton-sc-1vtrp04-0 lagfdI JHywm']")))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", icon_button)
    
    #点击L2
    L2_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='l2']")))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", L2_button)

    #获取有多少种货币 "//ul[@class='sc-jKTccl coOmWL']/li"

    #获取有多少个代币
    # tokens = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[9]/div/div[3]/ul/li')))
    # print("总共代币数量",len(list(tokens)))

    #总是获取第一个代币余额
    try:
        print("获取第一个代币余额")
        Amount = wait.until(EC.element_to_be_clickable((By.XPATH, "//ul/li[1]//div[@class='Text-sc-d4qk04-0 jXcvep']")))                                                             
        # time_sleep(3,"waiting 收回列表")
        # browser.execute_script("arguments[0].click();", icon_button)#收回列表
        return Amount.text
    except:
        print("可能只有一个代币")
        return None

#zk任务：zigzag上进行买和卖
#zk任务：zigzag上进行买和卖
def zk_zigzag_prepare_swap(browser, wait, L2_ETH_value, buy_or_sell):
    print("进入zk_zigzag_prepare_swap()，zigzag 准备进行交易")
    if buy_or_sell == "Sell":  #卖 ETH
        ## ============ 转账随机值
        print("准备进行“Sell”")
        min = float(L2_ETH_value) * 0.8
        max = float(L2_ETH_value) * 0.95
        point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(min, max), point)
        while input_value > float(L2_ETH_value):
            point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform(min, max), point)
        print("本次 zigzag swap的随机金额是：", input_value)

        ##  ============ 点击 sell模块
        try:
            sell_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='SELL']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", sell_button)
            print("已经点击sell")
        except:
            print("dinf sell button failed")

        ## ============ 输入交易金额
        input_swap_amount = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="react-grid-layout-wrap"]/div/div/div[1]/div/aside/div/div[3]/form/div[2]/div/div/input')))
        time_sleep(2)
        input_swap_amount.send_keys(str(input_value))
        time_sleep(8)

        ## ============ 确认交易
        confirm_sell_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="SELL ETH"]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", confirm_sell_button)

        time_sleep(5)
        print(f"本次：{input_value}  去兑换 USDC")
        return f"zigzag 点击【卖】成功，本次：{input_value} 去兑换 USDC"

    elif buy_or_sell == "Buy": #进行买操作
        ## ============ 转账随机值
        print("准备进行“Buy”")
        min = float(L2_ETH_value) * 0.7
        max = float(L2_ETH_value) * 0.9
        point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(min, max), point)
        while input_value > float(L2_ETH_value):
            point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform(min, max), point)
        print("本次 zigzag swap的随机金额是：", input_value)
        time_sleep(8)
        ## ========== 点击 Buy模块
        print("准备进行“Buy”")
        buy_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[text()='BUY']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", buy_button)
        print("已经点击buy")

        ## ============ 输入交易金额
        input_swap_amount = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="react-grid-layout-wrap"]/div/div/div[1]/div/aside/div/div[3]/form/div[2]/div/div/input')))
        time_sleep(1)
        input_swap_amount.send_keys(str(input_value))
        time_sleep(5)

        # ## ======== 输入交易金额， 法二：随机25%~100%
        # # a = random.randint(6, 8)  #注意这是闭区间 卖 50、75、100%
        # a = random.randint(8, 8)  # 注意这是闭区间 100%卖出去
        # a = random.randint(7, 7)  # 注意这是闭区间 75%卖出去
        # b = (a - 4) * 25 #交易的百分比
        # input_swap_amount = wait.until(EC.element_to_be_clickable(
        #     (By.CSS_SELECTOR, f'div.tab-content > form > div.spf_range > span > span:nth-child({a})')))
        # time_sleep(2)
        # ActionChains(browser).click(input_swap_amount).perform()  # 模拟鼠标点
        #
        # time_sleep(5)
        # print(f"本次：用 {b}% USDC余额去兑换 ETH")
        # return f"zigzag 点击【买】成功，本次：用 {b}% 余额去兑换 ETH"


        ##======== 确认交易
        confirm_buy_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//button[text()="BUY ETH"]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", confirm_buy_button)

        time_sleep(3)
        print(f"本次：用 {input_value} 去兑换 ETH")
        return f"zigzag 点击【买】成功，本次：用 {input_value} 去兑换 ETH"


    else:
        print("进入zk_zigzag_prepare_swap()，输入的【关键词有错】，请输入'Buy' 或者'Sell'")
        return "进入zk_zigzag_prepare_swap()，输入的【关键词有错】，请输入'Buy' 或者'Sell'"

    # except:
    #     print("进入zk_zigzag_prepare_swap()，【输入随机金额、提交订单失败】")
    #     return None
    #可能弹出警告
    # try:
    #     warn_button = wait.until(
    #         EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and text()='MOVE FUNDS TO ARBITRUM']")))
    #     time_sleep(2)
    #     browser.execute_script("arguments[0].click();", warn_button)
    # except:
    #     print("ARB_prepare_transfer(),没有找到警告弹窗，是否有影响？")
    # time_sleep(5)


#寻找excel 里，需要做zk任务的号。原理是看是否标记为 “1”
#参数 i和j 是excel 的行起点、终点
#返回一个列表
def find_excel_zk_account(i, j):
    zk_account_buffer = []
    for i in range(i, j):
        if Do_Excel('../eth1000_OP_操作后.xlsx').read(i, 7) == "1":  # zk 在第7列
            print(f"第{i}个号需要做zk")
            zk_account_buffer.append(i)
    return zk_account_buffer


zksybc_nft_url = "https://wallet.zksync.io/account/nft/"
# 该函数用于mint zk 的NFT，其中CID_text 是要输入的CID（预先准备）
def zksync_mint_NFT(browser, wait, CID_text):
    time_sleep(5, "准备打开 zksync NFT ， 请确保切换到了对应网络")
    new_tab(browser, zksybc_nft_url)
    time_sleep(10, "正在打开 zksync NFT ")
    browser.refresh()
    time_sleep(10, "refresh ")
    switch_tab_by_handle(browser, 2, 0)  # mac下调试，切换到被撸网站

    ##=======开始链接钱包
    zk_connect_wallet(browser, wait)

    #======开始mint NFT
    #1）点击NFT
    NFTs_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='NFTs']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", NFTs_button)

    #2)点击 mint NFT 页面
    mint_NFT_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()[contains(., 'Mint NFT')]]")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", mint_NFT_button)

    #3)输入CID号码
    input_CID_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='0x hash or CID']")))
    time_sleep(2)
    input_CID_box.send_keys(str(CID_text))  #

    #4）授权
    Authorize_to_Mint_NFT = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='Authorize to Mint NFT']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", Authorize_to_Mint_NFT)
    time_sleep(5, "等待小狐狸确认授权")

    #5)小狐狸确认授权
    switch_tab_by_handle(browser, 1, 1)  # mac下调试，切换到小狐狸
    fox_confirm_sign(browser, wait)

    #6）正式Mint NFT
    switch_tab_by_handle(browser, 2, 0)  # mac下调试，切换到被撸网站
    fromal_mint_NFT_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button//div[text()='Mint NFT']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", fromal_mint_NFT_button)

    try:
        time_sleep(5, "等待Proceed_to_Mint_NFT出现")
        Proceed_to_Mint_NFT_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()[contains(.,'Proceed')]]")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", Proceed_to_Mint_NFT_button)
        time_sleep(5, "等待小狐狸确认mint")
    except:
        print("可能是不需要等待 proceed")

    # 7)小狐狸确认授权
    switch_tab_by_handle(browser, 1, 1)  # mac下调试，切换到小狐狸
    fox_confirm_sign(browser, wait)

    # 8)查看是否mint成功
    time_sleep(10, "等待查看是否成功 mint")
    switch_tab_by_handle(browser, 2, 0)  # mac下调试，切换到被撸网站
    try:
        OK_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()[contains(.,'Ok')]]")))
        print("找到ok，确实 mint 成功！")
        return "成功"
    except:
        print("没有找到ok，可能mint失败")
        return "失败"


##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ zksync的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ARB的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#进入ARB官方桥，连接钱包
def ARB_connect_wallet(browser, wait):
    print("进入ARB_connect_wallet()，判断是不是要连接钱包？")
    try:
        MetaMask_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='sc-furwcr cMFuDJ web3modal-modal-card']/div[1]//div[@class='sc-dkPtRN eCZoDi web3modal-provider-description']")))
        browser.execute_script("arguments[0].click();", MetaMask_button)
    except:
        print("ARB_connect_wallet(),没有找到小狐狸，是否有影响？")
    time_sleep(10)


# ARB 连接完小狐狸后，可能会出现 I agree
def ARB_bridge_I_agree(browser, wait):
    print("进入ARB_bridge_I_agree()，ARB 连接完小狐狸后，可能会出现 I agree")
    try:
        I_agree_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='I agree']")))
        browser.execute_script("arguments[0].click();", I_agree_button)
    except:
        print("ARB_bridge_I_agree(),没有找到I agree，是否有影响？")
    time_sleep(3)

# ARB 准备从L1转到L2
def ARB_prepare_transfer(browser, wait, L1_ETH_value, L1_ETH_save_min ):
    print("进入ARB_prepare_transfer()，ARB 准备从L1转到L2")
    try:
        # 转账随机值
        point = random.randint(2, 4) #最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min), point)
        #如果随机值太大，则再次随机
        while input_value > float(L1_ETH_value):
            point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min),
                                point)
        deposit_value = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='number']")))
        time_sleep(2)
        print("本次ARB从L1转到L2的ETH金额是：", input_value)
        deposit_value.send_keys(str(input_value))
        time_sleep(3)

        #确认交易
        deposit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and text()='Deposit']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", deposit_button)
    except:
        print("ARB_prepare_transfer()，输入随机金额、提交订单失败")
    time_sleep(5)
    #可能弹出警告
    try:
        warn_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and text()='MOVE FUNDS TO ARBITRUM']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", warn_button)
    except:
        print("ARB_prepare_transfer(),没有找到警告弹窗，是否有影响？")
    return input_value



######↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ARB的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

###### ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ L1转账到到L2的一些方案 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
#主网——》lifi——》ARB
# def mainnet_lifi_ARB(browser, wait,)

#主网——》lifi——》OP
# def mainnet_lifi_OP()

#主网——》ARB官方桥——》ARB
# def mainnet_ARB_official_ARB()

#主网——》OP官方桥——》ARB
# def mainnet_OP_official_ARB()

#主网——》zk官方桥——》ARB
# def mainnet_ZK_official_ZK()




######↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ AL1转账到到L2的一些方案↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

###### ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ lifi的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#lifi判断是不是要连接钱包
def lifi_whether_connect_wallet(browser, wait):
    print("进入lifi_whether_connect_wallet()，判断是不是要连接钱包？")
    try:
        connect_your_wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='ant-btn ant-btn-primary wallet-buttons wallet-buttons-menu-full']")))
        return connect_your_wallet_button.text
    except:
        print("lifi_whether_connect_wallet(),没有找到connect wallet按钮，是否有影响？")

#lifi 确定要连接钱包
def lifi_connect_wallet(browser, wait):
    print("进入lifi_connect_wallet，确定要连接小狐狸钱包")
    try:
        connect_your_wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='ant-btn ant-btn-primary wallet-buttons wallet-buttons-menu-full']")))
        browser.execute_script("arguments[0].click();", connect_your_wallet_button)
    except:
        print("lifi_connect_wallet(), 点击connect your wallet出错，是否影响？")

    try:
        metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='MetaMask']")))
        browser.execute_script("arguments[0].click();", metamask_button)
    except:
        print("lifi_connect_wallet(), 点击metamask出错，是否影响？")

###  【 L1 转 L2】！！！！！！
# L1通过lifi转到L2 上输入金额，准备交易，
# L1_ETH_value是 L1的 ETH余额，用于确定随机数范围。
# L1_ETH_save_min, 是要在L1上保留的ETH范围
# from to 是想要转到的链，比如 "Optimistic Ethereum"、"Arbitrum One"、"Ethereum"
# 返回本次交易的金额
def L1_lifi_L2_prepare_transfer_coin(browser, wait, L1_ETH_value, L1_ETH_save_min, from_source, to_destination):
    print("我已经进入L1_lifi_L2_prepare_transfer_coin，L1 转到 L2")
    for i in range(1,4):
        try:#因为有可能刚进去lifi，显示不出链，则刷新重试
            # From——链
            From_chain_switch = browser.find_element_by_css_selector("#rc_select_0")
            From_chain_switch.send_keys("BSC")#随意输入的，为了保证能够选中下拉框
            From_chain_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//body/div[2]//div[@data-label='{from_source}']")))  # Ethereum   #Polygon Optimistic Ethereum
            browser.execute_script("arguments[0].click();", From_chain_button)
            time_sleep(2)
            break
        except:
            print("lifi_prepare_transfer_coin(), 没有找到链，刷新重试")
            browser.refresh()
            time_sleep(8)
            continue
    # From——币
    From_coin_switch = browser.find_element_by_css_selector("#rc_select_1")
    From_coin_switch.send_keys("ETH")
    From_coin_switch.send_keys(Keys.ENTER)
    time_sleep(2)

    # To——链
    To_chain_switch = browser.find_element_by_css_selector("#rc_select_2")
    To_chain_switch.send_keys("BSC")
    # 鼠标滚轮先滚动
    To_chain_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//body/div[4]//div[@data-label='Fantom']")))
    loc = To_chain_button.location_once_scrolled_into_view
    To_chain_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, f"//body/div[4]//div[@data-label='{to_destination}']")))  # Ethereum   #Polygon  #Arbitrum One
    browser.execute_script("arguments[0].click();", To_chain_button)

    time_sleep(2)

    # To——币
    To_coin_switch = browser.find_element_by_css_selector("#rc_select_3")
    To_coin_switch.send_keys("ETH")
    To_coin_switch.send_keys(Keys.ENTER)
    time_sleep(2)

    # 三、选择要换多少USDC钱
    BTH_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='ant-input ant-input-borderless']")))
    BTH_input_button.send_keys(Keys.BACKSPACE)
    time_sleep(1)
    BTH_input_button.send_keys(Keys.BACKSPACE)
    time_sleep(1)
    BTH_input_button.send_keys(Keys.BACKSPACE)

    #随机金额
    point = random.randint(2,4)  #小数点最起码要有2位，因为L1的ETH金额一般是两位小数以上
    input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min), point)
    while input_value > float(L1_ETH_value):
        point = random.randint(2, 4)  # 小数点最起码要有2位，因为L1的ETH金额一般是两位小数以上
        input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min),
                            point)
    print("随机转账的ETH金额是：",input_value)
    BTH_input_button.send_keys(str(input_value))  # 随机金额
    time_sleep(25)
    return input_value #返回本次交易金额


###  【 L2 转 L2】！！！！！！
# L2通过lifi转到L2 上输入金额，准备交易，
# L2_ETH_value是 L1的 ETH余额，用于确定随机数范围。
# rate：要转多少比例
# from, to 是想要转到的链，比如 "Optimistic Ethereum"、"Arbitrum One"、"Ethereum"
# 默认是转 ETH
# 返回本次交易的金额
def L2_lifi_L2_prepare_transfer_coin(browser, wait, L2_ETH_value, from_source, to_destination, from_coin="ETH", to_coin="ETH"):
    print("我已经进入L2_lifi_L2_prepare_transfer_coin，输入交易金额，准备交易")
    for i in range(1,4):
        try:#因为有可能刚进去lifi，显示不出链，则刷新重试
            # From——链

            From_chain_switch = browser.find_element_by_css_selector("#rc_select_0")
            time_sleep(3, '倒计时结束后，发送字符')
            From_chain_switch.send_keys("BSC")#随意输入的，为了保证能够选中下拉框
            time.sleep(5)
            # time_sleep(15, '倒计时结束后，模拟点击')
            # ActionChains(browser).click(From_chain_switch).perform()  # 模拟鼠标点击

            #用 Ethereum 或 Fantom 来定位，之后滚动滚轮
            Locate = wait.until(EC.element_to_be_clickable((By.XPATH, "//body/div[2]//div[@data-label='Fantom']")))  # Ethereum   #Polygon Optimistic Ethereum
            time_sleep(3, "倒计时结束后，滚动鼠标滚轮，防止选不中")
            loc= Locate.location_once_scrolled_into_view

            #开始寻找源路径
            From_chain_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//body/div[2]//div[@data-label='{from_source}']")))  # Ethereum   #Polygon Optimistic Ethereum
            time_sleep(2, "开始寻找源路径")
            browser.execute_script("arguments[0].click();", From_chain_button)
            break
        except:
            print("L2_lifi_L2_prepare_transfer_coin(), 没有找到 From 链，可能是没有滚动鼠标滚轮，刷新试一下")
            browser.refresh()
            time_sleep(8)
            continue
    # From——币
    From_coin_switch = browser.find_element_by_css_selector("#rc_select_1")
    time_sleep(2)
    From_coin_switch.send_keys(from_coin)
    time_sleep(2)
    From_coin_switch.send_keys(Keys.ENTER)
    time_sleep(2)

    # To——链
    To_chain_switch = browser.find_element_by_css_selector("#rc_select_2")
    time_sleep(3, '倒计时结束后，发送字符')
    To_chain_switch.send_keys("BSC")  #随意发送字符串
    # 鼠标滚轮先滚动
    To_chain_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//body/div[4]//div[@data-label='Fantom']")))
    time_sleep(3, "倒计时结束后，滚动鼠标滚轮，防止选不中")
    loc = To_chain_button.location_once_scrolled_into_view
    To_chain_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, f"//body/div[4]//div[@data-label='{to_destination}']")))  # Ethereum   #Polygon  #Arbitrum One
    time_sleep(2)
    browser.execute_script("arguments[0].click();", To_chain_button)

    time_sleep(2)

    # To——币
    To_coin_switch = browser.find_element_by_css_selector("#rc_select_3")
    time_sleep(2)
    To_coin_switch.send_keys(to_coin)
    time_sleep(2)
    To_coin_switch.send_keys(Keys.ENTER)
    time_sleep(2)

    # 三、选择要换多少钱
    #先删除默认值
    BTH_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='ant-input ant-input-borderless']")))
    BTH_input_button.send_keys(Keys.BACKSPACE)
    time_sleep(1)
    BTH_input_button.send_keys(Keys.BACKSPACE)
    time_sleep(1)
    BTH_input_button.send_keys(Keys.BACKSPACE)

    #随机金额
    point = random.randint(2,4)  #小数点最起码要有2位，因为L1的ETH金额一般是两位小数以上
    input_value = round(random.uniform((L2_ETH_value) * (0.8), (L2_ETH_value) * (0.95)), point)
    while input_value > float(L2_ETH_value):
        point = random.randint(2, 4)  # 小数点最起码要有2位，因为L1的ETH金额一般是两位小数以上
        input_value = round(random.uniform((L2_ETH_value) * (0.8), (L2_ETH_value) * (0.95)), point)

    print("本次随机转账的ETH金额是：",input_value)
    BTH_input_button.send_keys(str(input_value))  # 随机金额
    time_sleep(15)
    return input_value #返回本次交易金额

#等待 lifi 转账到账
def wait_lifi_complete(wait):
    print("进入wait_lifi_complete()，等待lifi 转账完成")
    for i in range(1,5):
        try:
            wait_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//strong[text()='Swap Successful!']")))
            break
        except:
            print("lifi 交易未结束，继续等待")
            time_sleep(10)

# 计算 lifi 上的交易费
def lifi_gas_fee(browser, wait, input_value):
    print("进入lifi_gas_fee()，估计 gas fee")
    get_prediction_value = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                  "//div[@style='display: flex; flex-direction: row; overflow-x: scroll;']/div[1]/div[@class='selected']//b")))
    browser.execute_script("arguments[0].click();", get_prediction_value)
    # a = get_prediction_value.text
    a = (get_prediction_value.text)[0:-4]  # 截取字符串长度
    gas_fee = (input_value - float(a)) / input_value
    print(f"预计到账：{a}, 预计gas fee：{gas_fee}")
    return gas_fee

#lifi点击switch进行网络切换
def lifi_switch_net(browser, wait):
    print("我已经进入lifi_switch_net，进行网络切换")
    switch_network_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    browser.execute_script("arguments[0].click();", switch_network_button)

#lifi点击swap进行交易
def lifi_swap(browser, wait):
    print("我已经进入lifi_swap，正式进行交易")
    # try:
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Swap']")))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", swap_button)
    # except:
    # print("lifi_swap()，点击swap失败")
    # try:

    # 可能会有gas fee 过高的警告
    try:
        ok_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='OK']")))
        time_sleep(3)
        browser.execute_script("arguments[0].click();", ok_button)
    except:
        print("没有出现gas fee 过高的警告")
    time_sleep(10, "倒计时结束后点击【Start Cross Chain Swap】")
    star_cross_chain_swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Start Cross Chain Swap']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", star_cross_chain_swap_button)
    time_sleep(10, "lifi_swap() 等待交易进行")
    # except:
    #     print("lifi_swap()，点击Start Cross Chain Swap失败")


#获取lifi按钮的文本，判断是连接网络还是swap
def lifi_switch_net_or_swap(browser, wait):
    print("我已经进入lifi_switch_net_or_swap，判断是连接网络还是swap")
    try:
        txt_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//form/span/div[2]/button")))
        print("lifi_switch_net_or_swap()找到的文字是：",txt_button.text)
        return txt_button.text
    except:
        print("lifi_switch_net_or_swap()，获取lifi按钮的文本时出错了")


##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ lifi 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ matcha 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

# 实时获取 op 上eth 的金额多少
def get_OP_ETH_by_matcha_and_prepare_from_to_token(browser, wait, from_token, to_token):
    print("我已进入get_OP_ETH_by_matcha_and_prepare_from_to_token，实时获取 OP 上 ETH 的金额多少")

    # ================= 一、先选择 to，兑换成什么货币
    # 点击下拉框
    receive_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[6]/div[3]/div[1]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", receive_button)
    time_sleep(2, " 我已经点击You receive下拉框")

    # 选择何种币
    try:
        choose_to_token = wait.until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='ChangeTokenActionSheet__TokenSymbolLabel-sc-1pl3sgj-12 kiChfM' and text()='{to_token}']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", choose_to_token)
        time_sleep(2, f"我已经选择to {to_token}")
    except:
        print(f"可能是之前选择了{to_token}")

    # ============= 二、再选择 from，源
    # 点击下拉框
    choose_token = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[4]/div/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", choose_token)
    time_sleep(2, " 我已经点击You Pay 下拉框")

    try:
        # 选择何种币
        choose_from_token = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                   f"//div[@class='ChangeTokenActionSheet__TokenSymbolLabel-sc-1pl3sgj-12 kiChfM' and text()='{from_token}']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", choose_from_token)
        time_sleep(2, f"我已经选择from {from_token}")
    except:
        print(f"可能是之前选择了{from_token}")
    #=============== 有可能出现 I understand
    try:
        I_understand_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH,
             "//button[text()='I understand']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", I_understand_button)
        print("get_OP_ETH_by_matcha_and_prepare_from_to_token, 已经点击 I understand")
    except:
        print("I understand 点击失败，可能没出现")

    #==============获取余额
    try:
        balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[3]/button')))
        print("matcha找到的按钮是", balance_button.text) #返回的是 " MAX 0.0332 ETH"
        a = balance_button.text
        b = a.split()
        c = b[1]
        print("分割后，最终找到的ETH金额是:", c)
        return float(c)
    except:
        print("get_OP_ETH_by_matcha_and_prepare_from_to_token，没有找到 OP 上ETH 的金额")

# 实时获取 op 上所有代币的金额多少，返回 from_token 以及最大金额
def get_OP_token_balance_by_matcha(browser, wait):
    print("我已进入get_OP_token_balance_by_matcha，实时获取 OP 上 ETH 的金额多少")

    # ================= 一、点击头像，查看有什么代币
    # 点击下拉框
    icon_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div[1]/div/nav/div[4]/div[2]/div[1]/button[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", icon_button)
    time_sleep(2, " 我已经点击头像下拉框")

    # ==============获取最大余额
    try:
        max_balance_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div[1]/div/nav/div[4]/div[2]/div[2]/div/div[2]/div/div[1]/div[1]/p')))
        print("matcha找到的按钮是", max_balance_button.text)  # 返回的是 " MAX 0.0332 ETH"
        a = max_balance_button.text
        b = a.split()
        max_balance = b[0]
        max_token = b[1]
        print(f"分割后，最终找到的最大代币是：{max_token}, 金额是：{max_balance}")

        # 点回头像
        browser.execute_script("arguments[0].click();", icon_button)
        return max_token, float(max_balance)

    except:
        print("get_OP_token_balance_by_matcha，没有找到 OP 上的最大金额和代币")


#判断抹茶要换网络
def matcha_change_net(browser, wait):
    print("我已经进入matcha_change_net()，抹茶要换网络")
    network_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//span[@class='SelectDropdown__StyledListboxButton-w3r5gj-3 kYrFzD']")))
    network_button.click()
    print("matcha_change_net()已点击下拉，展开网络")
    OP_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-label='Optimism']")))
    time_sleep(3)
    OP_button.click()
    print("matcha_change_net()，已选择OP")

#判断抹茶是不是要换网络
def matcha_whether_change_net(browser, wait):
    print("我已经进入matcha_whether_change_net()，判断抹茶是不是要换网络")
    # 一、Trade
    # Trade_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, "//a[@class='HeaderLink-qfrd7f-0 HeaderLinkWithHover-sc-1dg2h99-0 krXwdd zbhps'][1]")))
    Trade_button = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Trade")))
    browser.execute_script("arguments[0].click();", Trade_button)
    print("matcha_whether_change_net(), 我已经点击Trade")
    time_sleep(20)
    # 二、来到判断页面
    network_button = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//span[@class='SelectDropdown__StyledListboxButton-w3r5gj-3 kYrFzD']")))
    print("matcha_whether_change_net()找到的网络是：",network_button.text)
    return network_button.text
    # browser.execute_script("arguments[0].click();", network_button)

#抹茶上点击连接钱包
def matcha_connect_wallet(browser, wait):
    print("我已经进入matcha_connect_wallet，抹茶上点击连接钱包")
    connect_metamask_button = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[@class='ConnectWalletMenu__StyledWalletButton-sc-15fdpjx-2 fGcjf'][1]")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", connect_metamask_button)

#抹茶，判断是否转账成功
def matcha_whether_transfer_success(browser, wait):
    success_button = wait.until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div/div[2]/div[2]')))
    return success_button.text

#抹茶，准备换钱，输入货币、金额
# from_token，可选"Ethereum"， "USD Coin"，"sUSD"
# 法一. 以前的方法
def matcha_input_coin_amount_old(browser, wait, L2_balance, from_source, to_token):
    print("我已经进入matcha_input_coin_amount(), 准备换钱，输入货币、金额")

    # ================= 一、先选择 to，兑换成什么货币。防止选不上Ethereum
    # 点击下拉框
    receive_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                            '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[6]/div[3]/div[1]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", receive_button)
    time_sleep(2, " 我已经点击You receive下拉框")

    # 选择何种币
    try:
        choose_to_token = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                 f"//div[@class='ChangeTokenActionSheet__TokenSymbolLabel-sc-1pl3sgj-12 kiChfM' and text()='{to_token}']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", choose_to_token)
        time_sleep(2, f"我已经选择to {to_token}")
    except:
        print(f"可能是之前选择了{to_token}")

    # if to_token != "Ethereum":
    #     # =============== 有可能出现 I understand
    try:
        I_understand_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH,
                "//button[text()='I understand']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", I_understand_button)
        print("get_OP_ETH_by_matcha_and_prepare_from_to_token, 已经点击 I understand")
    except:
        print("to 时，尝试点击 I understand 点击失败，可能没出现")

    # ============= 二、再选择 from，源
    # 点击下拉框
    choose_token = wait.until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[4]/div/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", choose_token)
    time_sleep(2, " 我已经点击You Pay 下拉框")

    try:
        # 选择何种币
        choose_from_token = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                   f"//div[@class='ChangeTokenActionSheet__TokenSymbolLabel-sc-1pl3sgj-12 kiChfM' and text()='{from_token}']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", choose_from_token)
        time_sleep(2, f"我已经选择from {from_token}")
    except:
        print(f"可能是之前选择了{from_token}")

    # =============== 有可能出现 I understand
    try:
        I_understand_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH,
             "//button[text()='I understand']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", I_understand_button)
        print("==matcha_input_coin_amount, 已经点击 I understand")
    except:
        print("from时，尝试点击 I understand 点击失败，可能没出现")
    #================ 三、选择 from 要换多少
    try:
        pay_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div[4]/div[2]/div/input')))
        print("I find input")   
    except:
        print("Find intput botton wrong!!")

    point = random.randint(3, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    input_value = round(random.uniform(L2_balance * 0.7, L2_balance * 0.8), point)  #随机金额
    try_times = 0
    while float(L2_balance) - input_value < 0.005: #如果会使余额小于 0.005
        point = random.randint(3, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
        input_value = round(random.uniform(L2_balance * 0.3, L2_balance * 0.7), point)  # 随机金额
        try_times = try_times + 1
        if try_times == 100:
            print("余额不足，直接报错")
            return 0  # 说明余额不做，直接报错即可
    
    print(f"随机转账的{from_token}金额是：{input_value}，将来预估余额是：{float(L2_balance) - input_value}")
    pay_input_button.send_keys(str(input_value))
    time_sleep(2, f"matcha_input_coin_amount(), 已经输入要换的{from_token}")

    #================ 四、选择review order
    review_order_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text() = 'Review Order']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", review_order_button)
    print("matcha_input_coin_amount(), 已经点击review order")
    time_sleep(3)

    # =============== 五、可能会让你是否同意让其访问某种代币，比如 USDC、DAI
    try:
        confirm_token_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div/div[3]/button')))
        if "Approve" in confirm_token_button.text:
            print("可能会让你是否同意授权 Approve")
            browser.execute_script("arguments[0].click();", confirm_token_button)
            time_sleep(3)
            switch_tab_by_handle(browser, 1, 1)  # 切到小狐狸
            fox_confirm_L2_swap(browser, wait)
            switch_tab_by_handle(browser, 2, 0)  # 切到 matcha
    except:
        print("可能不需要授权")

    # ================ 六、选择 place order
    try:
        place_order_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[text() = 'Place Order']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", place_order_button)
        print("matcha_input_coin_amount(), 已经点击 place order")
        time_sleep(5)
        return f" OP matcha 完成，从 {from_token} 换 {to_token} 金额 {input_value}"
    except:
        print("尝试 place order 失败了")

# 法二. 用搜索的方法
def matcha_input_coin_amount_new(browser, wait, L2_balance, from_source, to_source):
    print("我已经进入matcha_input_coin_amount(), 准备换钱，输入货币、金额")

    #=====================先点击to的下拉列表
    to_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//button/div[text()[contains(.,'ETH')]]")))
    time_sleep(3, "to_list button found")
    browser.execute_script("arguments[0].click();", to_list)

    #输入框
    to_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='ChangeTokenActionSheet__StyledInput-sc-1pl3sgj-15 hIbkAm']")))
    time_sleep(3)
    to_button.send_keys(to_source)
    
    time_sleep(3,"再确认选择该代币, 第一个")
    first_token_xpath = "//div[@class='ChangeTokenActionSheet__AssetsContainer-sc-1pl3sgj-3 gUZwQm']/div[1]/div"
    select_first_token_button = wait.until(EC.element_to_be_clickable((By.XPATH, first_token_xpath)))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", select_first_token_button)
    time_sleep(6,"已经选择代币, 等待响应")

    try:
        print("尝试找 I understand")
        I_understand_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='I understand']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", I_understand_button)
        print("已经点击 I understand")
    except:
        print("to 时，尝试点击 I understand 点击失败，可能没出现")
    
    #===================再点击from的下拉列表
    from_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//button/div[text()[contains(.,'Choose')]]")))
    time_sleep(3, "from_list button found")
    browser.execute_script("arguments[0].click();", from_list)

    #输入框
    from_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class='ChangeTokenActionSheet__StyledInput-sc-1pl3sgj-15 hIbkAm']")))
    time_sleep(3)
    from_button.send_keys(from_source)
    
    time_sleep(3,"再确认选择该代币, 第一个")
    first_token_xpath = "//div[@class='ChangeTokenActionSheet__AssetsContainer-sc-1pl3sgj-3 gUZwQm']/div[1]/div"
    select_first_token_button = wait.until(EC.element_to_be_clickable((By.XPATH, first_token_xpath)))
    time_sleep(3)
    browser.execute_script("arguments[0].click();", select_first_token_button)
    time_sleep(5,"等待响应")

    try:
        print("尝试找 I understand")
        I_understand_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='I understand']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", I_understand_button)
        print("已经点击 I understand")
    except:
        print("from 时，尝试点击 I understand 点击失败，可能没出现")

    #================ 三、选择 from 要换多少
    try:
        pay_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[4]/div[2]/div[1]//input[@class='Input__SwapInput-jrcd0l-4 dOXKmv']")))
        print("找到了输入金额框")   
    except:
        print("没有找到输入金额的框Find intput botton wrong!!")

    point = random.randint(3, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    input_value = round(random.uniform(L2_balance * 0.7, L2_balance * 0.8), point)  #随机金额
    try_times = 0
    while float(L2_balance) - input_value < 0.005: #如果会使余额小于 0.005
        point = random.randint(3, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
        input_value = round(random.uniform(L2_balance * 0.3, L2_balance * 0.7), point)  # 随机金额
        try_times = try_times + 1
        if try_times == 100:
            print("余额不足，直接报错")
            return 0  # 说明余额不做，直接报错即可
    
    print(f"随机转账的{from_source}金额是：{input_value}，将来预估余额是：{float(L2_balance) - input_value}")
    pay_input_button.send_keys(str(input_value))
    time_sleep(5, f"matcha_input_coin_amount(), 已经输入要换的{from_source}")

    #================ 四、选择review order
    review_order_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text() = 'Review Order']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", review_order_button)
    print("matcha_input_coin_amount(), 已经点击review order")
    time_sleep(5)

    # =============== 五、可能会让你是否同意让其访问某种代币，比如 USDC、DAI
    try:
        confirm_token_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="trading-page-container"]/div[2]/div/div/aside/div[1]/div/div/div/div[3]/button')))
        if "Approve" in confirm_token_button.text:
            print("可能会让你是否同意授权 Approve")
            browser.execute_script("arguments[0].click();", confirm_token_button)
            time_sleep(3)
            switch_tab_by_handle(browser, 1, 1)  # 切到小狐狸
            fox_confirm_L2_swap(browser, wait)
            switch_tab_by_handle(browser, 2, 0)  # 切到 matcha
    except:
        print("可能不需要授权")

    # ================ 六、选择 place order
    try:
        place_order_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[text() = 'Place Order']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", place_order_button)
        print("matcha_input_coin_amount(), 已经点击 place order")
        time_sleep(5)
        return f" OP matcha 完成，从 {from_source} 换 {to_source} 金额 {input_value}"
    except:
        print("尝试 place order 失败了")
   

# 
def matcha_formal_change_coin(browser, wait):
    print("进入matcha_formal_change_coin，正式下单")
    place_order_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='CheckoutCommonLayout__BottomBarContainer-jgkqw7-6 hMlefr']/button")))
    time_sleep(5)
    place_order_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='CheckoutCommonLayout__BottomBarContainer-jgkqw7-6 hMlefr']/button")))
    browser.execute_script("arguments[0].click();", place_order_button)
    # place_order_button.click()
    print("已经点击")

#判断抹茶是否连接了小狐狸钱包，思路是看小狐狸的图片是否存在
def matcha_fox_icon_exit(browser, wait):
    print("进入matcha_fox_icon_exit，判断要不要连接小狐狸")
    a = 0
    for i in range(1,4):
        try:
            browser.find_element(By.XPATH, "//img[@src='data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMjAiIGhlaWdodD0iMjAiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+PGcgY2xpcC1wYXRoPSJ1cmwoI2NsaXAwKSI+PHBhdGggZD0iTTE5LjIxIDE2LjU3MmwtMS4yNi4yOTgtLjI2LjE5Mi0xMS4xMTUtMS40NDhIMGwyLjM0OC01LjYyOCAzLjgxNy4wMjEuOTI0IDEuNjI2IDMuMDQ5IDMuMDY2LS4zOTUtMS45MSAyLjcyLjQ2OSA1LjIyNiAzLjgwNC00LjEzLTMuODA0LS42NzgtMS44MTctMi4zODUtNS42MjhzLTEuODM0LTMuMTQ0LTEuODcxLTMuMThjLjA1Mi4wMDcgNC4yNTYtLjAxNCA0LjI1Ni0uMDE0bC43NjggMy4wMjQgMS42OTIgNS43OTguNDYyIDEuNTY4IDEuNjQgMS44MzktMS4zMzQtMS45NzMuODcyLS4zNTUuNDE4IDEuNzg4LjAzNy41NCAxLjc3NCAxLjcyNHoiIGZpbGw9IiNGNThDMjUiLz48cGF0aCBkPSJNNi4xNjUgMTAuMDA3bC0zLjgxNy0uMDIxLS45MDktLjkzLjU3NC0uMzctMS4wMjktLjUzMS44OC0uMjI4LTEuMjA4LS43MjMuNzIzLS4yOTkuMTgtMS4wNzggMi4wMDQtMy41NjNMNi4yNzcgMGwyLjM0IDIuNjMzIDEuODc5IDMuMTcyLTQuMzMxIDQuMjAyeiIgZmlsbD0iIzgxNDkxRiIvPjxwYXRoIGQ9Ik0wIDE1LjYxNGg2LjU4MmwxLjc5NyAyLjI4NXMtNC40NzMgMS4yODUtNi43MTcgMS45MzhBMzk1Ni44MiAzOTU2LjgyIDAgMDEwIDE1LjYxNHoiIGZpbGw9IiNFNDgwMjciLz48cGF0aCBkPSJNMTcuNjk3IDUuOTY5bC41MjkuNTgyLS4zODguNDQuNTM3LjM2OS0uMzg4LjQ5LjA5Ljc3My0uNjY0Ljc1Mi0zLjc2NC0zLjczMyA2LTQuNzM0LTEuOTUyIDUuMDZ6IiBmaWxsPSIjODE0OTFGIi8+PHBhdGggZD0iTTE3LjY5IDE3LjA2MmwuMDIyIDEuNDEyLTIuMTQgMS4wMy03LjItMS41OThzLS4wNTMuMDM2IDAgMGMzLjAxOS0uMjc2IDkuMzE4LS44NDQgOS4zMTgtLjg0NHoiIGZpbGw9IiNEOUM2QjciLz48cGF0aCBkPSJNMTkuNjUuOTA4bC02LjAwMSA0LjczNC0uNzY4LTMuMDIzIDYuNzY4LTEuNzF6IiBmaWxsPSIjRTQ4MDI3Ii8+PHBhdGggZD0iTTEzLjY0OSA1LjY0MmwzLjc1NyAzLjczMy0uMTU3IDEuOTM4LTEuOTA4LjEyOC0xLjY5Mi01Ljc5OXoiIGZpbGw9IiNFNTdGMjQiLz48cGF0aCBkPSJNOC4zNzEgMTcuODk5bDcuMjAxIDEuNTk3IDIuMTQtMS4wMyAxLjQ2OC0uMTItMS4xMTggMS4zOTEtMy4wMzQuMjYzLTYuNjU3LTIuMTAxeiIgZmlsbD0iI0M3QjdBQiIvPjxwYXRoIGQ9Ik04LjM3MiAxNy45bC0xLjc5LTIuMjg2IDMuNTU2LS45MTZMOC4zNzIgMTcuOXoiIGZpbGw9IiNENzZGMjEiLz48cGF0aCBkPSJNMTkuMTggMTguMzQ2Yy0uNDkyLjA0My0uOTc2LjA4NS0xLjQ2OC4xMmwtLjAyMy0xLjQxMi4yNjEtLjE5MSAxLjI2LS4yOTFzLjQyNS4yODQuNzkuNDgzYy0uMjkuNDU0LS41NTkuODcyLS44MiAxLjI5MXoiIGZpbGw9IiMyQTI2MjYiLz48cGF0aCBkPSJNMTUuMzQxIDExLjQ0bDEuOTA4LS4xMjctLjI2IDEuMjA3LS43MjQtLjgyNC0uMTU2IDEuMTc4IDEuMzM0IDEuOTczLTEuNjQtMS44NDUtLjQ2Mi0xLjU2MXoiIGZpbGw9IiNENjZGMjEiLz48cGF0aCBkPSJNNi4xNjUgMTAuMDA3bDQuMzMtNC4yMDIgMi4zODYgNS42MzYtNS43OTIuMTkxLS45MjQtMS42MjV6IiBmaWxsPSIjRTQ3RTI1Ii8+PHBhdGggZD0iTTguMzcxIDE3LjlsMS43NjctMy4yMDEgNy41NTEgMi4zNjMtOS4zMTguODM3eiIgZmlsbD0iI0U0N0YyNiIvPjxwYXRoIGQ9Ik03LjA5IDExLjYzMmw1Ljc5MS0uMTkxLS40MTcgMS44MTctLjk3Ny0xLjQzNC0xLjc0NC45NjUtMi42NTQtMS4xNTd6IiBmaWxsPSIjRDY2RjIxIi8+PHBhdGggZD0iTTEyLjQ2NCAxMy4yNThsLjQxNy0xLjgxNy42NzggMS44MTcgNC4xMyAzLjgwNHMtNC4wMzMtMi45MzItNS4yMjUtMy44MDR6IiBmaWxsPSIjRTQ3RjI1Ii8+PHBhdGggZD0iTTEyLjQ2NCAxMy4yNThsLTIuNzItLjQ2OSAxLjc0My0uOTY1Ljk3NyAxLjQzNHoiIGZpbGw9IiMzMjQwNEUiLz48cGF0aCBkPSJNOS43NDMgMTIuNzlsLjM5NSAxLjkwOS0zLjA0OS0zLjA2NyAyLjY1NCAxLjE1N3oiIGZpbGw9IiNFNTdGMjUiLz48cGF0aCBkPSJNMTYuMTA5IDEyLjg3NGwuMTU3LTEuMTc4LjcyMy44MjMtLjg4LjM1NXoiIGZpbGw9IiMzODQyNEQiLz48cGF0aCBkPSJNMTYuOTg5IDEyLjUybC4yNi0xLjIwNy4yNzcgMS4yODUtLjEyIDEuNzEtLjQxNy0xLjc4OHoiIGZpbGw9IiNFNTdGMjQiLz48L2c+PGRlZnM+PGNsaXBQYXRoIGlkPSJjbGlwMCI+PHBhdGggZmlsbD0iI2ZmZiIgZD0iTTAgMGgyMHYyMEgweiIvPjwvY2xpcFBhdGg+PC9kZWZzPjwvc3ZnPg==']")
            a = 1     #说明有小狐狸图标
            break
        except:
            browser.refresh()
            time_sleep(3)
            continue
    return a
##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ matcha 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ OP 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
def OP_prepare_transfer(browser, wait,L1_ETH_value, L1_ETH_save_min ):
    print("进入OP_prepare_transfer()，OP 准备从L1转到L2")
    try:
        # 确认交易
        Agree_to_terms = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Agree to terms']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", Agree_to_terms)
    except:
        print("没有找到 【Agree to terms】")
    try:
        # 转账随机值
        point = random.randint(3, 4) #最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min), point)
        while input_value > float(L1_ETH_value):
            point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform((L1_ETH_value - L1_ETH_save_min) * 0.9, L1_ETH_value - L1_ETH_save_min), point)

        deposit_value = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='number']")))
        time_sleep(2)
        print("本次OP从L1转到L2的ETH金额是：", input_value)
        deposit_value.send_keys(str(input_value))
        time_sleep(3)
    except:
        print("OP_prepare_transfer()，输入随机金额失败")

    try:
        #确认交易
        deposit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Deposit']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", deposit_button)
    except:
        print("OP_prepare_transfer()，提交订单失败")
    time_sleep(5)
    #可能弹出警告
    try:
        warn_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[3]/div/div[2]/button")))
        time_sleep(8)
        browser.execute_script("arguments[0].click();", warn_button)
    except:
        print("OP_prepare_transfer(),没有找到警告弹窗，是否有影响？")
    return input_value

#等待 OP 交易完成
def wait_OP_transfer_ok(wait):
    print("进入wait_OP_transfer_ok，等待 OP 转账完成")
    for i in range(1, 15):
        try:
            deposit_complete = wait.until(EC.element_to_be_clickable((By.XPATH, "//h1[@class='transaction-modal__complete-header']")))
            break
        except:
            time_sleep(6)
            continue
##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ OP 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ orb 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
#在小狐狸页面，获取orb的gas fee，因为gas fee很低，暂时不做这个功能。
# def fox_get_orb_gas_fee(browser, wait, input_value):
#     print("进入orb_gas_fee()，小狐狸估计 gas fee")
#     get_prediction_value = wait.until(EC.element_to_be_clickable((By.XPATH,
#                                                                   "//div[@style='display: flex; flex-direction: row; overflow-x: scroll;']/div[1]/div[@class='selected']//b")))
#     browser.execute_script("arguments[0].click();", get_prediction_value)
#     # a = get_prediction_value.text
#     a = (get_prediction_value.text)[0:-4]  # 截取字符串长度
#     gas_fee = (input_value - float(a)) / input_value
#     print(f"预计到账：{a}, 预计gas fee：{gas_fee}")
#     return gas_fee


#orb 判断是否连接钱包
def orb_whether_connect_wallet(wait):
    print("我已进入orb_whether_connect_wallet，判断 orb 是否要连接钱包")
    try:
        orb_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/span/span')))
        print("orb 找到的按钮是", orb_connect_wallet.text)
        return orb_connect_wallet.text
    except:
        print("没有找到 orb 上的 connect wallet 按钮")
        return None

#orb 连接狐狸钱包
def orb_connect_wallet(browser, wait):
    print("我已进入orb_connect_wallet()，orb 确实要连接钱包")
    try:
        # 点击连接钱包
        orb_connect_wallet =  wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/span/span')))
        time_sleep(2, "已经找到按钮")
        browser.execute_script("arguments[0].click();", orb_connect_wallet)

        # 选择小狐狸
        metamask_wallet = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[2]/div/div[2]/span')))
        time_sleep(2, "已经找到小狐狸")
        browser.execute_script("arguments[0].click();", metamask_wallet)
        print("orb_connect_wallet，连接钱包，点击小狐狸结束")
    except:
        print("orb_connect_wallet()，orb 连接钱包失败")
        return None

# orb 选择from 和 to
def L2_orb_L2_prepare_transfer_coin(browser, wait, from_source_value, to_source_value):
    print("我已经进入L2_orb_L2_prepare_transfer_coin，输入交易金额，准备交易")
    

    # 三、选择要换多少钱
    BTH_input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/div[2]/div[2]/div[2]/input')))

    #========================法一：随机金额
    # point = random.randint(3, 4)  # 小数点最起码要有3位，不然会被向上取整，导致 orb 无法交易
    # input_value = round(random.uniform((from_source_value) * (0.45), (from_source_value) * (0.5)), point)
    # while input_value > float(from_source_value):
    #     point = random.randint(2, 4)  # 小数点最起码要有2位，因为L1的ETH金额一般是两位小数以上
    #     input_value = round(random.uniform((from_source_value) * (0.85), (from_source_value) * (0.9)), point)
    # print("本次随机转账的ETH金额是：", input_value)
    
    # BTH_input_button.send_keys(str(input_value))  # 随机金额
    # time_sleep(5,"已经输入随机金额")
    #============================================

    #
    # #=========================法二：最大化
    # max_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="aliveRouter"]/div/div[2]/div[1]/div[2]/div[2]/button/span')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", max_button)
    # input_value = "全部"
    # time_sleep(5,"已经输入max金额")
    ## ============================================

    #========================法三：对半转账
    point = random.randint(3, 4)  # 小数点最起码要有3位，不然会被向上取整，导致 orb 无法交易
    average_value = (from_source_value + to_source_value)/2 
    input_differ_value = from_source_value - average_value # 从源转出去多少钱
    input_value = round(random.uniform((input_differ_value) * (0.8), (input_differ_value) * (0.9)), point)
    
    print("本次随机转账的ETH金额是：", input_value)
    
    BTH_input_button.send_keys(str(input_value))  # 随机金额
    time_sleep(5,"已经输入随机金额")
    
    #点击Send
    send_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//span[text()=' SEND ']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", send_button)
    time_sleep(8, "已经点击 send ，等待 orb 响应")
    # 获取orb gas fee
    orb_gas_fee = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[4]/div/div[2]/div[2]/span')))

    # 点击confirm and send
    confirm_and_send = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='CONFIRM AND SEND']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", confirm_and_send)
    time_sleep(10, "已经点击confirm and send，等待 orb 响应")

    return f" 本次交易额是：{input_value}, orb 显示 withholding fee 手续费是：{orb_gas_fee.text}"



# 从 orb 上获取生态的实时余额，返回的是浮点数
# Optimism、Arbitrum、zkSync
def get_L2_balance_from_orb(browser, wait, from_source, to_source):
    print("我已经进入get_L2_balance_from_orb，准备获取 L2 生态的实时金额")
    try:
        # 下拉按钮
        from_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                  '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/div[2]/div[2]/div[1]')))  # Ethereum   #Polygon  #Arbitrum One
        time_sleep(2,"开始找from_source的余额")
        browser.execute_script("arguments[0].click();", from_list_button)

        # 选择from chain
        From_chain_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH,
             f"//*[@id='app']/div[1]/div[2]/div/div[2]/div/div[5]/div[2]/div/div/div[2]/div//span[text()='{from_source}']")))  # "zkSync" "Arbitrum" "Optimism"
        time_sleep(2)
        browser.execute_script("arguments[0].click();", From_chain_button)
    except:
        print("L2_orb_L2_prepare_transfer_coin(), 没有找到 from 链")

    
    time_sleep(2, "开始找to_source的余额")

    try:
        # 下拉按钮
        to_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                  "//*[@id='app']/div[1]/div[2]/div/div[2]/div/div[3]/div[2]/div[1]")))  # Ethereum   #Polygon  #Arbitrum One
        time_sleep(2)
        browser.execute_script("arguments[0].click();", to_list_button)

        # 选择from chain
        to_chain_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH,
             f"//*[@id='app']/div[1]/div[2]/div/div[2]/div/div[6]/div[2]/div/div/div[2]/div//span[text()='{to_source}']")))  # "zkSync" "Arbitrum" "Optimism"
        time_sleep(2)
        browser.execute_script("arguments[0].click();", to_chain_button)
    except:
        print("L2_orb_L2_prepare_transfer_coin(), 没有找到 to 链")

    time_sleep(5, "等待orb显示出余额")
    from_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/div[2]/div[1]/div[2]/span')))
    to_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[2]/div/div[2]/div/div[3]/div[1]/div[2]/span')))
    

    return float(from_balance_button.text),float(to_balance_button.text)


##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ orb 的一些函数↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ bungee 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#

#如果要连接钱包的话，选择小狐狸
def bungee_whether_connect_wallet(browser, wait):
    print("我已经进入bungee_whether_connect_wallet，准备判断是否连接到钱包")
    try:
        # ====== 判断是否连接到钱包
        wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='pb-px text-xs text-white uppercase']")))
        print("bungee_whether_connect_wallet(), 找到的钱包按钮文本是：", wallet_button.text)
        if wallet_button.text == "Connect Wallet" or wallet_button.text == "CONNECT WALLET":
            time_sleep(2)
            browser.execute_script("arguments[0].click();", wallet_button)
            # ======== 选择 Metamask
            metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//p[@class='text-md font-medium capitalize' and text() = 'MetaMask']")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", metamask_button)
            time_sleep(5, "等待小狐狸钱包连接")
    except:
        print("bungee_whether_connect_wallet(), 没有找到钱包按钮，可能已经连接了钱包")


def bungee_Agree_and_continue(browser, wait):
    print("我已经进入bungee_Agree_and_continue，尝试点击 Agree and continue")
    # 点击 Agree and continue
    try:#可能有警告
        Agree_and_continue = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Agree and Continue']")))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", Agree_and_continue)
    except:
        print("没有警告")


# bungee选择from、to链、转账金额
def bungee_prepare_transfer_coin(browser, wait, from_source, to_destination, min_rate, max_rate):
    print("我已经进入bungee_prepare_transfer_coin，选择from、to，准备转账")
    #=====先点击from下拉框
    from_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[1]/div[1]/div/div[2]/div[1]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", from_list_button)
    time_sleep(5, "已经点击 from 下拉框, 等待bungee显示出各个网络")

    #===选择from chain
    your_from = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[1]/div[1]/div/div[2]//span[text()="{from_source}"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", your_from)
    time_sleep(5, "已经点击from chain, 等待bungee显示出各个网络")

    #==== 先点击to下拉框
    to_list_button = wait.until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[1]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", to_list_button)
    time_sleep(5, "已经点击 to 下拉框, 等待bungee显示出各个网络")
    your_to = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[1]/div[3]/div/div[2]//span[text()="{to_destination}"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", your_to)
    print("已经点击to chain")

    #========from coin list
    from_coin_list_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[3]/div[3]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", from_coin_list_button)
    time_sleep(2, "已经点击from coin list，等待bungee显示出各种币")

    # ==== 选择from coin
    from_coin = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[text()="ETH"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", from_coin)
    print("已经点击 from_coin")

    # ========to coin list
    to_coin_list_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[5]/div[3]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", to_coin_list_button)
    time_sleep(2, "已经点击to coin list, 等待bungee显示出各种币")

    # ========to coin
    to_coin = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[text()="ETH"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", to_coin)
    print("已经点击 to_coin")

    #获取 balance
    balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[2]/div[1]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("取到的amount是:", amount)
    L2_ETH_value = float(amount)
    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_ETH_value * min_rate, L2_ETH_value * max_rate), point)
    try_times = 0
    while input_value > float(L2_ETH_value) or input_value == 0 :
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * min_rate, L2_ETH_value * max_rate), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    #输入input
    input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/main/div/div[1]/div/div/div[3]/div[1]/input')))
    input_button.send_keys(str(input_value)) #要换成字符串
    time_sleep(3, f"已经输入input：{input_value}",)

    #选择 hop
    hop_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-row justify-between w-full relative items-center px-5']//span[@class='text-xs hidden md:inline' and contains(text(), 'Hop')]")))
    # '//*[@id="__next"]/div/div/main/div/div[2]/div/div[2]/div[1]/div[2]/div[3]/div/div[1]/span'
    browser.execute_script("arguments[0].click();", hop_button)
    time_sleep(10, "已经选择 hop，经过修改")

    # 点击 Proceed
    proceed_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Proceed']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", proceed_button)
    time_sleep(10, "已经点击 Proceed")

    # 点击 Bridge
    bridge_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Bridge']")))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", bridge_button)
    time_sleep(10, "已经点击 Bridge")
    return f"本次转账金额{input_value}"
##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ bungee 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ hop 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓#
def hop_prepare_transfer_coin(browser, wait, from_source, to_destination, min_rate, max_rate):
    print("我已经进入hop_prepare_transfer_coin，选择from、to，准备转账")
    #=====先点击from下拉框
    from_list_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#root > div > div.sc-gsDKAQ.dtArT > div > div > div:nth-child(2) > div.sc-dkPtRN.fmFFDZ > div.MuiInputBase-root.MuiInput-root > div')))
    time_sleep(2)
    ActionChains(browser).click(from_list_button).perform()  # 模拟鼠标点
    time_sleep(5, "已经点击 from 下拉框, 等待 hop 显示出各个网络")

    #===选择from chain
    your_from = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="menu-"]/div[3]/ul//h6[text() = "{from_source}"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", your_from)
    time_sleep(5, "已经点击from chain, 等待bungee显示出各个网络")

    #==== 先点击to下拉框
    to_list_button = wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, '#root > div > div.sc-gsDKAQ.dtArT > div > div > div:nth-child(4) > div.sc-dkPtRN.fmFFDZ > div.MuiInputBase-root.MuiInput-root > div > div > div')))
    time_sleep(2)
    ActionChains(browser).click(to_list_button).perform()  # 模拟鼠标点
    time_sleep(5, "已经点击 to 下拉框, 等待 hop 显示出各个网络")
    your_to = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="menu-"]/div[3]/ul//h6[text()="{to_destination}"]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", your_to)
    print("已经点击to chain")



    #获取 balance
    balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div[2]/div[1]/div/h6')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("取到的amount是:", amount)
    L2_ETH_value = float(amount)
    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_ETH_value * min_rate, L2_ETH_value * max_rate), point)
    try_times = 0
    while input_value > float(L2_ETH_value) or input_value == 0 :
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * min_rate, L2_ETH_value * max_rate), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    #输入input
    input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div[2]/div[2]/div[2]/div/input')))
    input_button.send_keys(str(input_value)) #要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击 send
    send_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div[7]/div/div/button/span[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", send_button)
    time_sleep(5, "已经点击 send")

    # 确认点击 send
    confirm_send_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", confirm_send_button)
    time_sleep(15, "已经再次点击 send")
    return f"本次转账金额{input_value}"

#如果要连接钱包的话，选择小狐狸
def hop_whether_connect_wallet(browser, wait):
    print("我已经进入 hop_whether_connect_wallet，准备判断是否连接到钱包")
    try:
        # ====== 判断是否连接到钱包
        wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[1]/div[3]/div[3]/div/button/span[1]')))
        print("hop_whether_connect_wallet(), 找到的钱包按钮文本是：", wallet_button.text)
        if wallet_button.text == "Connect a Wallet":
            time_sleep(2)
            browser.execute_script("arguments[0].click();", wallet_button)
            # ======== 选择 Metamask
            metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/aside/section/ul/li[1]/button")))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", metamask_button)
            time_sleep(5, "等待小狐狸钱包连接")
    except:
        print("bungee_whether_connect_wallet(), 没有找到的钱包按钮")


##↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ hop 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  gmx 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
gmx_trade_url = 'https://gmx.io/trade'
gmx_buy_url = 'https://gmx.io/buy_glp'
gmx_swap_url = 'https://gmx.io/trade'

#如果要连接钱包的话，选择小狐狸
def gmx_whether_connect_wallet(browser, wait):
    print("我已经进入 gmx_whether_connect_wallet，准备判断是否连接到钱包")
    try:
        # ====== 判断是否连接到钱包
        wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[2]/div/button/span')))
        print("gmx_whether_connect_wallet(), 找到的钱包按钮文本是：", wallet_button.text)
        if wallet_button.text == "Connect Wallet":
            time_sleep(2)
            browser.execute_script("arguments[0].click();", wallet_button)
            # ======== 选择 Metamask
            metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[4]/div[2]/div[3]/button[1]')))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", metamask_button)
            time_sleep(5, "等待小狐狸钱包连接")
    except:
        print("gmx_whether_connect_wallet(), 没有找到的钱包按钮，可能已经连接了钱包")

#准备杠杆交易，long
def gmx_prepare_input(browser, wait):
    print("我已经进入 gmx_prepare_input，准备输入")
    # 点击 trade
    trade_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[2]/div/div[1]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", trade_button)
    time_sleep(2, "已经点击 trade")

    # 选择 long
    long_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", long_button)
    time_sleep(2, "已经点击 long")
    #选择杠杆
    a = random.randint(1, 2)
    if a == 1:
        print("选择2倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(1)')))
    elif a == 2:
        print("选择5倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(2)')))
    time_sleep(2)
    ActionChains(browser).click(choose_leverage).perform()  # 模拟鼠标点
    # browser.execute_script("arguments[0].click();", choose_leverage)

    # 获取实时价格
    real_time_price_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]')))
    a = real_time_price_button.text
    real_time_price = float(a)
    print(f"取到的实时金额是：{real_time_price}")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(12/real_time_price, L2_ETH_value), point)
    print(f"输入金额：{input_value}，换算为美金是{input_value*real_time_price}")

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

def gmx_allow_leverage_long_trade(browser, wait):
    print("我已经进入gmx_allow_leverage_long_trade，允许使用杠杆 ")

    # 点击 trade
    trade_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[2]/div/div[1]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", trade_button)
    time_sleep(2, "已经点击 trade")

    # 选择 long
    long_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", long_button)
    time_sleep(2, "已经点击 long")
    #选择杠杆
    a = random.randint(1, 2)
    if a == 1:
        print("选择2倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(1)')))
    elif a == 2:
        print("选择5倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(2)')))
    time_sleep(2)
    ActionChains(browser).click(choose_leverage).perform()  # 模拟鼠标点
    # browser.execute_script("arguments[0].click();", choose_leverage)

    # 获取实时价格
    real_time_price_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]')))
    a = real_time_price_button.text
    real_time_price = float(a)
    print(f"取到的实时金额是：{real_time_price}")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(12/real_time_price, L2_ETH_value), point)
    print(f"输入金额：{input_value}，换算为美金是{input_value*real_time_price}")

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击允许杠杆
    enable_leverage_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", enable_leverage_button)
    time_sleep(10, "已经点击 enable leverage，等待小狐狸确认")

    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

def gmx_whether_allow_leverage_or_formal_long(browser, wait):
    print("我已经进入gmx_allow_leverage_or_formal_long，判断是否允许杠杆 ")
    # 查看是什么文本，从而判断是否要允许杠杆
    text_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    print(f"找到的文本按钮是：{text_button.text}")
    return text_button.text
    # browser.execute_script("arguments[0].click();", enable_leverage_button)
    # time_sleep(10, "已经点击 enable leverage，等待小狐狸确认")


#正式杠杆交易，long
def gmx_formal_long_trade(browser, wait):
    browser.refresh()
    time_sleep(10, "我已经进入gmx_formal_long_trade，开始正式交易")
    # 点击 trade
    trade_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[2]/div/div[1]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", trade_button)
    time_sleep(2, "已经点击 trade")

    # 选择 long
    long_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", long_button)
    time_sleep(2, "已经点击 long")
    #选择杠杆
    a = random.randint(1, 2)
    if a == 1:
        print("选择2倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(1)')))
    elif a == 2:
        print("选择5倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.positive > div > div.rc-slider-mark > span:nth-child(2)')))
    time_sleep(2)
    ActionChains(browser).click(choose_leverage).perform()  # 模拟鼠标点
    # browser.execute_script("arguments[0].click();", choose_leverage)

    # 获取实时价格
    real_time_price_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]')))
    a = real_time_price_button.text
    real_time_price = float(a)
    print(f"取到的实时金额是：{real_time_price}")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    b = random.randint(12, 15)
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(b/real_time_price, L2_ETH_value-0.005), point)
    print(f"输入金额：{input_value}，换算为美金是{input_value*real_time_price}")

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击long eth
    long_eth_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", long_eth_button)
    time_sleep(10, "已经点击 Long ETH，等待小狐狸确认")

    # 点击 long
    long_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[4]/div/div[2]/div[3]/div[2]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", long_button)
    time_sleep(5, "已经点击 long ")


#准备杠杆交易，short
def gmx_formal_short_trade(browser, wait):
    print("我已经进入gmx_formal_short_trade，准备开杠杆short")
    # 选择short
    short_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", short_button)
    time_sleep(2, "已经点击 short")
    swap = '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[3]'

    # 选择 token_list
    token_list = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", token_list)
    time_sleep(2, "已经点击 token_list")

    ETH_token = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", ETH_token)
    time_sleep(2, "已经点击 ETH_token")

    #选择杠杆
    a = random.randint(1, 2)
    if a == 1:
        print("选择2倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.negative > div > div.rc-slider-mark > span:nth-child(1)')))
    elif a == 2:
        print("选择5倍杠杆")
        choose_leverage = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '#root > div.App > div > div > div.Exchange-content > div.Exchange-right > div.Exchange-swap-box > div.Exchange-swap-box-inner.App-box-highlight > div.Exchange-leverage-box > div.Exchange-leverage-slider.App-slider.negative > div > div.rc-slider-mark > span:nth-child(2)')))
    time_sleep(2)
    ActionChains(browser).click(choose_leverage).perform()  # 模拟鼠标点
    # browser.execute_script("arguments[0].click();", choose_leverage)

    # 获取实时价格
    real_time_price_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]')))
    a = real_time_price_button.text
    real_time_price = float(a)
    print(f"取到的实时金额是：{real_time_price}")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    b = random.randint(12, 15) #随机美元数量
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(b/real_time_price, L2_ETH_value-0.005), point)
    print(f"输入金额：{input_value}，换算为美金是{input_value*real_time_price}")

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击long eth
    short_eth_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", short_eth_button)
    time_sleep(10, "已经点击 short ETH ")

    # 点击 short
    short_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[4]/div/div[2]/div[3]/div[2]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", short_button)
    time_sleep(5, "已经点击 short ")

    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

#准备 swap
def gmx_prepare_ETH_swap_token(browser, wait, token):
    print("我已经进入gmx_prepare_ETH_swap_token ")
    # 选择 swap
    swap_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[3]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(2, "已经点击 swap")

    # 选择from token_list
    from_token_list = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", from_token_list)
    time_sleep(2, "已经点击 token_list")

    ETH_token = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH,
             '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", ETH_token)
    time_sleep(2, "已经点击 ETH_token")

    # 选择receive token_list
    receive_token_list = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[4]/div[2]/div[2]/div/div')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", receive_token_list)
    time_sleep(2, "已经点击 token_list")
    if token == "USDT":
        your_token = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[4]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[8]')))
    elif token == "USDC":
        your_token = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH,
                 '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[4]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[7]')))

    time_sleep(2)
    browser.execute_script("arguments[0].click();", your_token)
    time_sleep(2, "已经点击 USDT_token")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_ETH_value * 0.8, L2_ETH_value * 0.85), point)
    try_times = 0
    while input_value > float(L2_ETH_value) or input_value == 0:
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击 swap
    swap_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(10, "已经点击 swap，等待小狐狸确认")

    # 点击
    confirm_swap_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[4]/div/div[2]/div[3]/div[2]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", confirm_swap_button)
    time_sleep(10, "已经点击 confirm swap，等待小狐狸确认")

    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

#准备 swap
def gmx_prepare_USDT_swap_ETH(browser, wait):
    print("我已经进入gmx_prepare_USDT_swap_ETH ")
    # 选择 swap
    swap_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[3]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(2, "已经点击 swap")

    # 选择from token_list
    from_token_list = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", from_token_list)
    time_sleep(2, "已经点击 token_list")

    USDT_token = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH,
             '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[8]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", USDT_token)
    time_sleep(2, "已经点击 USDT_token")

    # 选择receive token_list
    receive_token_list = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[4]/div[2]/div[2]/div/div')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", receive_token_list)
    time_sleep(2, "已经点击 token_list")

    ETH_token = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[4]/div[2]/div[2]/div/div[1]/div[2]/div[3]/div/div[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", ETH_token)
    time_sleep(2, "已经点击 ETH_token")

    # 获取 balance
    balance_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]')))
    a = balance_button.text
    print("取到的balance是：", a)
    b = a.split(": ")
    amount = b[-1]
    print("分割后，取到的amount是:", amount)
    L2_ETH_value = float(amount)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
    try_times = 0
    while input_value > float(L2_ETH_value) or input_value == 0:
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击
    swap_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(10, "已经点击 swap，等待小狐狸确认")

    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

#判断是否要输入推荐码
def gmx_whether_input_referal(browser, wait):
    print("我已经进入 gmx_whether_input_referal, 判断是否要输入输入邀请码")
    # 点击 referral
    referral_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[1]/div/div[5]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", referral_button)
    time_sleep(5, "已经点击 referral")
    # 推荐码
    try:
        input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div[1]/div[3]/div/form/input')))
        print("需要输入推荐码")
        return "yes to referal"
    except:
        print("不用需要再输入推荐码")
        return "no need to referal"


#输入推荐码
def gmx_input_referal(browser, wait):
    try:
        print("我已经进入gmx_try_referal，准备输入邀请码")
        # 点击 referral
        referral_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[1]/div/div[5]/a')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", referral_button)
        time_sleep(5, "已经点击 referral")

        # 推荐码
        input_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div[1]/div[3]/div/form/input')))
        input_button.send_keys("Bruce1ZY")  # 要换成字符串
        time_sleep(2, "已经输入邀请码")

        # 点击submit
        submit_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div[1]/div[3]/div/form/button')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", submit_button)
        time_sleep(5, "已经点击 submit")
    except:
        print("可能已经输入了邀请码")


def gmx_buy_or_sell_GLP(browser, wait, buy_or_sell):
    print("我已经进入gmx_GLP，准备提供流动性 ")

    # 点击 trade
    trade_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[2]/div/div[1]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", trade_button)
    time_sleep(2, "已经点击 trade")

    # 选择 buy
    buy_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/header/div[1]/div[1]/div/div[4]/a')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", buy_button)
    time_sleep(2, "已经点击 Buy 模块")

    # 选择 arb
    arb_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/a[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", arb_button)
    time_sleep(2, "已经点击 arb")

    #选择买/卖流动性
    if buy_or_sell == "buy":
        buy_or_sell_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[1]/div[2]/div[1]/div[1]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", buy_or_sell_button)
        time_sleep(2, f"已经点击 {buy_or_sell}流动性")

        #选择用ETH买
        buy_with_eth = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[3]/table/tbody/tr[1]/td[6]/button')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", buy_or_sell_button)
        time_sleep(2, f"已经点击用 ETH 买流动性")

        # 获取实时价格
        real_time_price_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[3]/table/tbody/tr[1]/td[2]')))
        a = real_time_price_button.text
        b = a.split("$")[-1] #以$作为分隔符，提取到'1,145.98'
        c = b.split(",")[0] #以 , 为分隔符，提取到前面数字
        d = b.split(",")[-1]  # 以 , 为分隔符，提取到后面数字
        e = c+d
        real_time_price = float(e)
        print(f"取到的实时价格是：{a}，分割前的价格{b}，分割后的价格：{real_time_price}")

        # 获取 balance
        balance_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/span[2]')))
        c = balance_button.text
        print("取到的balance是：", c)
        L2_ETH_value = float(c)

        # =======随机金额
        U_amount = random.randint(12, 14)

        point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(U_amount / real_time_price, L2_ETH_value-0.006), point)
        print(f"选择Uamout:{U_amount}，输入金额：{input_value}，换算为美金是{input_value * real_time_price}")

        # 输入input
        input_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[1]/div[2]/div[2]/div[2]/div[1]/input')))
        input_button.send_keys(str(input_value))  # 要换成字符串
        time_sleep(3, f"已经输入input{input_value}")

        #已经点击Buy
        buy_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[1]/div[2]/div[6]/button')))
        browser.execute_script("arguments[0].click();", buy_button)
        time_sleep(5, "已经点击 buy ")


    elif buy_or_sell == "sell":
        buy_or_sell_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[2]/div[1]/div[2]/div[1]/div[2]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", buy_or_sell_button)
        time_sleep(2, f"已经点击 {buy_or_sell}")




    # # 点击 send
    # enter_amount_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[1]/div/div/div[1]/div[2]/div[1]/div[1]/div[6]/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", enter_amount_button)
    # time_sleep(5, "已经点击 send")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"



## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  gmx的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  Yield的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
yield_pool_url = 'https://app.yieldprotocol.com/pool'
#如果要连接钱包的话，选择小狐狸
def yield_whether_connect_wallet(browser, wait):
    print("我已经进入 yield_whether_connect_wallet，准备判断是否连接到钱包")
    try:
        # ====== 判断是否连接到钱包
        wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/header/div/div[3]/div/span')))
        print("yield_whether_connect_wallet(), 找到的钱包按钮文本是：", wallet_button.text)
        if wallet_button.text == "Connect Wallet":
            time_sleep(2)
            browser.execute_script("arguments[0].click();", wallet_button)
            # ======== 选择 Metamask
            metamask_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/div/div[3]/div[1]/div/div/span')))
            time_sleep(2)
            browser.execute_script("arguments[0].click();", metamask_button)
            time_sleep(5, "等待小狐狸钱包连接")
    except:
        print("yield_whether_connect_wallet(), 没有找到的钱包按钮，可能已经连接了钱包")


#准备 pool
def yield_prepare_pool(browser, wait):
    print("我已经进入 yield_prepare_pool ")
    # 获取usdc余额
    balance_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/header/div/div[3]/div/div[1]/div/button/div/div/span[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", balance_button)
    time_sleep(2, "已经点击 balance")

    a = balance_button.text
    L2_ETH_value = float(a)
    print(f"取到的balance是：{a}，分割后是：{L2_ETH_value}", a)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(51, L2_ETH_value ), point)
    try_times = 0
    while input_value < 51: # 包括0 的情况
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="topsection"]/div/div[3]/div[1]/div[1]/div/div[1]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击 next
    next_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[3]/div/div/div[2]/div/div/div[3]/div/button/div/span/span')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", next_button)
    time_sleep(2, "已经点击 next，等待小狐狸确认")

    # 点击 I_understand_button
    I_understand_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="midSection"]/label/div/div')))
    browser.execute_script("arguments[0].click();", I_understand_button)
    # time_sleep(2)
    # I_understand_button = wait.until(EC.element_to_be_clickable(
    #     (By.CSS_SELECTOR, '#midSection > label > div > div')))
    # time_sleep(2)
    # ActionChains(browser).click(I_understand_button).perform()  # 模拟鼠标点

    time_sleep(2, "已经点击 I_understand_button ")

    # 点击 pool
    pool_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[3]/div/div/div[2]/div/div/div[3]/div/button')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", pool_button)
    time_sleep(10, "已经点击 pool")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"

#准备 lend
def yield_prepare_lend(browser, wait):
    print("我已经进入 yield_prepare_lend ")
    # 点击lend
    lend_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[1]/div[1]/span[2]/span')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", lend_button)
    time_sleep(2, "已经点击 lend 页面")

    # 获取usdc余额
    balance_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/header/div/div[3]/div/div[1]/div/button/div/div/span[2]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", balance_button)
    time_sleep(2, "已经点击 balance")

    a = balance_button.text
    L2_ETH_value = float(a)
    print(f"取到的balance是：{a}，分割后是：{L2_ETH_value}", a)

    # =======随机金额
    point = random.randint(2, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
    input_value = round(random.uniform(51, L2_ETH_value ), point)
    try_times = 0
    while input_value < 51: # 包括0 的情况
        point = random.randint(2, 6)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.7, L2_ETH_value * 0.8), point)
        try_times += 1
        if try_times == 10:
            print("超过最大重复取值次数")
            break

    # 输入input
    input_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="topsection"]/div/div[3]/div[1]/div/div[1]/div/div[1]/div[1]/input')))
    input_button.send_keys(str(input_value))  # 要换成字符串
    time_sleep(3, f"已经输入input{input_value}")

    # 点击 next
    orange_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH,
             '//*[@id="topsection"]/div/div[3]/div[3]/div[3]/div/div/div[1]')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", orange_button)
    time_sleep(2, "已经点击 orange 图标")

    # 点击 next
    next_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[3]/div/div/div[2]/div/div/div[3]/div/button/div/span/span')))
    time_sleep(2)
    browser.execute_script("arguments[0].click();", next_button)
    time_sleep(2, "已经点击 next，等待小狐狸确认")

    # 点击 lend
    lend_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[3]/div/div/div[2]/div/div/div[3]/div/button/span')))
    browser.execute_script("arguments[0].click();", lend_button)
    time_sleep(2, "已经点击lend")

    # # 点击 pool
    # pool_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/div/div[3]/div[3]/div/div/div[2]/div/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", pool_button)
    # time_sleep(10, "已经点击 pool")
    #
    # # 确认点击 send
    # confirm_send_button = wait.until(
    #     EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div[4]/div/div/div[2]/div/div[3]/div/button')))
    # time_sleep(2)
    # browser.execute_script("arguments[0].click();", confirm_send_button)
    # time_sleep(15, "已经再次点击 send")
    # return f"本次转账金额{input_value}"


#小狐狸【确认】showme
def fox_confirm_yield_signature(browser, wait):
    print('fox_confirm_yield_signature，【从小狐狸端】确认签名')
    browser.refresh()
    time_sleep(8)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[4]/button[2]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_gmx_referal 小狐狸点击【签名】失败了，是否影响？")
        return "fox_confirm_gmx_referal 小狐狸点击【签名】失败了，是否影响？"


## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  Yield的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  syncswap的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
#准备工作：登陆小狐狸，获取小狐狸账户个数。这是第0个标签
url_dashboard = 'https://clash.razord.top/#/proxies'
sync_swap_trade = "https://syncswap.xyz/swap" #用于做swap任务
sync_swap_pool = "https://syncswap.xyz/pool/add"#用于流动性任务
sync_swap_remove = "https://syncswap.xyz/pool/remove"#用于流动性任务


def open_zk_bridge_page(browser, wait):
    print("我已经进入 open_zk_bridge_page ")

    zktools_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="__layout"]/main/header/div[2]/div/div/div[2]/div/div/a')))
    time_sleep(1,"准备点击zktools_button")
    browser.execute_script("arguments[0].click();", zktools_button)

    zksync_v_2_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dropdown-menu-rf42ptxsx"]/a[2]')))
    time_sleep(1,"准备点击deposit")
    browser.execute_script("arguments[0].click();", zksync_v_2_button)


def goerli_login_metamask(browser, wait):
    print("我已经进入 goerli_login_metamask,选择小狐狸 ")
    try:
        fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/section/div[2]/button')))
        time_sleep(1,"准备点击小狐狸")
        browser.execute_script("arguments[0].click();", fox_button)
    except:
        print("可能已经进入了小狐狸")

def transfer_goerli_from_eth_to_zk(browser, wait):
    print("我已经进入 transfer_goerli_from_eth_to_zk，准备转goerli")
    for i in range(1,6):
        try:
            time_sleep(9, "准备寻找max")
            max_button = browser.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/main/div/div/section/div[2]/div/form/div[1]/div[2]/div[2]/div/button')
            time_sleep(1,"准备点击max")
            browser.execute_script("arguments[0].click();", max_button)

            # max_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/main/div/div/section/div[2]/div/form/div[1]/div[2]/div[2]/div/button')))
            # time_sleep(1,"准备点击max")
            # browser.execute_script("arguments[0].click();", max_button)

            deposit_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/main/div/div/section/div[2]/div/form/button[2]')))
            time_sleep(1,"准备点击deposit")
            browser.execute_script("arguments[0].click();", deposit_button)
            time_sleep(5)
            break
        except:
            print(f"第{i}次点击max金额出错,开始左右点击")
            withdraw_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/main/div/div/section/div[1]/nav/ul/li[3]/a')))
            time_sleep(1,"准备点击 withdraw_button")
            browser.execute_script("arguments[0].click();", withdraw_button)
            time_sleep(7)
            deposit_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/main/div/div/section/div[1]/nav/ul/li[2]/a')))
            time_sleep(1,"准备点击 deposit_button")
            browser.execute_script("arguments[0].click();", deposit_button)
            time_sleep(15)
    

#小狐狸【确认】
def fox_confirm_goerli_transfer(browser, wait):
    print('fox_confirm_goerli_transfer，【从小狐狸端】确认')
    browser.refresh()
    time_sleep(5)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[3]/footer/button[2]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_confirm_goerli_transfer 小狐狸点击【签名】失败了，是否影响？")
        try:#拒绝
            cancle_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/div[4]/footer/button[1]')))
            time_sleep(1)
            browser.execute_script("arguments[0].click();", cancle_button)
            print("fox失败, 点击拒绝交易")
            return "fox失败, 点击拒绝交易"
        except:
            print("小狐狸拒绝前面失败")
            return "fox失败"

#小狐狸【允许】
def fox_allow_use_goerli_usdc(browser, wait):
    print('fox_allow_goerli_usdc，【从小狐狸端】确认')
    browser.refresh()
    time_sleep(5)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/footer/button[2]')))
        time_sleep(2)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_allow_goerli_usdc 小狐狸点击【签名】失败了，是否影响？")
        try:#拒绝
            cancle_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[5]/footer/button[1]')))
            time_sleep(1)
            browser.execute_script("arguments[0].click();", cancle_button)
            print("fox失败, 点击拒绝交易")
            return "fox失败, 点击拒绝交易"
        except:
            print("小狐狸拒绝前面失败")
            return "fox失败"

#小狐狸【允许】liudx
def fox_allow_syncswap_use_LP(browser, wait):
    print('fox_allow_goerli_usdc，【从小狐狸端】确认')
    browser.refresh()
    time_sleep(5)
    try:
        # 签名
        sign_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[4]/button[2]')))
        time_sleep(1)
        browser.execute_script("arguments[0].click();", sign_button)
        print("小狐狸签名成功")
        return "小狐狸签名成功"
    except:
        print("fox_allow_goerli_usdc 小狐狸点击【签名】失败了，是否影响？")
        try:#拒绝
            cancle_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[4]/button[1]')))
            time_sleep(1)
            browser.execute_script("arguments[0].click();", cancle_button)
            print("fox失败, 点击拒绝交易")
            return "fox失败, 点击拒绝交易"
        except:
            print("小狐狸拒绝前面失败")
            return "fox失败了"



#任务1:ETH转USDC.
def ETH_swap_USDC(browser, wait, excel_path,excel_row, write_excel_column): #后面两个参数用于记录信息到excel
    #新建标签页,准备转goerli
    new_tab(browser, sync_swap_trade)
    time_sleep(5,"等待网络加载")
    switch_tab_by_handle(browser, 2, 0)

    #连接小狐狸钱包
    wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    time_sleep(1, "准备点击连接钱包")
    browser.execute_script("arguments[0].click();", wallet_button)
    time_sleep(5, "等等小狐狸出现")
    fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div[2]/div[2]/div[1]')))
    time_sleep(1, "小狐狸出现了,准备点击")
    browser.execute_script("arguments[0].click();", fox_button)
    time_sleep(8, "已经点击小狐狸,准备获取余额")
    
    #获取余额
    balance = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[2]/div[2]/p')))
    a = balance.text   #Balance : 0.99909
    b = a.split()[-1]
    L2_balance = float(b)
    print(f"获取到的余额是: {L2_balance}")
    point = random.randint(2, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    input_value = round(random.uniform(L2_balance * 0.2, L2_balance * 0.6), point)  #随机金额

    #输入余额
    pay_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[1]/input')))
    time_sleep(1, f"准备输入余额:{input_value}")
    pay_input_button.send_keys(str(input_value))

    #准备交易
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    time_sleep(1, "准备点击swap")
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(10, "已经点击swap")

    #切換到小狐狸,准备确认
    switch_tab_by_handle(browser, 1, 1)
    fox_statue = fox_confirm_goerli_transfer(browser, wait)
    if "失败" in fox_statue:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号小狐狸签名失败了, 已经记录下来")

    elif "成功" in fox_statue:
        # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
        print("小狐狸签名成功!! ")
    time_sleep(15)

    #切換到网站, 查看是否成功
    switch_tab_by_handle(browser, 2, 0)
    time_sleep(20,"等一会网页")

    web_statue = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div/div[2]/p')))
    a = web_statue.text   #Balance : 0.99909 = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[2]/div[2]/p')))
    print(a)
    if "Confirmed" not in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号失败了,已经记录下来")

    elif "Confirmed" in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "成功")
        print("成功!! 已经记录到excel")

#任务2:USDC转ETH.
def USDC_swap_ETH(browser, wait, excel_path,excel_row, write_excel_column, mode): #后面两个参数用于记录信息到excel
    #新建标签页,准备转goerli
    new_tab(browser, sync_swap_trade)
    time_sleep(5,"等待网络加载")
    switch_tab_by_handle(browser, 2, 0)

    #连接小狐狸钱包
    wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    time_sleep(5, "准备点击连接钱包")
    browser.execute_script("arguments[0].click();", wallet_button)
    time_sleep(5, "等等小狐狸出现")
    fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div[2]/div[2]/div[1]')))
    time_sleep(5, "小狐狸出现了,准备点击")
    browser.execute_script("arguments[0].click();", fox_button)
    time_sleep(13, "已经点击小狐狸")
    
    #调换一下USDC和ETH的位置
    position_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[2]')))
    time_sleep(1, "准备调换USDC和ETH位置")
    browser.execute_script("arguments[0].click();", position_button)

    if mode == 0:
        print("模式0, 随机转USDC (25~75%)")
        percent = random.randint(1, 3)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
        percent_button = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[3]/button[{percent}]')))
        time_sleep(8, "准备点击USDC比例")
        browser.execute_script("arguments[0].click();", percent_button)
        print("已经点击转出比例")

    elif mode == 1:
        print("模式1, 转走全部USDC (100%)")
        percent_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[3]/button[4]')))
        time_sleep(8, "USDC比例100%")
        browser.execute_script("arguments[0].click();", percent_button)
        print("已经点击全部转出")
    
    your_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    if "Approve" in your_button.text:
        #准备批准usdc
        time_sleep(1, "确实要批准使用usdc代币")
        browser.execute_script("arguments[0].click();", your_button)
        time_sleep(15, "已经点击allow")

        #切換到小狐狸,准备确认
        switch_tab_by_handle(browser, 1, 1)
        fox_statue = fox_allow_use_goerli_usdc(browser, wait)
        time_sleep(10)
        switch_tab_by_handle(browser, 2, 0) #切换回网页
        time_sleep(6)
        #关闭提醒
        close_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div/div[2]/button')))
        time_sleep(1, "准备关闭提醒")
        browser.execute_script("arguments[0].click();", close_button)
        if "失败" in fox_statue:
            Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
            print("这个号小狐狸签名失败了, 已经记录下来")

        elif "成功" in fox_statue:
            # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
            print("小狐狸签名成功!! ")
    
    
    #准备交易
    switch_tab_by_handle(browser, 2, 0) #切换回网页
    time_sleep(8, "准备寻找swap") #必须加延时,否则找不到元素
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    time_sleep(8, "准备点击swap")
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(10, "已经点击swap")

    #切換到小狐狸,准备确认
    switch_tab_by_handle(browser, 1, 1)
    fox_statue = fox_confirm_goerli_transfer(browser, wait)
    if "失败" in fox_statue:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号小狐狸签名失败了, 已经记录下来")

    elif "成功" in fox_statue:
        # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
        print("小狐狸签名成功!! ")
    time_sleep(15)

    #切換到网站, 查看是否成功
    switch_tab_by_handle(browser, 2, 0)
    time_sleep(20,"等一会网页")
    web_statue = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div/div[2]/p')))
    a = web_statue.text   #Balance : 0.99909 = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[2]/div[2]/p')))
    print(a)
    if "Confirmed" not in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号失败了,已经记录下来")

    elif "Confirmed" in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "成功")
        print("成功!! 已经记录到excel")


#任务3:提供流动性
def syncswap_provide_LP(browser, wait,excel_path, excel_row, write_excel_column): #后面两个参数用于记录信息到excel
    #新建标签页,准备转goerli
    new_tab(browser, sync_swap_pool)
    time_sleep(5,"等待网络加载")
    switch_tab_by_handle(browser, 2, 0)
    wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/button')))
    print("找到的文字是:",wallet_button.text)
    
    if "Connect" in wallet_button.text:#如果没有链接钱包,则链接钱包
        #连接小狐狸钱包
        time_sleep(1, "准备点击连接钱包")
        browser.execute_script("arguments[0].click();", wallet_button)
        time_sleep(5, "等等小狐狸出现")
        fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div[2]/div[2]/div[1]')))
        time_sleep(1, "小狐狸出现了,准备点击")
        browser.execute_script("arguments[0].click();", fox_button)
        time_sleep(10, "已经点击小狐狸")
    
    # #输入交易额,法一:随机ETH值
    # #获取余额
    # balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div/div[2]/div/div')))
    # a = balance_button.text   #Balance : 0.99909
    # b = a.split()[-1]
    # L2_balance = float(b)
    # print(f"获取到的余额是: {L2_balance}")
    # point = random.randint(2, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
    # input_value = round(random.uniform(L2_balance * 0.2, L2_balance * 0.6), point)  #随机金额

    # #输入余额
    # pay_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div/div[2]/input')))
    # time_sleep(1, f"准备输入余额:{input_value}")
    # pay_input_button.send_keys(str(input_value))

    #法二: 先尝试点击 max USDC,如果提示不足,则点击 eth
    max_usdc_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div/div/div[2]/div/button')))
    time_sleep(1)
    browser.execute_script("arguments[0].click();", max_usdc_button)
    time_sleep(1, "已经点击 max usdc")
    #如果提示不足,则说明要maxETH
    try:
        text_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/button')))
        if "fficient" in text_button.text:
            print("ETH不足,需要随机输入ETH金额")
            #获取ETH余额
            eth_balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div/div[2]/div/div')))
            a = eth_balance_button.text   #Balance : 0.99909
            b = a.split()[-1]
            L2_balance = float(b)
            print(f"获取到的余额是: {L2_balance}")
            point = random.randint(2, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
            input_value = round(random.uniform(L2_balance * 0.2, L2_balance * 0.6), point)  #随机金额

            # #输入余额
            pay_input_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div/div[2]/input')))
            time.sleep(2)
            #先发送清除
            pay_input_button.clear()
            time_sleep(1, f"准备输入余额:{input_value}")
            pay_input_button.send_keys(str(input_value))
            time_sleep(1, "已经输入 eth")
    except:
        print("没有遇到ETH不充足的情况")


    #准备交易
    add_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div[2]/div[1]/div[1]/div/div[2]/button')))
    time_sleep(1, "准备点击add")
    browser.execute_script("arguments[0].click();", add_button)
    time_sleep(10, "已经点击add")

    #切換到小狐狸,准备确认
    switch_tab_by_handle(browser, 1, 1)
    fox_statue = fox_confirm_goerli_transfer(browser, wait)
    if "失败" in fox_statue:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号小狐狸签名失败了, 已经记录下来")

    elif "成功" in fox_statue:
        # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
        print("小狐狸签名成功!! ")
    time_sleep(15)

    #切換到网站, 查看是否成功
    switch_tab_by_handle(browser, 2, 0)
    time_sleep(15,"等一会网页")

    web_statue = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div/div[1]/div[2]/div/p')))
    a = web_statue.text   #Balance : 0.99909 = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[2]/div[2]/p')))
    print(a)
    if "Add" not in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号失败了,已经记录下来")

    elif "Add" in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "成功")
        print("成功!! 已经记录到excel")

#任务4:解除流动性
def syncswap_remove_LP(browser, wait, excel_path,excel_row, write_excel_column, mode): #后面两个参数用于记录信息到excel
    #新建标签页,准备转goerli
    new_tab(browser, sync_swap_remove)
    time_sleep(5,"等待网络加载")
    switch_tab_by_handle(browser, 2, 0)
    try:
        path = "//button[@class='MuiButton-root MuiButton-outlined MuiButton-outlinedPrimary MuiButton-sizeMedium MuiButton-outlinedSizeMedium MuiButtonBase-root  css-1wwmuyk']"
        wallet_button = wait.until(EC.element_to_be_clickable((By.XPATH, path)))
        print("找到的文字是:",wallet_button.text)
    
        if "Connect" in wallet_button.text:#如果没有链接钱包,则链接钱包
            #连接小狐狸钱包
            time_sleep(1, "准备点击连接钱包")
            browser.execute_script("arguments[0].click();", wallet_button)
            time_sleep(5, "等等小狐狸出现")
            fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div[2]/div[2]/div[1]')))
            time_sleep(1, "小狐狸出现了,准备点击")
            browser.execute_script("arguments[0].click();", fox_button)
            time_sleep(10, "已经点击小狐狸")
    except:
        print("没有找到钱包按钮, 可能是不要链接钱包?")
    
    time_sleep(2, "准备重新打开网站")
    browser.get(sync_swap_remove)
    time_sleep(8,"等待网络加载")
    switch_tab_by_handle(browser, 2, 0)

    # 移除某个比例的流动性
    if mode == 0:
        print("模式0, 移除随机比例的流动性 (25~100%)")
        percent = random.randint(1, 4)  # ETH时，小数点最起码要有2位，因为L1的金额一般是两位小数以上
        percent_button = wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="container"]/div/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div/div[1]/div[2]/div/div/div/button[{percent}]')))
        time_sleep(2, "准备点击比例")
        browser.execute_script("arguments[0].click();", percent_button)
        print("已经点击转出比例")

    elif mode == 1:
        remove_max_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div/div[1]/div[2]/div/div/div/button[4]')))
        time_sleep(1)
        browser.execute_script("arguments[0].click();", remove_max_button)
        print("已经点击全部转出")   

    try:#签名允许解除流动性
        allow_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/button')))
        if "Sign" in allow_button.text:
            time_sleep(1,"确实要授权")
            browser.execute_script("arguments[0].click();", allow_button)
            time_sleep(3, "已经点击授权")

            #切換到小狐狸,准备确认
            switch_tab_by_handle(browser, 1, 1)
            fox_statue = fox_allow_syncswap_use_LP(browser, wait)
            if "失败" in fox_statue:
                Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
                print("这个号小狐狸签名失败了, 已经记录下来")

            elif "成功" in fox_statue:
                # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
                print("小狐狸签名成功!! ")
            time_sleep(10)

    except:
        print("可能是已经授权")

    #准备交易
    switch_tab_by_handle(browser, 2, 0)
    remove_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div[2]/div/div/div/div[2]/div[1]/div/button')))
    time_sleep(1, "准备点击 remove")
    browser.execute_script("arguments[0].click();", remove_button)
    time_sleep(3, "已经点击 remove")

    #切換到小狐狸,准备确认
    switch_tab_by_handle(browser, 1, 1)
    fox_statue = fox_confirm_goerli_transfer(browser, wait)
    if "失败" in fox_statue:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号小狐狸签名失败了, 已经记录下来")

    elif "成功" in fox_statue:
        # Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, excel_column, "成功")
        print("小狐狸签名成功!! ")
    time_sleep(10)

    #切換到网站, 查看是否成功
    switch_tab_by_handle(browser, 2, 0)
    time_sleep(10,"等一会网页")

    web_statue = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[2]/div/div/div/div[1]/div[2]/div/p')))
    a = web_statue.text   #Balance : 0.99909 = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="container"]/div/div/div[5]/div/div/div/div[2]/div[1]/div[2]/div[2]/p')))
    print(a)
    if "Remove" not in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "×")
        print("这个号失败了,已经记录下来")

    elif "Remove" in a:
        Do_Excel(excel_path, sheetname="Sheet1").plain_write(excel_row, write_excel_column, "成功")
        print("成功!! 已经记录到excel")

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  syncswap的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  Galaxy领取NFT上的项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

Odyssey_url = "https://galaxy.eco/arbitrum/campaign/GCCNzUtQiW/"  # 2022,8,5,领取odessey NFT
galaxy_op_hop_NFT = "https://galaxy.eco/HopProtocol/campaign/GCmydUtumN/" #2022，8，11，领取op的hop NFT

# 领取ORB的奥德赛 NFT
def galaxy_claim_orb_odyssey_NFT(browser, wait):
    time_sleep(2, "准备打开 orb ")
    new_tab(browser, Odyssey_url)
    time_sleep(10, "正在打开 orb ")
    switch_tab_by_handle(browser, 2, 0)  # 切换到被撸网站
    time_sleep(6, "waiting")
    browser.refresh()
    time_sleep(6, "waiting")
    browser.refresh()
    time_sleep(6, "waiting")
    browser.refresh()
    time_sleep(6, "waiting")  # 多刷新几次，防止说不能cliam

    # 连接小狐狸
    # try:
    #     connect_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="topNavbar"]/div/div[2]/div[2]/div[1]/div[2]/button/span')))
    #     if "Connect Wallet" in connect_button.text:
    #         time_sleep(2)
    #         browser.execute_script("arguments[0].click();", connect_button)

    #         fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[3]/div/div/div/div[2]/div[2]/div')))
    #         time_sleep(2)
    #         browser.execute_script("arguments[0].click();", fox_button)
    # except:
    #     print("可能已经连接了小狐狸")
    try:
        claim_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                              '//*[@id="app"]/div/main/div/div/div/div/div[1]/div[1]/div[2]/div[2]/div/div[1]/div/div/button/span')))
        time_sleep(7, "button found")
        browser.execute_script("arguments[0].click();", claim_button)
        time_sleep(15)
        switch_tab_by_handle(browser, 1, 1)  # 切换到被撸网站
        fox_info = fox_confirm_galaxy(browser, wait)
        time_sleep(20)
        switch_tab_by_handle(browser, 2, 0)  # 切换到被撸网站

        if "成功" in fox_info:
            try:
                submitted_button = wait.until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[3]/div/div/div/div[1]')))
                if "submitted" in submitted_button.text:
                    print("领取成功!")
                    return "领取成功"
            except:  # 小狐狸虽然确认成功了，但实际没有cliam成功
                return "失败,小狐狸虽然确认成功了，但实际没有cliam成功"
        else:
            print("这个号领不了NFT")
            return "失败"
    except:
        print("这个号领不了NFT")
        return "失败"

#由于这个任务不需要小狐狸确认，所以通过是否已经claimed 来判断是否领取成功
def galaxy_claim_op_hop_NFT(browser, wait):
    time_sleep(2, "准备打开 orb ")
    new_tab(browser, galaxy_op_hop_NFT)
    time_sleep(30, "正在打开 orb ")
    # switch_tab_by_handle(browser, 1, 0)  # Mac下调试，切换到被撸网站
    switch_tab_by_handle(browser, 2, 0)  # 切换到被撸网站
    # time_sleep(6, "waiting")
    # browser.refresh()
    # time_sleep(6, "waiting")
    # browser.refresh()
    # time_sleep(6, "waiting")
    # browser.refresh()
    # time_sleep(6, "waiting")  # 多刷新几次，防止说不能cliam

    # 连接小狐狸
    # try:
    #     connect_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="topNavbar"]/div/div[2]/div[2]/div[1]/div[2]/button/span')))
    #     if "Connect Wallet" in connect_button.text:
    #         time_sleep(2)
    #         browser.execute_script("arguments[0].click();", connect_button)

    #         fox_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[3]/div/div/div/div[2]/div[2]/div')))
    #         time_sleep(2)
    #         browser.execute_script("arguments[0].click();", fox_button)
    # except:
    #     print("可能已经连接了小狐狸")
    try:
        #==================  查看有多少个数量可以领
        available_num_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[@class='text-base ml-1 text-bold']")))
        claimed_num_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='text-base ml-1 mr-3 text-bold']")))
        if claimed_num_button.text != "0": #说明这个号已经领取成功
            print("这个号领取成功!")
            return "领取成功"
        else:
            if available_num_button.text != "0": #说明这个号确实可以领，但还没有领
                print("找到的available数量是：",available_num_button.text)
                for i in range(1,5):
                    print(f"开始第{i}次领取")
                    claim_button = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="app"]/div/main/div/div/div/div/div[1]/div[1]/div[2]/div[2]/div/div[1]/div/div/button/span')))
                    time_sleep(7, "claim button found")
                    browser.execute_script("arguments[0].click();", claim_button)

                    time_sleep(60,"60秒后查看是否到账")
                    claimed_num_button = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='text-base ml-1 mr-3 text-bold']")))
                    if claimed_num_button.text == "1":  # 说明这个号已经领取成功
                        print("领取成功!")
                        return "领取成功"
                    else:
                        time_sleep(2, "已经点击领取，但60秒后还未到账")
                        browser.refresh()
                        time_sleep(20, "刷新浏览器后继续")
                        if i == 4:
                            print("尝试领取4次，依旧失败")
                            return "失败"
                        continue
            else:
                print("这个号领不了NFT")
                return "失败"
    except:
        print("这个号领不了NFT")
        browser.quit()
        return "失败"


## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   uniswap的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #



## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  uniswap的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ OP项目 的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

####一些url
lifi_url = "https://transferto.xyz/swap"
ARB_url = "https://bridge.arbitrum.io/"
ZK_url = "https://wallet.zksync.io/transaction/deposit"
OP_url = "https://app.optimism.io/bridge"
test_url = "https://www.baidu.com/"

#==OP上的项目
op_pika_url = "https://app.pikaprotocol.com/trade/ETH-USD"
op_showme_url = "https://optimismair.showme.fan/"
op_zipswap_url = "https://zipswap.fi/#/swap"
op_clipper_url = "https://clipper.exchange/app/swap"
op_uniswap_url = "https://app.uniswap.org/#/swap?chain=optimism"
op_sushiswap_url = "https://app.sushi.com/swap?chainId=10"
op_curve_url = "https://optimism.curve.fi/"

matcha_url = "https://matcha.xyz/login"



#==================================下面是 OP 的任务
# from_token，可选"Ethereum"， "USD Coin"，"sUSD"
def OP_matcha(browser, wait, from_source, to_source):

    new_tab(browser, matcha_url)
    time_sleep(5, "等待抹茶加载")
    switch_tab_by_handle(browser, 2, 0)  # 切换到抹茶
    time_sleep(8, "等待抹茶加载")

    # 如果没有连接小狐狸钱包，则先去连接钱包
    time_sleep(3, "准备判断小狐狸是否连接钱包")
    if matcha_fox_icon_exit(browser, wait) != 1: # 返回值是1，说明有小狐狸图标，说明已经连接了小狐狸
        print("Match没有连接小狐狸，先去连接")
        matcha_connect_wallet(browser, wait)  # 先连接钱包
         
    time_sleep(2, "提醒：请记得先把抹茶上的网络切为 OP")
    if matcha_whether_change_net(browser, wait) != "Optimism":
        matcha_change_net(browser, wait)

    # 获取余额
    max_token, L2_balance = get_OP_token_balance_by_matcha(browser, wait)
    # 选择 from , to
    matcha_input_coin_amount_new(browser, wait, L2_balance, from_source, to_source)  #准备换钱

    print("切换到小狐狸")
    switch_tab_by_handle(browser, 1, 1)  
    fox_info = fox_confirm_OP(browser, wait)  # 小狐狸确认交易
    time_sleep(5)
    switch_tab_by_handle(browser, 2, 0)  # 切换到matcha
    time_sleep(3)

    # 转账后，20秒后，查询OP上的USDC、DAI是否到账
    time_sleep(20,"准备查询是否 matcha Success？")
    transfer_status = matcha_whether_transfer_success(browser, wait)

    if "成功" in fox_info:
        if transfer_status == "Success!":
            print("matcha 确实换钱 Success！")
            return "OP matcha 任务完成；" + fox_info
        else:
            return " matcha 还是换钱失败" 
    elif "失败" in fox_info:
        return fox_info

#做 OP上的 zipswap
def OP_zipswap(browser, wait, from_source, to_source):
    print("开始OP_zipswap任务")
    new_tab(browser, op_zipswap_url)
    time_sleep(15, "等待zipswap加载")
    switch_tab_by_handle(browser, 2, 0)  # 切换到zipswap

    #======连接钱包
    if zipswap_whether_connect_wallet(browser, wait) == "Connect Wallet":
        print("需要连接钱包")
        zipswap_connect_wallet(browser, wait)
        time.sleep(5)
        switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
        fox_confirm_connect_network(browser, wait)  # 小狐狸【批准】、【切换网络】
        switch_tab_by_handle(browser, 2, 0)  # 切换到第1个标签页：被撸网站

    #=========实时获取 OP上的ETH金额，返回浮点型数据
    # L2_ETH_value = get_OP_ETH_by_zipswap(wait)
    L2_max_value = get_OP_ETH_and_select_from_to_token_by_zipswap(browser, wait, from_source, to_source)

    #======选择代币、输入金额、确认交易， #from，to可选"ETH"， "USDC"，"USDT"
    detail = zipswap_prepare_transfer(browser, wait, L2_max_value, from_source, to_source)
    time_sleep(8)

    #========小狐狸确认
    switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
    fox_info = fox_confirm_OP(browser, wait)

    #关闭zipswap
    switch_tab_by_handle(browser, 2, 0)  # 切换到第
    time_sleep(5)

    #=======判断是否成功
    time_sleep(15, "准备检查clipper是否成功")
    try:
        success_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='sc-1jvskkm-2 flClZu']/div")))
    except:
        success_info = "失败"
        print("没有成功")

    if "Submitted" in success_info.text:
        print("zigswap 成功")
        return " OP zigswap 任务成功" + detail + fox_info
    else:
        return "失败" + detail + fox_info

#做 OP上的 clipper
def OP_clipper(browser, wait, from_source, to_source):
    print("开始OP_clipper任务")
    new_tab(browser, op_clipper_url)
    time_sleep(15, "等待 clipper 加载")
    switch_tab_by_handle(browser, 2, 0)
    time_sleep(3)

    #连接钱包
    if clipper_whether_connect_wallet(browser, wait) == "Connect your wallet":
        print("需要连接钱包")
        clipper_connect_wallet(browser, wait)
        time_sleep(5)
 
    #============选择代币、输入金额、确认交易，返回浮点数
    # L2_ETH_value = get_OP_ETH_by_clipper(wait)
    L2_max_value = get_OP_ETH_and_select_from_to_token_by_clipper(browser, wait, from_source, to_source)

    detail = clipper_prepare_transfer(browser, wait, L2_max_value, from_source, to_source)
    time_sleep(3)

    #=======小狐狸确认交易
    switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
    info = fox_confirm_OP(browser, wait)
    time_sleep(5)
    switch_tab_by_handle(browser, 2, 0)  # 切换到
    
    #=======判断是否成功
    time_sleep(15, "准备检查clipper是否成功")
    try:
        success_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='success_notification']//h2")))
    except:
        success_info = "失败"
        print("没有成功")

    if "Submitted" in success_info.text:
        return " OP clipper 任务成功"  
    else:
        return "失败" 

#做 OP 上的 uniswap
def OP_uniswap(browser, wait, from_source, to_source):
    #========== 小狐狸网络切换成OP
    # switch_tab_by_handle(browser, 1, 0)
    # fox_change_network(browser, wait, "Optimism")
    # time_sleep(5)
    major_token = "" #记录经过swap后的主要代币是什么
    print("开始OP_uniswap任务")
    new_tab(browser, op_uniswap_url)
    time_sleep(5, "等待 uniswao 加载")
    switch_tab_by_handle(browser, 2, 0) 
    time_sleep(10, "等待 uniswao 加载")

    # 由于登录就有小狐狸帐号,所以不要重新连接小狐狸
    # 搜索from和to
    #先点击from的下拉列表
    from_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='swap-currency-input']//button/span")))
    time_sleep(3, "from_list button found")
    browser.execute_script("arguments[0].click();", from_list)

    from_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(3)
    from_button.send_keys(from_source)
    time_sleep(3,"再发送回车确认")
    from_button.send_keys(Keys.ENTER)

    #  =======找余额. 放在上面步骤from之下, 这样可以避免和下面to的余额冲突
    time_sleep(5,"准备寻找金额")
    L2_ETH_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='sc-18nh1jk-0 horBTe css-1lpcios']")))
    a = L2_ETH_button.text #类似 ---> 余额： 0.009156
    b = a.split()[-1]
    print("提取到的纯金额是: ", b)
    L2_ETH_value = float(b)

    #再点击to的下拉列表
    to_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='swap-currency-output']//button/span")))
    time_sleep(3, "to_list button found")
    browser.execute_script("arguments[0].click();", to_list)

    to_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(3)
    to_button.send_keys(to_source)
    time_sleep(3,"再发送回车确认")
    to_button.send_keys(Keys.ENTER)
    time_sleep(5,"等待响应")
    
    # 输入金额
    #法一: 选择最大金额. 但用ETH作为from时有bug:点击max, 但实际没有max
    if from_source != "ETH":
        max_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()[contains(.,'Max')]]")))
        time_sleep(3, "max_button button found")
        browser.execute_script("arguments[0].click();", max_button)
        time_sleep(8, "已经点击max,等待响应")
    else: #法二:选择比例法
        # =======输入金额框
        input_amount_box = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='swap-currency-input']//input")))
        time_sleep(3,"找金额输入框")
        
        # ====随机比例
        point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_ETH_value * 0.6, L2_ETH_value * 0.8), point)
        try_times = 1
        while float(L2_ETH_value) - input_value < 0.005: #如果会使余额小于 0.005
            point = random.randint(3, 4)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
            input_value = round(random.uniform(L2_ETH_value * 0.2, L2_ETH_value * 0.7), point)
            try_times = try_times + 1
            if try_times == 100:
                print("余额不足，直接报错")
                return 0  #说明余额不做，直接报错即可
        input_amount_box.send_keys(str(input_value))
        time_sleep(8, "已经输入金额,等待响应")

    #准备确认交易. 可能会出现 allow 
    if from_source != "ETH":
        try:
            uniswap_allow_token(browser, wait)
        except:
            print("可能是之前allow 过了,不需要重新授权")
    
    #点击 swap 按钮
    time_sleep(5,"准备查看swap按钮")
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id = 'swap-button']")))
    time_sleep(3, "swap button button found")
    browser.execute_script("arguments[0].click();", swap_button)

    #点击confirm swap按钮
    flag = 0
    while flag == 0:
        try:
            confirm_swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id= 'confirm-swap-or-send']")))
            time_sleep(3, "confirm_swap_button button found")
            browser.execute_script("arguments[0].click();", confirm_swap_button)
            flag = 1 #说明确实点击到了confirm_swap
            break
        except:
            #尝试点击accept按钮,有时价格会更新
            accept_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()= 'Accept']")))
            time_sleep(3, "accept_button button found")
            browser.execute_script("arguments[0].click();", accept_button)
            continue
    #切换到小狐狸确认
    switch_tab_by_handle(browser, 1, 1) 
    fox_confirm_allow_token(browser, wait)

    #切换回被鲁网站,查看是否成功
    time_sleep(10,"准备查看是否成功")
    switch_tab_by_handle(browser, 2, 0) 
    time_sleep(10,"准备查看是否成功")
    try:
        success_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()= 'Success']")))
        return "成功"
    except:
        return "失败"
    
#mode. 1-->最大化ETH模式    0-->比例化USDC等代币
def OP_sushiswap(browser, wait, from_source, to_source, mode):

    print("开始 OP_sushiswap 任务")
    new_tab(browser, op_sushiswap_url)
    time_sleep(5, "等待 sushiswap 加载")
    switch_tab_by_handle(browser, 2, 0) 
    time_sleep(10, "等待 sushiswap 加载")

    # 由于登录就有小狐狸帐号,所以不要重新连接小狐狸
    # 搜索from和to
    #先点击from的下拉列表
    from_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='asset-select-trigger-0']")))
    time_sleep(3, "from_list button found")
    browser.execute_script("arguments[0].click();", from_list)

    from_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(3)
    from_button.send_keys(from_source)
    time_sleep(3,"再发送回车确认")
    from_button.send_keys(Keys.ENTER)


    #再点击to的下拉列表
    to_list = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='asset-select-trigger-1']")))
    time_sleep(3, "to_list button found")
    browser.execute_script("arguments[0].click();", to_list)

    to_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='token-search-input']")))
    time_sleep(3)
    to_button.send_keys(to_source)
    time_sleep(3,"再发送回车确认")
    to_button.send_keys(Keys.ENTER)
    time_sleep(5,"等待响应")
    
    # 输入金额.
    if mode == "max_all":
        #法一: 选择最大金额. 但用ETH作为from时有bug:点击max, 但实际没有max
        print("本次采用最大化金额法")
        max_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-col gap-3']/div[1]//div[text()[contains(.,'Balance')]]")))
        time_sleep(3, "max_button button found")
        browser.execute_script("arguments[0].click();", max_button)
        time_sleep(8, "已经点击max,等待响应")
    
    elif mode == "save_some_token":
        #法二: 部分金额法.比如源是OP, 保留3~5个OP代币
        print("本次采用比例化金额法")
        time_sleep(3, "提取金额....")
        balance_box_path = "//div[@class='relative filter z-10']/div[1]/div[@class='flex flex-col gap-3']/div[1]//div[@class='text-sm leading-5 font-medium cursor-pointer select-none flex text-secondary whitespace-nowrap']"
        balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, balance_box_path)))
        balance_string = balance_button.text #类似 Balance:13.3942
        balance_a = balance_string.split()[-1]
        L2_max_value = float(balance_a)
        print("恭喜提取到的金额是:", L2_max_value)

        # =======找输入金额框
        time_sleep(3,"准备找输入金额框")
        input_amount_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='relative filter z-10']/div[1]/div[@class='flex flex-col gap-3']/div[1]//input")))
        time_sleep(2,"已经找到了输入框")

        # 随机金额. 留取3~5个OP币
        point = random.randint(1, 2)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(L2_max_value - 5, L2_max_value -3), point)
        print(f"本次从 {from_source} 转到 {to_source} 的随机金额是{input_value}，将来预估余额是：{float(L2_max_value) - input_value}")
        input_amount_box.send_keys(str(input_value))
        time_sleep(10,"已经输入金额")

    elif mode == "buy_some_token":#只买指定数量的目标币,如5个OP, 则输入0.001 ETH
        # =======找输入金额框
        time_sleep(3,"准备找输入金额框")
        input_amount_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='relative filter z-10']/div[1]/div[@class='flex flex-col gap-3']/div[1]//input")))
        time_sleep(2,"已经找到了输入框")

        # 随机金额. 
        point = random.randint(3, 5)  # 最起码保留2位小数，因为L1的ETH范围是0.05~0.08
        input_value = round(random.uniform(0.001, 0.003), point)
        print(f"本次从 {from_source} 转到 {to_source} 的随机金额是{input_value}，将来预估余额是：{float(L2_max_value) - input_value}")
        input_amount_box.send_keys(str(input_value))
        time_sleep(10,"已经输入金额")

    #准备确认交易. from 不是ETH时可能会出现 Approve BentoBox , Approve USDC
    if from_source != "ETH":
        try:
            #查看是哪一种allow
            allow_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-col gap-3']/button[1]")))
            if "BentoBox" in allow_button.text:
                sushiswap_allow_Bentobox(browser, wait)

        except:
            print("可能是之前allow 过了,不需要重新授权")
    if from_source != "ETH":
        try:
            #查看是哪一种allow
            allow_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-col gap-3']/button[1]")))

            if "Approve" in allow_button.text:
                #无论如何都要试一试使用USDC
                sushiswap_allow_token(browser, wait)
        except:
            print("可能是之前allow 过了,不需要重新授权")
            
    else: #如果是用ETH转出去,可能会有一个 Bentobox
        try:
            sushiswap_allow_Bentobox(browser, wait)
        except:
            print("可能是不需要授权Bentobox")
    
    # 可能出现swap anyway
    try: #
        print("尝试 swap anyway")
        swap_anyway_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()[contains(.,'Swap Anyway')]]")))
        print(swap_anyway_button.text)
        browser.quit()
        return "失败流动性不足, 取消交易"
    except:
        print("可能不要swap anyway")
        time_sleep(3, "waiting")

    #点击 swap 按钮
    time_sleep(5,"准备查看swap按钮")
    swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id = 'swap-button']")))
    time_sleep(3, "swap button button found")
    browser.execute_script("arguments[0].click();", swap_button)
    time_sleep(10, "已经点击 swap button")

    
    #点击confirm swap按钮
    confirm_swap_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id= 'confirm-swap-or-send']")))
    time_sleep(3, "confirm_swap_button button found")
    browser.execute_script("arguments[0].click();", confirm_swap_button)
    time_sleep(5, "已经点击 confirm swap button")


    #切换到小狐狸确认
    switch_tab_by_handle(browser, 1, 1) 
    fox_confirm_allow_token(browser, wait)

    #切换回被鲁网站,查看是否成功
    time_sleep(10,"准备查看是否成功")
    switch_tab_by_handle(browser, 2, 0) 
    time_sleep(10,"准备查看是否成功")
    try:
        success_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()= 'Transaction submitted']")))
        return "成功"
    except:
        return "失败"
 


## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  uniswap 项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
#允许 uniswap是使用token
def uniswap_allow_token(browser, wait):
    #点击allow
    allow_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()[contains(.,'Allow')]]")))
    time_sleep(3, "allow_button button found")
    browser.execute_script("arguments[0].click();", allow_button)

    #切换到小狐狸
    switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
    fox_confirm_allow_token(browser, wait)
    time_sleep(2)
    switch_tab_by_handle(browser, 2, 0)  # 切换到第0个标签页：小狐狸
    time_sleep(2)


## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  uniswap 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  sushiswap 项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
#允许 sushiswap 使用token
def sushiswap_allow_token(browser, wait):
    #点击allow. 这里必须用 //div[@class='flex flex-col gap-3']/button[1] 这种方法.因为如果用id的话,可能会有两个
    allow_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-col gap-3']/button[1]")))
    time_sleep(3, "allow button found")
    browser.execute_script("arguments[0].click();", allow_button)

    #切换到小狐狸
    switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
    fox_confirm_allow_token(browser, wait)
    time_sleep(2)
    switch_tab_by_handle(browser, 2, 0)  # 切换到第0个标签页
    time_sleep(3)

#允许 sushiswap 使用 Bentobox
def sushiswap_allow_Bentobox(browser, wait):
    #点击allow
    allow_Bentobox_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='flex flex-col gap-3']/button[1]")))
    time_sleep(3, "allow Bentobox button found")
    browser.execute_script("arguments[0].click();", allow_Bentobox_button)

    #切换到小狐
    switch_tab_by_handle(browser, 1, 1)  # 切换到第0个标签页：小狐狸
    fox_confirm_allow_bentobox(browser, wait)
    time_sleep(2)
    switch_tab_by_handle(browser, 2, 0)  # 切换到第0个标签页
    time_sleep(3)
## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  sushiswap 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #




## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  debank 项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

debank_url = "https://debank.com/"

def debank_whether_connect_wallet(browser, wait):
    print("开始 debank_whether_connect_wallet 任务")
    try:
        login_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Log in')]]")))
        time_sleep(3, "login button found")
        browser.execute_script("arguments[0].click();", login_button)

        fox_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='WalletSelectorModal_walletList__1R_dC']/div[2]/button")))
        time_sleep(3, "fox button found")
        browser.execute_script("arguments[0].click();", fox_button)
    except:
        print("可能是不需要链接小狐狸")
        


#network_name可选:Ethereum, Optimism, Arbitrum等
def get_balance_from_debank(browser, wait, network_name):
    print("开始 get_balance_from_debank 任务")
    new_tab(browser, debank_url)
    time_sleep(15, "等待 debank 加载")
    switch_tab_by_handle(browser, 2, 0)  # 一定要切换!!切换到

    #链接小狐狸
    debank_whether_connect_wallet(browser, wait)
    
    #选择需要查看的网络
    xpath_path = f"//div[text()[contains(.,'{network_name}')]]"
    click_network = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_path)))
    time_sleep(3, "找到了要查看余额的网络,即将点击")
    browser.execute_script("arguments[0].click();", click_network)
    time_sleep(5,"等待切换到对应网络")
   
    # 查看有多少个代币
    coin_list = wait.until(EC.presence_of_all_elements_located((By.XPATH,"//div[@class='db-table-body']/div")))
    print("总共有多少个代币:", len(coin_list))

    # #获取代币余额
    balance_dict = {} #创建一个空字典,用于装余额
    for i in range(1, len(coin_list) + 1):
        xpath_way = f"//div[@class='db-table-body']/div[{i}]//div[@class='db-table-cell']/div/div[2]"
        coin_name_button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_way)))
        coin_name = coin_name_button.text
        print(f"第{i}个代币名字是: ", coin_name)

        #余额的路径
        xpath_way = f"//div[@class='db-table-body']/div[{i}]/div[@class='db-table-row']/div[4]"
        balance_button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_way)))
        a = balance_button.text # 形如 $4
        coin_balance = a.split("$")[-1]
        print(f"第{i}个代币的余额是: ", coin_balance)
        balance_dict[coin_name] = int(coin_balance) #添加到字典里
    print("所有代币及金额是:",balance_dict)
    max_token = max(balance_dict,key=balance_dict.get)
    print("金额最多的代币是:", max_token)
    browser.close() #关闭这个 debank 网页
    time_sleep(3,"已经关闭了网页")
    switch_tab_by_handle(browser, 1, 0)  # 一定要切换回去,给selenium一个指针,否则无法新建标签页.
    time_sleep(3,"准备打开新网页")

    return max_token, balance_dict

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  debank的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓  Alchemy 项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

#Alchemy填写项目描述
def fill_in_alchemy_project_des(browser, wait):
    print("开始填写alchemy项目描述, 刚开始注册帐号的时候需要")
    des1_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='My Team']")))
    time_sleep(2,"准备输入描述1")
    des1_button.send_keys(fake.sentence())

    next_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Next']")))
    time_sleep(2,"准备点击Next")
    browser.execute_script("arguments[0].click();", next_button)

    des2_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='...']")))
    time_sleep(2,"准备输入描述2")
    des2_button.send_keys("NFT")

    next_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Next']")))
    time_sleep(2,"准备点击Next")
    browser.execute_script("arguments[0].click();", next_button)

    Ethereum_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()='Ethereum']")))
    time_sleep(2,"准备选择Ethereum")
    browser.execute_script("arguments[0].click();", Ethereum_button)

    next_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Next']")))
    time_sleep(2,"准备点击Next")
    browser.execute_script("arguments[0].click();", next_button)

    free_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()='Free']")))
    time_sleep(2,"准备选择Free")
    browser.execute_script("arguments[0].click();", free_button)

    next_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Next']")))
    time_sleep(2,"准备点击Next")
    browser.execute_script("arguments[0].click();", next_button)

    #跳过信用卡
    skip_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Skip for now']")))
    time_sleep(2,"准备选择Free")
    browser.execute_script("arguments[0].click();", skip_button)

    build_button = wait.until(EC.element_to_be_clickable((By.XPATH,'''//button[text()="Let's Build!"]''')))
    time_sleep(2,"准备选择 build ")
    browser.execute_script("arguments[0].click();", build_button)
    time_sleep(20,"已经点击 build ")

#Alchemy创建rinkeby项目
def alchemy_create_rinkeby_app(browser, wait):
    print("开始创建rinkeby项目")
    create_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Create')]]")))
    time_sleep(2,"准备创建")
    browser.execute_script("arguments[0].click();", create_button)

    name_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='name']")))
    time_sleep(2,"准备输入name")
    app_name = fake.last_name()
    name_button.send_keys(app_name)

    desc_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='description']")))
    time_sleep(2,"准备输入desc")
    desc_button.send_keys(fake.first_name())

    #下拉列表
    try:
        down_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='select-input-container css-1tyu61v']//div[@class='css-1hkumgc']/span")))
        time_sleep(2,"准备点击下拉")
        # browser.execute_script("arguments[0].click();", down_list_button)
        ActionChains(browser).click(down_list_button).perform()  # 必须用模拟鼠标点
    except:
        print("下拉没找到哦啊")

    # #选择rinkeby
    rinkeby_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='css-1hkumgc']/span[text()[contains(.,'Rinkeby')]]")))
    time_sleep(2,"准备选择 rinkeby")
    browser.execute_script("arguments[0].click();", rinkeby_button)
    # ActionChains(browser).click(rinkeby_button).perform()  # 模拟鼠标点

    #确定创建
    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']")))
    time_sleep(2,"准备点击创建")
    browser.execute_script("arguments[0].click();", confirm_login)
    time_sleep(20,"已经点击创建")
    return app_name

#Alchemy创建goerli项目
def alchemy_create_goerli_app(browser, wait):
    print("开始创建goerli项目")
    create_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Create')]]")))
    time_sleep(2,"准备创建")
    browser.execute_script("arguments[0].click();", create_button)

    name_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='name']")))
    time_sleep(2,"准备输入name")
    app_name = fake.last_name()
    name_button.send_keys(app_name)

    desc_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='description']")))
    time_sleep(2,"准备输入desc")
    desc_button.send_keys(fake.first_name())

    #下拉列表
    try:
        down_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='select-input-container css-1tyu61v']//div[@class='css-1hkumgc']/span")))
        time_sleep(2,"准备点击下拉")
        # browser.execute_script("arguments[0].click();", down_list_button)
        ActionChains(browser).click(down_list_button).perform()  # 必须用模拟鼠标点
    except:
        print("下拉没找到哦啊")

    # #选择rinkeby
    rinkeby_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='css-1hkumgc']/span[text()[contains(.,'Goerli')]]")))
    time_sleep(2,"准备选择 goerli ")
    browser.execute_script("arguments[0].click();", rinkeby_button)
    # ActionChains(browser).click(rinkeby_button).perform()  # 模拟鼠标点

    #确定创建
    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']")))
    time_sleep(2,"准备点击创建")
    browser.execute_script("arguments[0].click();", confirm_login)
    time_sleep(20,"已经点击创建")
    return app_name

#Alchemy创建goerli项目
def alchemy_create_Mainnet_app(browser, wait):
    print("开始创建 Mainnet 项目")
    create_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Create')]]")))
    time_sleep(2,"准备创建")
    browser.execute_script("arguments[0].click();", create_button)

    name_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='name']")))
    time_sleep(2,"准备输入name")
    app_name = fake.last_name()
    name_button.send_keys(app_name)

    desc_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='description']")))
    time_sleep(2,"准备输入desc")
    desc_button.send_keys(fake.first_name())

    # #下拉列表
    # try:
    #     down_list_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='select-input-container css-1tyu61v']//div[@class='css-1hkumgc']/span")))
    #     time_sleep(2,"准备点击下拉")
    #     # browser.execute_script("arguments[0].click();", down_list_button)
    #     ActionChains(browser).click(down_list_button).perform()  # 必须用模拟鼠标点
    # except:
    #     print("下拉没找到哦啊")

    # # #选择rinkeby
    # rinkeby_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='css-1hkumgc']/span[text()[contains(.,'Goerli')]]")))
    # time_sleep(2,"准备选择 goerli ")
    # browser.execute_script("arguments[0].click();", rinkeby_button)
    # # ActionChains(browser).click(rinkeby_button).perform()  # 模拟鼠标点

    #确定创建
    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']")))
    time_sleep(2,"准备点击创建")
    browser.execute_script("arguments[0].click();", confirm_login)
    time_sleep(20,"已经点击创建")
    return app_name


#注册Alchemy时,填写随机信息
def signup_alchemy_random_info(browser, wait, email_account, email_pw):
    first_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Gavin']")))
    time_sleep(2,"准备输入姓")
    first_name.send_keys(fake.last_name())
    time.sleep(2)

    second_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Belson']")))
    time_sleep(2,"准备输入名")
    second_name.send_keys(fake.first_name())

    email_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='gavin@hooli.com']")))
    time_sleep(2,"准备输入邮箱")
    email_button.send_keys(email_account)

    pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='••••••••']")))
    time_sleep(2,"准备输入密码")
    pw_button.send_keys(email_pw)

    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//form/button[text()='Sign up']")))
    time_sleep(2,"准备点击登录")
    browser.execute_script("arguments[0].click();", confirm_login)
    time_sleep(15,"等待查看resent按钮")

    #=========循环查看是不是有resent按钮
    check_resent_flag = True
    active_email_flag = False #初始定义不要去邮箱检查激活链接
    try_times = 0
    while check_resent_flag:
        try:
            try_times +=1  
            resent_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Resend email']")))
            print("找到了resent按钮, 待激活")
            check_resent_flag = False #不用再检查resent了
            active_email_flag = True #需要去激活邮件
            return active_email_flag
            
        except:
            time_sleep(6, "**********暂时还没有找到resent链接, 尝试再次点击sign up")
            try:#尝试再次点击login
                confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//form/button[text()='Sign up']")))
                time_sleep(2,"======准备再次点击登录")
                browser.execute_script("arguments[0].click();", confirm_login)
                print("===========已经再次点击sign up")
            except:
                print("===尝试再次点击sign up,失败!!!")

        if try_times == 50:
            check_resent_flag = False #不要再检查resent了
            browser.quit()
            time_sleep(5, "等待了5分钟, 没有找到resent按钮,可能是有验证, 记录到excel")

# 登录alchemy时,填写帐号密码
def alchemy_login(browser, wait, email_to_login, email_pw):
    email_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='gavin@hooli.com']")))
    time_sleep(2,"准备输入邮箱")
    email_button.send_keys(email_to_login)

    pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='••••••••']")))
    time_sleep(2,"准备输入密码")
    pw_button.send_keys(email_pw)

    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']")))
    time_sleep(2,"准备点击登录")
    # browser.execute_script("arguments[0].click();", confirm_login)
    ActionChains(browser).click(confirm_login).perform()  # 用模拟鼠标点
    # time_sleep(2,"再点一次")
    # ActionChains(browser).click(confirm_login).perform()  # 用模拟鼠标点

#官方限制最多5个app, 删除一些不必要的app
def alchemy_delete_app(browser, wait, keyword):
    #查看第一个app的名字
    try:
        first_app_name = wait.until(EC.element_to_be_clickable((By.XPATH,"//tbody[@class='table-body']/tr[1]/td[1]/div/a")))
        if first_app_name.text == "Demo App":
            time_sleep(2,"是官方app, 需要进去删除")
            browser.execute_script("arguments[0].click();", first_app_name)

            Delete_App = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Delete App')]]")))
            time_sleep(2,"准备点击删除")
            browser.execute_script("arguments[0].click();", Delete_App)
            #
            Delete_App_input = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Demo App']")))
            time_sleep(2,"找到输入框")
            Delete_App_input.send_keys("Demo App")

            Confirm_Delete_App = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Delete Forever')]]")))
            time_sleep(2,"准备确认删除")
            browser.execute_script("arguments[0].click();", Confirm_Delete_App)
            time_sleep(30,"已经确认")
    except:
        print("第一个app不是管方app")
    try:
        # 删除其他不必要的app, 先查看有多少个app
        app_nums = wait.until(EC.presence_of_all_elements_located((By.XPATH,"//tbody[@class='table-body']/tr")))
        num_result = len(list(app_nums))
        print("App个数：", num_result)
    except:
        print("没有找到app列表, 可能是没有app")
        num_result = 1
    #循环删除不要的app, 比如非 Rinkeby
    for i in range(1,num_result+1):
        try:
            network_xpath = f"//tbody[@class='table-body']/tr[{i}]/td[3]//span"
            App_network = wait.until(EC.element_to_be_clickable((By.XPATH, network_xpath)))
            
            view_detail_xpath = f"//tbody[@class='table-body']/tr[{i}]//span[text()='View Details']"
            view_detail_button = wait.until(EC.element_to_be_clickable((By.XPATH, view_detail_xpath)))

            app_net_name = App_network.text #查看App是在那个网络
            print("===找到的app网络是:", app_net_name)
            if app_net_name not in keyword:
                print(f"第{i}个 app 需要删除")
                time_sleep(2,f"需要删除非 {keyword} 的网络. ")
                browser.execute_script("arguments[0].click();", view_detail_button)
                
                security_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='button']/span")))
                time_sleep(2,"准备点击删除1")
                browser.execute_script("arguments[0].click();", security_button)

                Delete_App = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Delete App')]]")))
                time_sleep(2,"准备点击删除2")
                browser.execute_script("arguments[0].click();", Delete_App)

                Delete_App_input = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='input-container']//input[@type='text']")))
                input_text = Delete_App_input.get_attribute("placeholder")
                time_sleep(2,f"找到输入框,输入: {input_text}")
                Delete_App_input.send_keys(input_text)

                Confirm_Delete_App = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Delete Forever')]]")))
                time_sleep(2,"准备确认删除")
                browser.execute_script("arguments[0].click();", Confirm_Delete_App)
                time_sleep(30,"已经确认删除")
        except:
            time_sleep(1, "删除app出错了")
            Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()[contains(.,'Dashboard')]]")))
            time_sleep(2,"准备点击dashboard")
            browser.execute_script("arguments[0].click();", Dashboard_button)

#获取app的https 和 api等信息
def get_alchemy_app_info(browser, wait, app_name):
    print("开始去获取 app 的 https 和 api等信息")
    app_name_xpath = f"//tbody//a[text()[contains(.,'{app_name}')]]"
    App_name_button = wait.until(EC.element_to_be_clickable((By.XPATH, app_name_xpath)))    
    time_sleep(2, "找到了新建的app")
    browser.execute_script("arguments[0].click();", App_name_button)

    viw_key_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'View Key')]]")))
    time_sleep(2,"准备点击view key")
    browser.execute_script("arguments[0].click();", viw_key_button)

    api_info_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='app-url-widget m-3']/div[1]//input")))
    https_info_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='app-url-widget m-3']/div[2]//input")))
    api_info = api_info_button.get_attribute("value")
    https_info = https_info_button.get_attribute("value")
    print("api获取到的数据是", api_info)
    print("https_info获取到的数据是", https_info)

    return https_info, api_info

#从alchemy出领取测试币
def get_alchemy_faucet(browser, wait):
    #=================================== 领水
    app_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//tbody[@class='table-body']/tr[3]/td[1]/a")))
    time_sleep(2,"准备点击进入app, 打算去领水")
    browser.execute_script("arguments[0].click();", app_button)


    get_test_ETH = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[text()[contains(.,'Get Test ETH')]]")))
    time_sleep(2,"准备点击进入水龙头")
    browser.execute_script("arguments[0].click();", get_test_ETH)

    switch_tab_by_handle(browser, 2, 0)  # 切换到

    #============查看是不是要登录
    login_flag = True
    try_times = 0
    while login_flag:
        #去检查需不需要登录
        browser.refresh()
        time_sleep(10, "领水网页刷新")
        try:
            login_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()[contains(.,'Please signup or login')]]")))
            print("需要登录login")
            login_flag = True
        except:
            print("可能是已经登录了")
            login_flag = False
        #需要的话去登录
        if login_flag:
            try:
                try_times +=1
                login_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()[contains(.,'Alchemy Login')]]")))
                time_sleep(2,"准备登录")
                browser.execute_script("arguments[0].click();", login_button)
                time_sleep(20,"已经点击登录alchemy")

            except:
                browser.refresh()
                time_sleep(15, "登录失败, 重新登录")
        print(f"第{try_times}次领水尝试登录")        
        if try_times == 5:
            browser.quit()


    #=========循环领取
    faucet_flag = True
    try_times = 0
    while faucet_flag:
        browser.refresh()
        time_sleep(10, "进入领水, 等待刷新")
        #去检查需不需要领. 如果是空,说明需要领取
        try:
            blank_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='alchemy-faucet-table-data col']")))
            print("领取列表是空的,需要领水")
        except:
            print("可能是已经领到了")
            faucet_flag = False
        #需要的话去领
        if faucet_flag:
            try:
                try_times +=1
                address_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='address']")))
                time_sleep(2,"准备输入小狐狸帐号")
                address_button.send_keys(fox_address)

                confirm_send = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='button']//span")))
                time_sleep(2,"准备点击send")
                # browser.execute_script("arguments[0].click();", confirm_login)
                ActionChains(browser).click(confirm_send).perform()  # 用模拟鼠标点

                time_sleep(30,"已经点击send")
            except:
                browser.refresh()
                time_sleep(15, "领取失败, 重新领")
        print(f"第{try_times}次领水~~~")   
        if try_times == 5:
            browser.quit()

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  Alchemy 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   filebase的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
#=============== 修改json文件
def edit_json_file(json_path, pic_CID):
    whole_pic_CID = f'https://ipfs.filebase.io/ipfs/{pic_CID}'
    with open(json_path, 'r+') as f:
        # 读取demp.json文件内容
        data = json.load(f)
        # print(data)
        # print('================')
        # print(data["image"])

        # 修改CID的值
        data["image"] = whole_pic_CID
        data["name"] = RandomWords().get_random_word()  # 随机一个名字
        print("====修改 image 的链接为:", whole_pic_CID)
        # print('================')
        # print(data)
        with open(json_path, 'w') as f2:
            json.dump(data, f2)  # 写入f2文件到本地
            f2.close() #打开后需要关闭，否则文件无变化，导致CID一直一样
        time.sleep(2)
        f.close()
        time.sleep(2)

def login_filebase(browser, wait, filebase_email, filebase_pw):
    print("尝试登录")
    send_password = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='user_email']")))
    time_sleep(2,"准备输入用户名")
    send_password.send_keys(filebase_email)
    time.sleep(2)

    send_password = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='user_password']")))
    time_sleep(2,"准备输入密码")
    send_password.send_keys(filebase_pw)

    remember_me = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='user_remember_me']")))
    time_sleep(2,"准备点击记住我")
    browser.execute_script("arguments[0].click();", remember_me)


    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']")))
    time_sleep(2,"准备点击登录")
    browser.execute_script("arguments[0].click();", confirm_login)  


#=========在进入bucket后,上传图片文件,并获得图片CID
def filebase_upload_pic_file_in_bucket(browser, wait, pic_path):
    print("===上传图片")
    upload_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@id='menu-button']")))
    time_sleep(2,"准备点击upload")
    browser.execute_script("arguments[0].click();", upload_button)

    #==========选择file按钮
    file_button  = browser.find_element(By.XPATH, "//label[text()[contains(.,'File')]]//span/input")
    time_sleep(2,"准备选择file上传")
    file_button.send_keys(pic_path)

    time_sleep(20,"等待上传图片......")

    #获取图片的CID
    pic_CID = wait.until(EC.element_to_be_clickable((By.XPATH, "//tbody/tr[1]//span[@id='ipfs_cid']")))
    pic_CID_text = pic_CID.text 
    print("========获取到的图片CID是:", pic_CID_text)
    return pic_CID_text

#=======================上传json文件,并获得CID
def filebase_upload_json_file_in_bucket(browser, wait, json_path):
    print("===上传json")
    upload_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@id='menu-button']")))
    time_sleep(2,"准备点击upload")
    browser.execute_script("arguments[0].click();", upload_button)

    file_button  = browser.find_element(By.XPATH, "//label[text()[contains(.,'File')]]//span/input")
    time_sleep(2,"准备选择file上传")
    file_button.send_keys(json_path)
    time_sleep(20,"等待上传json文件......")

    # 获取json 的CID
    json_CID = wait.until(EC.element_to_be_clickable((By.XPATH, "//tbody/tr[2]//span[@id='ipfs_cid']")))
    json_CID_text = json_CID.text 
    print("============获取到json的CID是:", json_CID_text)
    return json_CID_text

def filebase_random_create_bucket_and_enter(browser, wait):
    print("===随机创建bucket并进入")
    create_bucket = wait.until(EC.element_to_be_clickable((By.XPATH,"//span//button[text()[contains(.,'Create Bucket')]]")))
    time_sleep(2,"准备创建bucket")
    browser.execute_script("arguments[0].click();", create_bucket)

    c = fake.md5()
    print("bucket 名字是:",c)

    send_bucket_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='bucket_name']")))
    time_sleep(2,"准备输入bucket name")
    send_bucket_name.send_keys(c)
    time.sleep(2)

    #创建bucket
    confirm_bucket_name = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@id='save_bucket']")))
    time_sleep(2,"准备创建bucket")
    browser.execute_script("arguments[0].click();", confirm_bucket_name)
    time_sleep(10, "已经点击创建bucket")

    #======================进入bucket
    # c = 'cd1183'
    login_bucket = wait.until(EC.element_to_be_clickable((By.XPATH,f"//p[text()[contains(.,'{c}')]]")))
    time_sleep(2,"准备进入bucket")
    browser.execute_script("arguments[0].click();", login_bucket)
    time_sleep(10, "已经点进进入bucket")

def filebase_return_to_bucket(browser, wait):
    #=============回到bucket首页
    try:
        back_to_bucket = wait.until(EC.element_to_be_clickable((By.XPATH,"//nav/a[text()[contains(.,' Buckets')]]")))
        time_sleep(2,"准备回到bucket")
        browser.execute_script("arguments[0].click();", back_to_bucket)
        print("已经点击回到bucket")
    except:
        print("程序一开始, 点击返回bucket失败")

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑  Replit 的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   Replit项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

#找 Replit 的注册激活的邮件 id
def cuiqiu_find_replit_activate_email_id(email_to_be_activate, email_from, email_subject):
    #循环检索邮箱. email_to_be_activate 表示待激活的邮箱
    not_find_yet = True
    try_times = 0
    while not_find_yet:
        #获取邮件列表
        url = "https://domain-open-api.cuiqiu.com/v1/box/list"
        payload={'mail_id': cuiqiu_mail_id,
        'token': cuiqiu_token,
        'start_time': '2022-08-24',
        'end_time': '2023-08-25',
        'page': '1',
        'limit': '100'}
        files=[

        ]
        headers = {}
        response = requests.request("POST", url, headers=headers, data=payload, files=files)
        result = response.text #原始字符串格式
        result_to_json = json.loads(result) #字符串转json
        
        #========这是一个列表集合, 先把列表轮寻一遍, 提取邮件id
        #如果失败, 则再请求下一个列表
        email_list = result_to_json['data']['list'] #取值
        # print("=========所有的邮件在这里: ", email_list)
        for email in email_list:
            if email["to"] == email_to_be_activate:
                if email["from"] == email_from:
                    if email["subject"]== email_subject:
                        print("====找到了这封邮件:", email)
                        result_email_id = email["id"]
                        print("===========待激活的邮件id是:", result_email_id)
                        not_find_yet = False #防止死循环
                        return result_email_id

        try_times += 1
        time_sleep(5,f"尝试{try_times}次, 最多100次. 是不是参数 limit 太少了? ")
        if try_times == 100:
            print("找邮件重试了10分钟,还是失败")
            not_find_yet = False #防止死循环
            result_email_id = False
            return result_email_id

#找到邮件 id 后, 开始提取链接
def cuiqiu_extract_replit_link_from_email_id(email_id):
    # # box_id 通过 v1/box/list 获取邮箱列表接口获取
    url = "https://domain-open-api.cuiqiu.com/v1/box/detail"
    payload={'mail_id': cuiqiu_mail_id,
    'token': cuiqiu_token,
    'box_id': email_id}
    files=[

    ]
    headers = {}

    response = requests.request("POST", url, headers=headers, data=payload, files=files)
    result = response.text
    # print("======request邮件内容初始格式", result)

    to_json = json.loads(result)
    print("转化为json后, 如果有html 则可以xml解析, 很关键!!")

    html_body = to_json['data']['content']['body']
    # print("===提取html_body, 输入到etree", html_body)

    html = etree.HTML(html_body)
    # print("=====解析到的网页是:", type(html), html)
    
    activate_link = html.xpath("//a[text()='Verify Now']/@href")
    print("=====提取到的链接是, 这是一个列表, 需要转为字符串", activate_link, type(activate_link))
    return str(activate_link[0])



#注册 Replit 时,填写随机信息
def signup_replit_random_info(browser, wait, email_account, email_pw):
    first_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Username']")))
    time_sleep(2,"准备输入姓名")
    #fake.first_name()+fake.last_name()
    name = email_account.split("@")[0]
    while len(name) > 15: #网站限制不能超过15个字符
        name = name[:15]
    first_name.send_keys(name)
    time.sleep(2)

    email_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Email']")))
    time_sleep(2,"准备输入邮箱")
    email_button.send_keys(email_account)

    pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@placeholder='Password']")))
    time_sleep(2,"准备输入密码")
    pw_button.send_keys(email_pw)

    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Create account']")))
    time_sleep(2,"准备点击登录")
    browser.execute_script("arguments[0].click();", confirm_login)
    time_sleep(2,"已经点击登录")

    #=========循环查看是不是有resent按钮
    check_resent_flag = True
    active_email_flag = False #初始定义不要去邮箱检查激活链接
    try_times = 0
    while check_resent_flag:
        try:
            try_times +=1  
            print("=========准备不断去查找on board")
            onboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Skip']")))
            check_resent_flag = False #不用再检查resent了
            active_email_flag = True #需要去激活邮件
            # browser.quit()#不要关闭浏览器, 直接打开激活链接, 防止有验证码
            return active_email_flag
        except:
            time_sleep(15, "**********暂时还没有找到resent链接, 等等")
            
        if try_times == 32: #最多等待8分钟
            print("没有找到首页按钮,可能是有验证, 记录到excel")
            check_resent_flag = False #不要再检查resent了
            browser.quit()

#开始打开浏览器激活
def cuiqiu_browser_active_replit_link(wait, browser, activate_link, email_account, email_pw):
    print("开始激活 replit 发来的邮件")

    dashboard_flag = False
    ##============ 准备浏览器, 激活帐号
    print("登录 Replit")
 
    #=======清理下缓存
    # delete_cookie(browser)

    #=============正式开始
    browser.get(activate_link)
    time_sleep(5, "等待  加载")
    switch_tab_by_handle(browser, 1, 0)  # 切换到
    #尝试输入帐号密码登录
    # try:
    #     email_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='1val-input']")))
    #     time_sleep(2,"准备输入邮箱")
    #     email_button.send_keys(email_account)

    #     pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='2val-input']")))
    #     time_sleep(2,"准备输入密码")
    #     pw_button.send_keys(email_pw)

    #     confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Log in']")))
    #     time_sleep(2,"准备点击登录")
    #     browser.execute_script("arguments[0].click();", confirm_login)
    #     time_sleep(10,"cuiqiu_browser_active_replit_link 纯倒计时,等待网页加载, 查看 on boarding 按钮")
    # except:
    #     print("可能不要登录")

    try:
        #如果能找到首页,说明已经进入了
        time_sleep(15, "===尝试查看能不能找到哦啊首页")
        Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//h2[text()='Get started with Replit!']")))
        dashboard_flag = True
        print("=======已经登录了首页")
    except:
        print("======再试试能不能找到 @ ====")
        try:
            #如果能找到@,说明已经进入
            Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()[contains(.,'@')]]")))
            dashboard_flag = True
            print("=======已经找到了@")
        except:
            print("======没有进入首页====")

    return dashboard_flag
## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   Replit项目的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #

## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   polygonscan项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #
#找 polysacn 的注册激活的邮件 id
def cuiqiu_find_polyscan_activate_email_id(email_to_be_activate, email_from, email_subject):
    #循环检索邮箱. email_to_be_activate 表示待激活的邮箱
    not_find_yet = True
    try_times = 0
    while not_find_yet:
        #获取邮件列表
        url = "https://domain-open-api.cuiqiu.com/v1/box/list"
        payload={'mail_id': cuiqiu_mail_id,
        'token': cuiqiu_token,
        'start_time': '2022-08-24',
        'end_time': '2023-08-25',
        'page': '1',
        'limit': '100'}
        files=[

        ]
        headers = {}
        response = requests.request("POST", url, headers=headers, data=payload, files=files)
        result = response.text #原始字符串格式
        result_to_json = json.loads(result) #字符串转json
        
        #========这是一个列表集合, 先把列表轮寻一遍, 提取邮件id
        #如果失败, 则再请求下一个列表
        email_list = result_to_json['data']['list'] #取值
        # print("=========所有的邮件在这里: ", email_list)
        for email in email_list:
            if email["to"] == email_to_be_activate:
                if email["from"] == email_from:
                    if email["subject"]== email_subject:
                        print("====找到了这封邮件:", email)
                        result_email_id = email["id"]
                        print("===========待激活的邮件id是:", result_email_id)
                        not_find_yet = False #防止死循环
                        return result_email_id

        try_times += 1
        time_sleep(5,f"尝试{try_times}次, 最多100次. 是不是参数 limit 太少了? ")
        if try_times == 100:
            print("找邮件重试了10分钟,还是失败")
            not_find_yet = False #防止死循环
            result_email_id = False
            return result_email_id

#找到邮件 id 后, 开始提取链接
def cuiqiu_extract_polyscan_link_from_email_id(email_id):
    # # box_id 通过 v1/box/list 获取邮箱列表接口获取
    url = "https://domain-open-api.cuiqiu.com/v1/box/detail"
    payload={'mail_id': cuiqiu_mail_id,
    'token': cuiqiu_token,
    'box_id': email_id}
    files=[

    ]
    headers = {}

    response = requests.request("POST", url, headers=headers, data=payload, files=files)
    result = response.text
    # print("======request邮件内容初始格式", result)

    to_json = json.loads(result)
    print("转化为json后, 如果有html 则可以xml解析, 很关键!!")
    # print("======转为json后", to_json)  #这个决定了能不能用xml解析, 如果有html 则可以解析

    text_info = to_json['data']['content']['body']  #这是字符串
    print("===========格式为",type(text_info), text_info)

    #由于返回值没有html, 只能字符串提取
    activate_link = text_info.split("Link : ")[1].split("***")[0]
    print("=======切割后的链接是:", activate_link)
    return activate_link


#注册 polysacn 时,填写随机信息
def signup_polygonscan_random_info(browser, wait, email_account, email_pw, signup_name):
    user_name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='ContentPlaceHolder1_txtUserName']")))
    time_sleep(2,"准备输入姓名")
    # name = fake.first_name()+fake.last_name()
    # # name = email_account.split("@")[0]
    # while len(name) > 29: #网站限制不能超过15个字符
    #     name = name[:28]
    user_name.send_keys(signup_name)
    time.sleep(2)

    email_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_txtEmail']")))
    time_sleep(2,"准备输入邮箱")
    email_button.send_keys(email_account)

    pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_txtPassword']")))
    time_sleep(2,"准备输入密码")
    pw_button.send_keys(email_pw)

    pw_button_2 = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_txtPassword2']")))
    time_sleep(2,"准备输入密码")
    pw_button_2.send_keys(email_pw)

    check_box_1 = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='I agree to the ']")))
    time_sleep(2,"准备打勾1")
    browser.execute_script("arguments[0].click();", check_box_1)
    # ActionChains(browser).click(check_box_1).perform()  # 模拟鼠标点

    check_box_2 = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='I agree to receive the PolygonScan newsletter and understand that I can unsubscribe at any time.']")))
    time_sleep(2,"准备打勾2")
    # ActionChains(browser).click(check_box_2).perform()  # 模拟鼠标点
    browser.execute_script("arguments[0].click();", check_box_2)
    
    captcha_info = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='captcha-solver-info']")))
    captcha_info_now = captcha_info.text
    
    try_times = 0
    while captcha_info_now == "Solving...":
        captcha_info = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='captcha-solver-info']")))
        captcha_info_now = captcha_info.text
        time_sleep(5, "**********暂时还没有解决验证码, 等等")
        try_times += 1
        if try_times == 120: #最多等待10分钟
            print("=======还是没有通过验证码,失败!!")
            browser.quit()
    
    #如果解决了验证码, 则点击注册
    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_btnRegister']")))
    time_sleep(2,"准备点击登录")
    browser.execute_script("arguments[0].click();", confirm_login)
    # ActionChains(browser).click(confirm_login).perform()  # 模拟鼠标点
    time_sleep(2,"已经点击登录")


    #点击login后, 查看是否进入主页
    #=========循环查看是不是有resent按钮
    check_resent_flag = True
    already_in_flag = False #初始定义不要flase
    try_times = 0
    while check_resent_flag:
        try:
            try_times +=1  
            print("=========准备不断去查找标志!!")
            onboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()[contains(.,'We have sent')]]")))
            check_resent_flag = False #不用再检查resent了
            already_in_flag = True #需要去激活邮件
            # browser.quit()#不要关闭浏览器, 直接打开激活链接, 防止有验证码
            time_sleep(3, f"++++++++++已经找到了sent 标志, 等待时间")
            return already_in_flag #需要记录用户名
        except:
            time_sleep(5, "**********暂时还没有找到主页, 等等")
            
        if try_times == 60: #最多等待5分钟
            print("没有找到首页按钮,可能是有验证, 记录到excel")
            check_resent_flag = False #不要再检查resent了
            browser.quit()
            return already_in_flag

#开始打开浏览器激活
def cuiqiu_browser_active_polygonscan_link(wait, browser, activate_link):
    dashboard_flag = False
    ##============ 准备浏览器, 激活帐号
    print("登录 polygonscan")

    #=============正式开始
    browser.get(activate_link)
    time_sleep(5, "等待  加载")
    switch_tab_by_handle(browser, 1, 0)  # 切换到
    try:
        #如果能找到首页,说明已经进入了
        time_sleep(15, "尝试查看是不是已经进入首页")
        Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//h1[text()[contains(.,'Confirm')]]")))
        dashboard_flag = True
        print("=======已经登录了首页")
    except:
        print("======没有进入首页====")
        
    return dashboard_flag

# polyscan登录
def login_polygonscan(wait, browser, email_pw, user_name):
    click_login_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@value='Click to Login']")))
    time_sleep(2,"准备点击click to login")
    browser.execute_script("arguments[0].click();", click_login_button)

    user_name_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_txtUserName']")))
    time_sleep(2,"准备输入邮箱")
    user_name_button.send_keys(user_name)

    pw_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_txtPassword']")))
    time_sleep(2,"准备输入密码")
    pw_button.send_keys(email_pw)

    # check_box_1 = wait.until(EC.element_to_be_clickable((By.XPATH,"//label[@class='custom-control-label']")))
    # time_sleep(2,"准备打勾1")
    # browser.execute_script("arguments[0].click();", check_box_1)
    # ActionChains(browser).click(check_box_1).perform()  # 模拟鼠标点

    try_times = 0
    captcha_info_now = "Solving..."#设置初始值为Solving...
    print("======开始等待验证码")
    while captcha_info_now == "Solving...":
        captcha_info = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='captcha-solver-info']")))
        captcha_info_now = captcha_info.text
        time_sleep(5, "**********暂时还没有解决验证码, 等等")
        try_times += 1
        if try_times == 120: #最多等待10分钟
            print("=======还是没有通过验证码,失败!!")
            browser.quit()
    print("======验证码通过!!")
    #如果解决了验证码, 则点击注册
    confirm_login = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_btnLogin']")))
    time_sleep(2,"准备点击登录")
    browser.execute_script("arguments[0].click();", confirm_login)

    #=========循环查看是不是有sent按钮
    check_home_flag = True
    already_home_flag = False #初始定义不要去邮箱检查激活链接
    try_times = 0
    while check_home_flag:
        try:
            try_times +=1  
            print("=========准备不断去查找on board")
            onboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[text()[contains(.,'API-KEYs')]]")))
            check_home_flag = False #不用再检查resent了
            already_home_flag = True #
            # browser.quit()#不要关闭浏览器, 直接打开激活链接, 防止有验证码
            return already_home_flag
        except:
            time_sleep(5, f"**********暂时还没有找到主页, 等等{try_times}/50")
        if try_times == 50:
            print("========登录失败!!")
            check_home_flag = False
            browser.quit()
            
#polyscan获取API
def pologanscan_get_API(wait, browser):
    api_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[text()[contains(.,'API-KEYs')]]")))
    time_sleep(2,"准备点击 api_button ")
    browser.execute_script("arguments[0].click();", api_button)

    add_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@id='ContentPlaceHolder1_addnew']")))
    time_sleep(2,"准备点击 Add")
    browser.execute_script("arguments[0].click();", add_button)

    continue_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ContentPlaceHolder1_btnSubmit']")))
    time_sleep(2,"准备点击continue")
    browser.execute_script("arguments[0].click();", continue_button)
    
    api_box = wait.until(EC.element_to_be_clickable((By.XPATH,"//table/tbody[1]//td[2]")))
    api_result = api_box.text
    print("提取到的API信息是", api_result)

    return api_result

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   polygonscan项目的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   polygonscan项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #



#开始打开浏览器激活
def cuiqiu_browser_active_polyscan_link(browser, wait, activate_link):
    print("开始激活alchemy发来的邮件")
    dashboard_flag = False
    
    browser.get(activate_link)
    time_sleep(5, "等待 alchemy 加载")
    switch_tab_by_handle(browser, 1, 0)  # 切换到

    #=================激活帐号
    try:
        Verify_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='Verify']")))
        time_sleep(2,"准备点击 Verify")
        browser.execute_script("arguments[0].click();", Verify_button)
        time_sleep(30,"=====已经点击 Verify")
    except:
        print("点击verify失败, 会不会有影响?")
 
    # ==================填写描述
    try:
        fill_in_alchemy_project_des(browser, wait)
    except:
        print("可能是不需要填写alchemy项目描述, 或哪里出错了")

    try:
        #如果能找到Alchemy的首页,说明已经进入了
        time_sleep(20, "尝试查看是不是已经进入 Alchemy 了")
        Dashboard_button = wait.until(EC.element_to_be_clickable((By.XPATH,"//div[text()[contains(.,'Dashboard')]]")))
        dashboard_flag = True
        print("=======已经登录了 Alchemy")
        return dashboard_flag
    except:
        print("======没有找到dashboard=====")
    

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #


## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #



## ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓   项目的一些函数 ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ ↓ #

## ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑   的一些函数 ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ ↑ #